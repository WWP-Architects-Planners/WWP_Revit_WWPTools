#! python3

import os
import re
import sys
import System
import traceback

from pyrevit import DB, revit, script


output = None
_material_cache = None

OVERWRITE_MODE_ASK = "ask"
OVERWRITE_MODE_ALL = "overwrite_all"
OVERWRITE_MODE_SKIP = "skip_all"
MISSING_MATERIALS_ABORT = "abort"
MISSING_MATERIALS_IGNORE = "ignore"


def add_lib_path():
    script_dir = os.path.dirname(__file__)
    lib_path = os.path.abspath(os.path.join(script_dir, "..", "..", "..", "..", "lib"))
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def load_uiutils():
    add_lib_path()
    import WWP_uiUtils as ui
    return ui


def init_output():
    global output
    try:
        output = script.get_output()
    except Exception:
        output = None


def log_info(message):
    print("[INFO] {}".format(message))


def log_warn(message):
    print("[WARN] {}".format(message))


def log_error(message):
    print("[ERROR] {}".format(message))


def log_section(title):
    print("")
    print("=== {} ===".format(title))


def choose_overwrite_mode(ui):
    options = [
        "Ask Each Time",
        "Overwrite All Changed Types",
        "Skip All Changed Types",
    ]
    selected = ui.uiUtils_select_indices(
        options,
        title="Import Type Layers",
        prompt="Overwrite behavior for changed existing types:",
        multiselect=False,
        width=520,
        height=360,
    )
    if not selected:
        return None
    idx = selected[0]
    if idx == 1:
        return OVERWRITE_MODE_ALL
    if idx == 2:
        return OVERWRITE_MODE_SKIP
    return OVERWRITE_MODE_ASK


def choose_missing_material_mode(ui, missing_materials):
    material_names = sorted(missing_materials.keys())
    affected_types = set()
    preview = []
    for name in material_names[:12]:
        type_names = sorted(missing_materials.get(name) or [])
        affected_types.update(type_names)
        if type_names:
            preview.append("{} [{}]".format(name, ", ".join(type_names[:2])))
        else:
            preview.append(name)
    for names in missing_materials.values():
        affected_types.update(names or [])

    prompt_lines = [
        "Missing materials were found in the workbook.",
        "",
        "Missing materials: {}".format(len(material_names)),
        "Affected types: {}".format(len(affected_types)),
    ]
    if preview:
        prompt_lines.append("")
        prompt_lines.append("Examples:")
        prompt_lines.extend(preview)
        if len(material_names) > len(preview):
            prompt_lines.append("... and {} more".format(len(material_names) - len(preview)))
    prompt_lines.append("")
    prompt_lines.append("Choose how to proceed:")

    options = [
        "Do Not Import, Materials First",
        "Ignore Missing Materials",
    ]
    selected = ui.uiUtils_select_indices(
        options,
        title="Import Type Layers",
        prompt="\n".join(prompt_lines),
        multiselect=False,
        width=760,
        height=520,
    )
    if not selected:
        return None
    if selected[0] == 0:
        return MISSING_MATERIALS_ABORT
    return MISSING_MATERIALS_IGNORE


def element_id_value(elem_id):
    if elem_id is None:
        return -1
    if hasattr(elem_id, "IntegerValue"):
        return elem_id.IntegerValue
    if hasattr(elem_id, "Value"):
        return elem_id.Value
    try:
        return int(elem_id)
    except Exception:
        return -1


def pick_excel_path(doc, ui):
    initial_dir = ""
    try:
        if doc.PathName:
            initial_dir = os.path.dirname(doc.PathName)
    except Exception:
        initial_dir = ""
    if not initial_dir or not os.path.isdir(initial_dir):
        initial_dir = ""
    return ui.uiUtils_open_file_dialog(
        title="Select Type Layers Excel Workbook",
        filter_text="Excel Files (*.xlsx)|*.xlsx|All Files (*.*)|*.*",
        multiselect=False,
        initial_directory=initial_dir,
    )


def get_length_unit_id(doc):
    try:
        units = doc.GetUnits()
        return units.GetFormatOptions(DB.SpecTypeId.Length).GetUnitTypeId()
    except Exception:
        return None


def get_millimeter_unit_id():
    try:
        return DB.UnitTypeId.Millimeters
    except Exception:
        return None


def to_internal_length(doc, value):
    unit_id = get_length_unit_id(doc)
    if unit_id:
        try:
            return DB.UnitUtils.ConvertToInternalUnits(value, unit_id)
        except Exception:
            pass
    metric_unit_id = get_millimeter_unit_id()
    if metric_unit_id:
        try:
            return DB.UnitUtils.ConvertToInternalUnits(value, metric_unit_id)
        except Exception:
            pass
    try:
        return DB.UnitUtils.ConvertToInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def to_internal_length_mm(value):
    metric_unit_id = get_millimeter_unit_id()
    if metric_unit_id:
        try:
            return DB.UnitUtils.ConvertToInternalUnits(value, metric_unit_id)
        except Exception:
            pass
    try:
        return DB.UnitUtils.ConvertToInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def parse_display_length(doc, value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        units = doc.GetUnits()
        return DB.UnitFormatUtils.Parse(units, DB.SpecTypeId.Length, text)
    except Exception:
        pass
    cleaned = text.replace(",", "")
    match = re.search(r"-?\d+(\.\d+)?", cleaned)
    if not match:
        return None
    try:
        return to_internal_length(doc, float(match.group(0)))
    except Exception:
        return None


def parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in ("true", "yes", "1")


def parse_function(value):
    if value is None or value == "":
        return DB.MaterialFunctionAssignment.Structure
    if isinstance(value, (int, float)):
        try:
            return DB.MaterialFunctionAssignment(int(value))
        except Exception:
            pass
    try:
        int_val = int(str(value).strip())
        return DB.MaterialFunctionAssignment(int_val)
    except Exception:
        pass
    name = str(value).strip()
    if not name:
        return DB.MaterialFunctionAssignment.Structure
    bracket_match = re.search(r"\[(\d+)\]", name)
    if bracket_match:
        try:
            return DB.MaterialFunctionAssignment(int(bracket_match.group(1)))
        except Exception:
            pass
    if "." in name:
        name = name.split(".")[-1]
    normalized = re.sub(r"[^A-Za-z0-9]+", "", name).lower()
    aliases = {
        "none": "None",
        "structure": "Structure",
        "substrate": "Substrate",
        "thermalairlayer": "Insulation",
        "insulation": "Insulation",
        "finish1": "Finish1",
        "finish2": "Finish2",
        "membranelayer": "Membrane",
        "membrane": "Membrane",
        "structuraldeck": "StructuralDeck",
    }
    enum_name = aliases.get(normalized)
    if enum_name:
        try:
            return getattr(DB.MaterialFunctionAssignment, enum_name)
        except Exception:
            pass
    try:
        return getattr(DB.MaterialFunctionAssignment, name)
    except Exception:
        return DB.MaterialFunctionAssignment.Structure


def describe_layer_row(row):
    return "index={}; function={}; material={}; thickness={}".format(
        row.get("LayerIndex"),
        row.get("LayerFunctionId") or row.get("LayerFunction"),
        row.get("MaterialName"),
        first_non_none(
            row.get("ThicknessMetric"),
            row.get("ThicknessProject"),
            row.get("ThicknessDisplay"),
            row.get("ThicknessInternal"),
        ),
    )


def describe_group(group):
    return "type='{}'; source='{}'; type_id={}; rows={}".format(
        group.get("TypeName"),
        group.get("SourceTypeName"),
        group.get("TypeId"),
        len(group.get("Layers", [])),
    )


def first_non_none(*values):
    for value in values:
        if value is not None and value != "":
            return value
    return None


def normalize_lookup_text(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def tokenize_lookup_text(value):
    if value is None:
        return []
    return [token for token in re.split(r"[^A-Za-z0-9]+", str(value).strip().lower()) if token]


def get_material_cache(doc):
    global _material_cache
    if _material_cache is not None:
        return _material_cache

    exact = {}
    lower = {}
    normalized = {}
    materials = []
    for mat in DB.FilteredElementCollector(doc).OfClass(DB.Material):
        try:
            name = mat.Name
        except Exception:
            continue
        if not name:
            continue
        norm = normalize_lookup_text(name)
        exact[name] = mat
        lower[name.lower()] = mat
        normalized.setdefault(norm, []).append(mat)
        materials.append((mat, name, norm, set(tokenize_lookup_text(name))))

    _material_cache = {
        "exact": exact,
        "lower": lower,
        "normalized": normalized,
        "materials": materials,
    }
    return _material_cache


def material_name_aliases(mat_name):
    text = str(mat_name or "").strip()
    if not text:
        return []

    aliases = [text]
    parts = [part.strip() for part in text.split("_") if part and part.strip()]
    tail_candidates = []
    if parts:
        tail_candidates.append(parts[-1])
        if len(parts) >= 2:
            tail_candidates.append(" ".join(parts[-2:]))
        if len(parts) >= 3:
            tail_candidates.append(" ".join(parts[-3:]))
    for alias in tail_candidates:
        if alias and alias not in aliases:
            aliases.append(alias)
    return aliases


def find_material_id(doc, mat_id_value, mat_name):
    if mat_id_value and mat_id_value > 0:
        try:
            mat = doc.GetElement(DB.ElementId(mat_id_value))
            if mat:
                return mat.Id, ""
        except Exception:
            pass
    if not mat_name:
        return DB.ElementId.InvalidElementId, ""

    cache = get_material_cache(doc)
    name = str(mat_name).strip()
    exact = cache["exact"].get(name)
    if exact:
        return exact.Id, ""

    lower = cache["lower"].get(name.lower())
    if lower:
        return lower.Id, "Material matched case-insensitively: '{}' -> '{}'.".format(mat_name, lower.Name)

    aliases = material_name_aliases(name)
    for alias in aliases:
        norm = normalize_lookup_text(alias)
        mats = cache["normalized"].get(norm) or []
        if len(mats) == 1:
            mat = mats[0]
            if mat.Name == name:
                return mat.Id, ""
            return mat.Id, "Material matched by normalized name: '{}' -> '{}'.".format(mat_name, mat.Name)

    contains_candidates = []
    for alias in aliases:
        alias_norm = normalize_lookup_text(alias)
        if not alias_norm:
            continue
        for mat, project_name, project_norm, _ in cache["materials"]:
            if alias_norm == project_norm:
                continue
            if project_norm and (project_norm in alias_norm or alias_norm in project_norm):
                score = abs(len(project_norm) - len(alias_norm))
                contains_candidates.append((score, project_name, mat))
    if contains_candidates:
        contains_candidates.sort(key=lambda item: (item[0], item[1]))
        best = contains_candidates[0][2]
        return best.Id, "Material matched by partial normalized name: '{}' -> '{}'.".format(mat_name, best.Name)

    token_candidates = []
    source_tokens = set(tokenize_lookup_text(aliases[0] if aliases else name))
    source_tokens.update(tokenize_lookup_text(aliases[1] if len(aliases) > 1 else ""))
    if source_tokens:
        for mat, project_name, _, project_tokens in cache["materials"]:
            shared = len(source_tokens.intersection(project_tokens))
            if shared >= 2:
                token_candidates.append((-shared, len(project_tokens), project_name, mat))
    if token_candidates:
        token_candidates.sort(key=lambda item: (item[0], item[1], item[2]))
        best = token_candidates[0][3]
        return best.Id, "Material matched by token overlap: '{}' -> '{}'.".format(mat_name, best.Name)

    return DB.ElementId.InvalidElementId, ""


def is_invalid_element_id(elem_id):
    if elem_id is None:
        return True
    try:
        if elem_id == DB.ElementId.InvalidElementId:
            return True
    except Exception:
        pass
    return element_id_value(elem_id) == -1


def resolve_duplicated_type(doc, duplicated):
    if duplicated is None:
        return None
    try:
        if isinstance(duplicated, DB.ElementId):
            return doc.GetElement(duplicated)
    except Exception:
        pass
    if hasattr(duplicated, "Id") and hasattr(duplicated, "Name"):
        return duplicated
    try:
        return doc.GetElement(duplicated)
    except Exception:
        return None


def read_workbook(path, ui):
    add_lib_path()
    try:
        import openpyxl
    except Exception as exc:
        ui.uiUtils_alert("openpyxl is not available.\n{}".format(exc), title="Import Type Layers")
        return None
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        ui.uiUtils_alert("Failed to open workbook.\n{}".format(exc), title="Import Type Layers")
        return None


def extract_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(val).strip() if val is not None else "" for val in rows[0]]
    aliases = {
        "category": "Category",
        "id": "IdeateId",
        "typename": "TypeName",
        "type name": "TypeName",
        "type: type name": "TypeName",
        "family and type": "TypeName",
        "type: family and type": "TypeName",
        "type: type mark": "TypeMark",
        "sourcetypename": "SourceTypeName",
        "source type name": "SourceTypeName",
        "originaltypename": "SourceTypeName",
        "original type name": "SourceTypeName",
        "familyname": "FamilyName",
        "family name": "FamilyName",
        "typeid": "TypeId",
        "type id": "TypeId",
        "typeuniqueid": "TypeUniqueId",
        "type unique id": "TypeUniqueId",
        "layerindex": "LayerIndex",
        "layer index": "LayerIndex",
        "index": "LayerIndex",
        "layerfunction": "LayerFunction",
        "layer function": "LayerFunction",
        "layerfunctionid": "LayerFunctionId",
        "layer function id": "LayerFunctionId",
        "function": "LayerFunction",
        "materialname": "MaterialName",
        "material name": "MaterialName",
        "materialid": "MaterialId",
        "material id": "MaterialId",
        "material: name": "MaterialName",
        "thicknessinternal": "ThicknessInternal",
        "thickness internal": "ThicknessInternal",
        "thicknessmetric": "ThicknessMetric",
        "thickness metric": "ThicknessMetric",
        "thickness": "ThicknessMetric",
        "thicknessdisplay": "ThicknessDisplay",
        "thickness display": "ThicknessDisplay",
        "thicknessproject": "ThicknessProject",
        "thickness project": "ThicknessProject",
        "isvariable": "IsVariable",
        "is variable": "IsVariable",
        "variable": "IsVariable",
        "hascompoundstructure": "HasCompoundStructure",
        "has compound structure": "HasCompoundStructure",
    }
    header_map = {}
    for idx, name in enumerate(headers):
        if not name:
            continue
        key = name.strip()
        lower = key.lower()
        canonical = aliases.get(lower, key)
        header_map[canonical] = idx
    data = []
    for raw in rows[1:]:
        if raw is None:
            continue
        if all(cell in (None, "") for cell in raw):
            continue
        row = {}
        for key, idx in header_map.items():
            if idx < len(raw):
                row[key] = raw[idx]
            else:
                row[key] = None
        type_name = first_non_none(row.get("TypeName"), row.get("SourceTypeName"), row.get("TypeMark"))
        has_layer_data = first_non_none(
            row.get("LayerIndex"),
            row.get("LayerFunction"),
            row.get("LayerFunctionId"),
            row.get("ThicknessMetric"),
            row.get("ThicknessDisplay"),
            row.get("ThicknessProject"),
            row.get("ThicknessInternal"),
        )
        raw_id = row.get("IdeateId")
        parsed = parse_ideate_row_id(raw_id)
        row["RowGroupKey"] = parsed["group_key"]
        row["RowIsNewPlaceholder"] = parsed["is_new_placeholder"]
        row["RowReferencesExistingId"] = parsed["references_existing_id"]
        if parsed["type_id"] is not None and row.get("TypeId") in (None, ""):
            row["TypeId"] = parsed["type_id"]
        if parsed["layer_index"] is not None and row.get("LayerIndex") in (None, ""):
            row["LayerIndex"] = parsed["layer_index"]
        if not type_name and not has_layer_data and raw_id in (None, ""):
            continue
        data.append(row)
    return data


def parse_ideate_row_id(value):
    text = "" if value is None else str(value).strip()
    result = {
        "raw": text,
        "group_key": text,
        "type_id": None,
        "layer_index": None,
        "is_new_placeholder": False,
        "references_existing_id": False,
    }
    if not text:
        result["group_key"] = None
        result["is_new_placeholder"] = True
        return result

    if text.lower() == "new":
        result["group_key"] = None
        result["is_new_placeholder"] = True
        return result

    match = re.match(r"^\[(\d+)\]\[(\d+)\]$", text)
    if match:
        result["type_id"] = int(match.group(1))
        result["layer_index"] = int(match.group(2))
        result["group_key"] = match.group(1)
        result["references_existing_id"] = True
        return result

    match = re.match(r"^\[(\d+)\]$", text)
    if match:
        result["type_id"] = int(match.group(1))
        result["group_key"] = match.group(1)
        result["references_existing_id"] = True
        return result

    match = re.match(r"^(new\[\d+\])$", text, re.IGNORECASE)
    if match:
        layer_match = re.match(r"^new\[(\d+)\]$", text, re.IGNORECASE)
        if layer_match:
            result["layer_index"] = int(layer_match.group(1))
        result["group_key"] = None
        result["is_new_placeholder"] = True
        return result

    return result


def map_category(sheet_name):
    name = sheet_name.strip().lower()
    if name in ("wall types", "walls", "wall", "walls_layers", "walls layers", "wall_layers", "wall layers"):
        return ("Wall Types", DB.WallType)
    if name in ("ceiling types", "ceilings", "ceiling", "ceilings_layers", "ceilings layers", "ceiling_layers", "ceiling layers"):
        return ("Ceiling Types", DB.CeilingType)
    if name in ("floor types", "floors", "floor", "floors_layers", "floors layers", "floor_layers", "floor layers"):
        return ("Floor Types", DB.FloorType)
    if name in ("roof types", "roofs", "roof", "roofs_layers", "roofs layers", "roof_layers", "roof layers"):
        return ("Roof Types", DB.RoofType)
    return None


def collect_types_by_category(doc, type_class):
    items = list(DB.FilteredElementCollector(doc).OfClass(type_class).WhereElementIsElementType())
    by_unique = {}
    by_id = {}
    by_name = {}
    for item in items:
        try:
            by_unique[item.UniqueId] = item
        except Exception:
            pass
        by_id[element_id_value(item.Id)] = item
        by_name[item.Name] = item
    return items, by_unique, by_id, by_name


def build_layer(layer_row, doc):
    width = None
    metric = parse_float(layer_row.get("ThicknessMetric"))
    if metric is not None:
        width = to_internal_length_mm(metric)
    if width is None:
        width = parse_float(layer_row.get("ThicknessInternal"))
    if width is None:
        width = parse_float(layer_row.get("ThicknessProject"))
        if width is not None:
            width = to_internal_length(doc, width)
    if width is None:
        width = parse_display_length(doc, layer_row.get("ThicknessDisplay"))
    func_value = layer_row.get("LayerFunctionId")
    if func_value in (None, ""):
        func_value = layer_row.get("LayerFunction")
    func = parse_function(func_value)

    if width is None:
        return None, "No thickness value found."
    if width < 0:
        return None, "Thickness must be zero or greater."
    if width == 0 and func != DB.MaterialFunctionAssignment.Membrane:
        return None, "Zero thickness is only valid for membrane layers."

    mat_id_val = parse_float(layer_row.get("MaterialId"))
    mat_id_val = int(mat_id_val) if mat_id_val is not None else None
    mat_name = layer_row.get("MaterialName")
    mat_id, material_lookup_message = find_material_id(doc, mat_id_val, mat_name)

    layer = DB.CompoundStructureLayer(width, func, mat_id)
    if "IsVariable" in layer_row:
        try:
            layer.IsVariable = parse_bool(layer_row.get("IsVariable"))
        except Exception:
            pass
    material_message = material_lookup_message
    if mat_name and is_invalid_element_id(mat_id) and str(mat_name).strip() != "<By Category>":
        material_message = "{} Material not found in project: '{}'.".format(material_message, mat_name).strip()
    return layer, material_message


def update_type_layers(doc, elem_type, layers):
    if not layers:
        return False, "No valid layers found."

    net_layers = System.Collections.Generic.List[DB.CompoundStructureLayer]()
    for layer in layers:
        net_layers.Add(layer)

    compound = None
    set_layers_error = None
    create_error = None
    try:
        compound = elem_type.GetCompoundStructure()
    except Exception:
        compound = None

    if compound is not None:
        try:
            compound.SetLayers(net_layers)
        except Exception as exc:
            set_layers_error = str(exc)
            compound = None

    if compound is None:
        try:
            compound = DB.CompoundStructure.CreateSimpleCompoundStructure(net_layers)
        except Exception as exc:
            create_error = str(exc)
            compound = None

    if compound is None:
        errors = []
        if set_layers_error:
            errors.append("SetLayers failed: {}".format(set_layers_error))
        if create_error:
            errors.append("CreateSimpleCompoundStructure failed: {}".format(create_error))
        if errors:
            return False, "Compound structure creation failed. {}".format(" | ".join(errors))
        return False, "Compound structure creation failed."

    try:
        elem_type.SetCompoundStructure(compound)
    except Exception as exc:
        return False, str(exc)
    return True, ""


def rename_type(elem_type, new_name):
    if not new_name:
        return
    try:
        if elem_type.Name != new_name:
            elem_type.Name = new_name
    except Exception:
        pass


def get_existing_layers(elem_type):
    try:
        compound = elem_type.GetCompoundStructure()
    except Exception:
        compound = None
    if compound is None:
        return []
    try:
        return list(compound.GetLayers())
    except Exception:
        return []


def layers_thickness_changed(existing_layers, new_layers, tol=1e-4):
    if len(existing_layers) != len(new_layers):
        return True
    for old, new in zip(existing_layers, new_layers):
        try:
            if abs(old.Width - new.Width) > tol:
                return True
        except Exception:
            return True
    return False


def has_compound_structure_flag(row):
    if "HasCompoundStructure" not in row:
        return True
    return parse_bool(row.get("HasCompoundStructure"))


def pick_seed_type(types, family_name=None):
    if not types:
        return None

    if family_name:
        family_matches = [t for t in types if (getattr(t, "FamilyName", "") or "") == family_name]
    else:
        family_matches = []

    candidates = family_matches or list(types)
    for elem_type in candidates:
        try:
            if elem_type.GetCompoundStructure() is not None:
                return elem_type
        except Exception:
            continue
    return candidates[0]


def group_rows(rows):
    grouped = {}
    for row in rows:
        row_group_key = row.get("RowGroupKey")
        is_new_placeholder = parse_bool(row.get("RowIsNewPlaceholder"))
        references_existing_id = parse_bool(row.get("RowReferencesExistingId"))
        unique_id = row.get("TypeUniqueId")
        type_id = row.get("TypeId")
        type_name = row.get("TypeName")
        type_mark = row.get("TypeMark")
        source_name = row.get("SourceTypeName")
        if is_new_placeholder:
            key = ("new", type_name or source_name or type_mark)
        else:
            key = row_group_key or unique_id or type_id or source_name or type_name
        if not key:
            continue
        group = grouped.get(key)
        if group is None:
            group = {
                "RowGroupKey": row_group_key,
                "RowIsNewPlaceholder": is_new_placeholder,
                "RowReferencesExistingId": references_existing_id,
                "TypeUniqueId": unique_id,
                "TypeId": type_id,
                "TypeName": type_name,
                "TypeMark": type_mark,
                "SourceTypeName": source_name,
                "FamilyName": row.get("FamilyName"),
                "HasCompoundStructure": has_compound_structure_flag(row),
                "Layers": [],
            }
            grouped[key] = group
        else:
            if references_existing_id:
                group["RowReferencesExistingId"] = True
            if is_new_placeholder:
                group["RowIsNewPlaceholder"] = True
        group["Layers"].append(row)
    return grouped


def collect_missing_materials(doc, category_data):
    missing = {}
    for _, _, rows in category_data:
        for row in rows:
            mat_name = row.get("MaterialName")
            if mat_name is None:
                continue
            mat_name = str(mat_name).strip()
            if not mat_name or mat_name == "<By Category>":
                continue
            mat_id_val = parse_float(row.get("MaterialId"))
            mat_id_val = int(mat_id_val) if mat_id_val is not None else None
            mat_id, _ = find_material_id(doc, mat_id_val, mat_name)
            if is_invalid_element_id(mat_id):
                type_name = first_non_none(row.get("TypeName"), row.get("SourceTypeName"), row.get("TypeMark"), "<Unknown Type>")
                missing.setdefault(mat_name, set()).add(str(type_name))
    return missing


def main():
    doc = revit.doc
    init_output()
    ui = load_uiutils()
    log_section("Start")
    path = pick_excel_path(doc, ui)
    if not path:
        log_warn("No workbook selected. Command cancelled.")
        return
    log_info("Workbook selected: {}".format(path))

    workbook = read_workbook(path, ui)
    if workbook is None:
        log_error("Workbook could not be loaded.")
        return
    log_info("Workbook sheets: {}".format(", ".join(workbook.sheetnames)))

    category_data = []
    for sheet_name in workbook.sheetnames:
        mapping = map_category(sheet_name)
        if not mapping:
            log_info("Skipping sheet '{}': no supported category mapping.".format(sheet_name))
            continue
        label, type_class = mapping
        rows = extract_rows(workbook[sheet_name])
        if rows:
            log_info("Loaded sheet '{}': {} row(s) mapped to {}.".format(sheet_name, len(rows), label))
            category_data.append((label, type_class, rows))
        else:
            log_warn("Sheet '{}' mapped to {} but contained no usable rows.".format(sheet_name, label))

    if not category_data:
        log_error("No matching category sheets found.")
        ui.uiUtils_alert("No matching category sheets found.", title="Import Type Layers")
        return

    missing_materials = collect_missing_materials(doc, category_data)
    if missing_materials:
        log_warn(
            "Missing materials detected before import: {} material(s) across {} type(s).".format(
                len(missing_materials),
                len(set(name for names in missing_materials.values() for name in names)),
            )
        )
        missing_material_mode = choose_missing_material_mode(ui, missing_materials)
        if not missing_material_mode:
            log_warn("No missing-material option selected. Command cancelled.")
            return
        if missing_material_mode == MISSING_MATERIALS_ABORT:
            summary = "Please import the list of materials first.\n\nMissing materials:\n"
            material_names = sorted(missing_materials.keys())
            summary += "\n".join(material_names[:20])
            if len(material_names) > 20:
                summary += "\n... ({} more)".format(len(material_names) - 20)
            log_section("Summary")
            for line in summary.splitlines():
                if line.strip():
                    log_info(line)
            ui.uiUtils_alert(summary, title="Import Type Layers")
            return
        log_warn("User chose to ignore missing materials. Unresolved layers will use <By Category>.")

    overwrite_mode = choose_overwrite_mode(ui)
    if not overwrite_mode:
        log_warn("No overwrite mode selected. Command cancelled.")
        return
    log_info("Overwrite mode: {}.".format(overwrite_mode))

    updated = 0
    created = 0
    skipped = 0
    errors = []
    missing_types_summary = []
    overwrite_skipped = 0

    with revit.Transaction("Import Type Layers"):
        for label, type_class, rows in category_data:
            log_section(label)
            types, by_unique, by_id, by_name = collect_types_by_category(doc, type_class)
            grouped = group_rows(rows)
            excel_ids = set()
            log_info("Existing Revit types: {}. Excel groups: {}.".format(len(types), len(grouped)))

            for key, group in grouped.items():
                unique_id = group.get("TypeUniqueId")
                type_id = group.get("TypeId")
                type_name = group.get("TypeName")
                source_name = group.get("SourceTypeName")
                is_new_placeholder = parse_bool(group.get("RowIsNewPlaceholder"))
                references_existing_id = parse_bool(group.get("RowReferencesExistingId"))
                elem_type = None
                match_reason = None
                was_created = False
                force_create_new = is_new_placeholder

                log_info("Processing group {}.".format(describe_group(group)))

                if not force_create_new and unique_id and unique_id in by_unique:
                    elem_type = by_unique[unique_id]
                    match_reason = "unique id"
                else:
                    type_id_val = parse_float(type_id)
                    if not force_create_new and type_id_val is not None and int(type_id_val) in by_id:
                        elem_type = by_id[int(type_id_val)]
                        match_reason = "type id"
                    elif references_existing_id and type_id_val is not None and int(type_id_val) not in by_id:
                        force_create_new = True
                        log_info(
                            "Source id '{}' was not found in the project. Treating '{}' as a new type.".format(
                                int(type_id_val),
                                type_name or source_name or key,
                            )
                        )
                if not force_create_new and elem_type is None and source_name and source_name in by_name:
                    elem_type = by_name[source_name]
                    match_reason = "source type name"
                if not force_create_new and elem_type is None and type_name and type_name in by_name:
                    elem_type = by_name[type_name]
                    match_reason = "type name"

                if elem_type is None:
                    base_type = pick_seed_type(types, group.get("FamilyName"))
                    if base_type and type_name:
                        try:
                            log_info("No existing type matched. Creating '{}' from seed '{}'.".format(type_name, base_type.Name))
                            duplicated = base_type.Duplicate(type_name)
                            elem_type = resolve_duplicated_type(doc, duplicated)
                            if elem_type is None:
                                raise Exception("Duplicate returned an unsupported object: {}".format(type(duplicated)))
                            created += 1
                            was_created = True
                            match_reason = "created from seed"
                        except Exception as exc:
                            log_error("Create failed for '{}': {}.".format(type_name, exc))
                            errors.append("{}: create failed ({})".format(type_name, exc))
                            skipped += 1
                            continue
                    else:
                        log_warn("Skipped group '{}': no matching type and no seed type available.".format(type_name or source_name or key))
                        skipped += 1
                        continue
                else:
                    log_info("Matched Revit type '{}' by {}.".format(elem_type.Name, match_reason))

                excel_ids.add(element_id_value(elem_type.Id))
                existing_name = elem_type.Name

                sorted_rows = sorted(
                    group["Layers"],
                    key=lambda r: parse_float(r.get("LayerIndex")) or 0,
                )
                layers = []
                for row in sorted_rows:
                    layer, layer_message = build_layer(row, doc)
                    if layer:
                        layers.append(layer)
                        log_info("Accepted layer for '{}': {}.".format(type_name or existing_name, describe_layer_row(row)))
                        if layer_message:
                            log_warn("Layer note for '{}': {}".format(type_name or existing_name, layer_message.strip()))
                    else:
                        log_warn("Ignored layer for '{}': {} ({})".format(type_name or existing_name, describe_layer_row(row), layer_message))

                name_changed = bool(type_name and existing_name != type_name)

                if not layers:
                    if not group.get("HasCompoundStructure", True):
                        log_warn("Skipped '{}': no layers provided and group is marked without compound structure.".format(type_name or source_name or existing_name))
                        skipped += 1
                        continue
                    if name_changed:
                        prompt = (
                            "Type '{}' has a name change ({} -> {}).\n"
                            "No valid layer data was found.\n\n"
                            "Rename anyway?"
                        ).format(existing_name, existing_name, type_name)
                        allow_rename = False
                        if overwrite_mode == OVERWRITE_MODE_ALL:
                            allow_rename = True
                        elif overwrite_mode == OVERWRITE_MODE_SKIP:
                            allow_rename = False
                        else:
                            allow_rename = ui.uiUtils_confirm(prompt, title="Import Type Layers")
                        if allow_rename:
                            log_info("User accepted rename-only change: '{}' -> '{}'.".format(existing_name, type_name))
                            rename_type(elem_type, type_name)
                        else:
                            log_warn("User declined rename-only change for '{}'.".format(existing_name))
                            overwrite_skipped += 1
                            skipped += 1
                            continue
                    skipped += 1
                    log_warn("Skipped '{}': no valid layers found after parsing.".format(type_name or source_name or existing_name))
                    errors.append("{}: No valid layers found.".format(type_name or source_name or existing_name))
                    continue

                existing_layers = get_existing_layers(elem_type)
                thickness_changed = layers_thickness_changed(existing_layers, layers)
                if was_created:
                    log_info("Applying imported layers to newly created type '{}' without overwrite prompt.".format(elem_type.Name))
                elif name_changed or thickness_changed:
                    prompt_parts = []
                    if name_changed:
                        prompt_parts.append("name ({} -> {})".format(existing_name, type_name))
                    if thickness_changed:
                        prompt_parts.append("layer thickness")
                    prompt = (
                        "Type '{}' has changes in {}.\n\n"
                        "Overwrite from Excel?"
                    ).format(elem_type.Name, " and ".join(prompt_parts))
                    allow_overwrite = False
                    if overwrite_mode == OVERWRITE_MODE_ALL:
                        allow_overwrite = True
                    elif overwrite_mode == OVERWRITE_MODE_SKIP:
                        allow_overwrite = False
                    else:
                        allow_overwrite = ui.uiUtils_confirm(prompt, title="Import Type Layers")
                    if not allow_overwrite:
                        log_warn("User declined overwrite for '{}'.".format(elem_type.Name))
                        overwrite_skipped += 1
                        skipped += 1
                        continue
                    log_info("User accepted overwrite for '{}' ({}).".format(elem_type.Name, " and ".join(prompt_parts)))

                if name_changed:
                    log_info("Renaming type '{}' -> '{}'.".format(existing_name, type_name))
                    rename_type(elem_type, type_name)

                ok, err = update_type_layers(doc, elem_type, layers)
                if ok:
                    log_info("Updated '{}' with {} layer(s).".format(elem_type.Name, len(layers)))
                    updated += 1
                else:
                    skipped += 1
                    log_error("Failed to update '{}': {}.".format(elem_type.Name, err))
                    if err:
                        errors.append("{}: {}".format(elem_type.Name, err))

            missing = [t for t in types if element_id_value(t.Id) not in excel_ids]
            if missing:
                log_warn("{} type(s) exist in Revit but not in Excel for {}.".format(len(missing), label))
                missing_types_summary.append(
                    (label, [t.Name for t in missing])
                )

    if missing_types_summary:
        confirm = ui.uiUtils_confirm(
            "Some Revit types are not listed in the Excel file.\n"
            "Do you want to delete those types?\n\n"
            "This cannot be undone.",
            title="Import Type Layers",
        )
        if confirm:
            log_section("Delete Missing Types")
            with revit.Transaction("Delete Types Not In Excel"):
                for label, names in missing_types_summary:
                    items, _, _, by_name = collect_types_by_category(
                        doc,
                        DB.WallType if label == "Wall Types" else
                        DB.CeilingType if label == "Ceiling Types" else
                        DB.FloorType if label == "Floor Types" else
                        DB.RoofType
                    )
                    for name in names:
                        elem = by_name.get(name)
                        if elem:
                            try:
                                doc.Delete(elem.Id)
                                log_info("Deleted {} type '{}' because it was not found in Excel.".format(label, name))
                            except Exception:
                                log_warn("Failed to delete {} type '{}'.".format(label, name))
                                pass
        else:
            log_info("User kept Revit types that were missing from Excel.")

    summary = "Updated: {}\nCreated: {}\nSkipped: {}".format(updated, created, skipped)
    if overwrite_skipped:
        summary += "\nSkipped (user chose not to overwrite): {}".format(overwrite_skipped)
    if errors:
        summary += "\n\nErrors:\n" + "\n".join(errors[:10])
        if len(errors) > 10:
            summary += "\n... ({}) more".format(len(errors) - 10)
    log_section("Summary")
    for line in summary.splitlines():
        if line.strip():
            log_info(line)
    ui.uiUtils_alert(summary, title="Import Type Layers")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log_section("Unhandled Exception")
        log_error(traceback.format_exc())
        raise
