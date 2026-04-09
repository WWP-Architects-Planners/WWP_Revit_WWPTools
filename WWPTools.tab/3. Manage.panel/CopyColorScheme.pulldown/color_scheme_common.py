# -*- coding: utf-8 -*-

import os
import re
from pyrevit import DB
from System import Activator, Type
from System.Runtime.InteropServices import Marshal
try:
    from System.Reflection import BindingFlags
except Exception:
    BindingFlags = None

import WWP_colorSchemeUtils as csu
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    get_column_letter = None

EXCEL_SHEET_NAME = "ColorScheme"
FORMAT_VERSION = "1"
_BASE_FLAGS = None


def _binding_flag(name):
    if BindingFlags is None:
        return None
    try:
        return getattr(BindingFlags, name)
    except Exception:
        return None


def _base_flags():
    public_flag = _binding_flag("Public")
    instance_flag = _binding_flag("Instance")
    optional_flag = _binding_flag("OptionalParamBinding")
    if public_flag is None or instance_flag is None or optional_flag is None:
        return None
    try:
        return public_flag | instance_flag | optional_flag
    except Exception:
        return None


def elem_id_int(elem_id):
    try:
        return int(elem_id.IntegerValue)
    except Exception:
        pass
    try:
        return int(elem_id.Value)  # Revit 2024+ uses .Value instead of .IntegerValue
    except Exception:
        pass
    try:
        return int(elem_id)
    except Exception:
        return None


def scheme_area_scheme_id(scheme):
    try:
        area_scheme_id = getattr(scheme, "AreaSchemeId", None)
        if area_scheme_id and area_scheme_id != DB.ElementId.InvalidElementId:
            return area_scheme_id
    except Exception:
        pass
    try:
        getter = getattr(scheme, "GetAreaSchemeId", None)
        if callable(getter):
            area_scheme_id = getter()
            if area_scheme_id and area_scheme_id != DB.ElementId.InvalidElementId:
                return area_scheme_id
    except Exception:
        pass
    return None


def scheme_area_scheme_name(doc, scheme):
    try:
        area_scheme_id = scheme_area_scheme_id(scheme)
        if area_scheme_id:
            area_scheme = doc.GetElement(area_scheme_id)
            if area_scheme:
                return getattr(area_scheme, "Name", "") or ""
    except Exception:
        pass
    return ""


def category_name(doc, category_id):
    if category_id is None:
        return "Unknown Category"
    cat_int = elem_id_int(category_id)
    try:
        cat = doc.GetElement(category_id)
        cat_name = getattr(cat, "Name", "") if cat else ""
        if cat_name:
            return cat_name
    except Exception:
        pass
    try:
        categories = getattr(getattr(doc, "Settings", None), "Categories", None)
        if categories is not None:
            cat = categories.get_Item(category_id)
            cat_name = getattr(cat, "Name", "") if cat else ""
            if cat_name:
                return cat_name
    except Exception:
        pass
    try:
        import System
        bic = System.Enum.ToObject(DB.BuiltInCategory, cat_int)
        categories = getattr(getattr(doc, "Settings", None), "Categories", None)
        if categories is not None:
            cat = categories.get_Item(bic)
            cat_name = getattr(cat, "Name", "") if cat else ""
            if cat_name:
                return cat_name
    except Exception:
        pass
    try:
        import System
        bic = System.Enum.ToObject(DB.BuiltInCategory, cat_int)
        label = DB.LabelUtils.GetLabelFor(bic)
        if label:
            return label
    except Exception:
        pass
    try:
        label = DB.LabelUtils.GetLabelFor(category_id)
        if label:
            return label
    except Exception:
        pass
    if cat_int is not None:
        return "Category {}".format(cat_int)
    return "Unknown Category"


def scheme_scope_label(doc, scheme):
    area_name = scheme_area_scheme_name(doc, scheme)
    if area_name:
        return "Area({})".format(area_name)
    return category_name(doc, getattr(scheme, "CategoryId", None))


def scheme_display_name(doc, scheme):
    scheme_name = getattr(scheme, "Name", "") or "Color Scheme"
    return "{}: {}".format(scheme_scope_label(doc, scheme), scheme_name)


def describe_scheme(scheme):
    return "scheme='{}' id={} categoryId={} areaSchemeId={}".format(
        getattr(scheme, "Name", "<unnamed>"),
        elem_id_int(getattr(scheme, "Id", None)),
        elem_id_int(getattr(scheme, "CategoryId", None)),
        elem_id_int(scheme_area_scheme_id(scheme)),
    )


def storage_type_token(storage_type):
    if storage_type == DB.StorageType.String:
        return "String"
    if storage_type == DB.StorageType.Integer:
        return "Integer"
    if storage_type == DB.StorageType.Double:
        return "Double"
    if storage_type == DB.StorageType.ElementId:
        return "ElementId"
    return "Unknown"


def parse_storage_type(token):
    value = (token or "").strip().lower()
    if value == "string":
        return DB.StorageType.String
    if value == "integer":
        return DB.StorageType.Integer
    if value == "double":
        return DB.StorageType.Double
    if value == "elementid":
        return DB.StorageType.ElementId
    return None


def entry_value_to_text(entry, storage_type, doc=None):
    value = csu._entry_get_value(entry, storage_type)
    if storage_type == DB.StorageType.Double:
        try:
            return "{:.12g}".format(float(value))
        except Exception:
            return ""
    if storage_type == DB.StorageType.ElementId:
        if value is None:
            return ""
        # For key-schedule parameters the entry stores the key row ElementId.
        # Resolve to the element's Name so callers can match by human-readable key
        # name rather than raw integer ID.
        if doc is not None:
            try:
                elem = doc.GetElement(value)
                name = getattr(elem, "Name", None) if elem else None
                if name:
                    return str(name)
            except Exception:
                pass
        return str(elem_id_int(value))
    return "" if value is None else str(value)


def _scheme_parameter_id(scheme):
    try:
        pid = getattr(scheme, "ParameterId", None)
        if pid is not None:
            return pid
    except Exception:
        pass
    return None


def _scheme_parameter_id_int(scheme):
    return elem_id_int(_scheme_parameter_id(scheme))


def _scheme_parameter_label(doc, scheme):
    pid = _scheme_parameter_id(scheme)
    if pid is None:
        return ""
    try:
        if pid == DB.ElementId.InvalidElementId:
            return ""
    except Exception:
        pass
    try:
        elem = doc.GetElement(pid)
        name = getattr(elem, "Name", "") if elem else ""
        if name:
            return name
    except Exception:
        pass
    try:
        binding = getattr(scheme, "ParameterDefinition", None)
        name = getattr(binding, "Name", "") if binding else ""
        if name:
            return name
    except Exception:
        pass
    pid_int = elem_id_int(pid)
    return "" if pid_int is None else "ParameterId {}".format(pid_int)


def build_payload_from_scheme(doc, scheme):
    payload = {
        "format_version": FORMAT_VERSION,
        "scheme_name": getattr(scheme, "Name", "") or "Color Scheme",
        "category_name": category_name(doc, getattr(scheme, "CategoryId", None)),
        "category_id": elem_id_int(getattr(scheme, "CategoryId", None)),
        "area_scheme_name": scheme_area_scheme_name(doc, scheme),
        "title": getattr(scheme, "Title", "") or "",
        "is_by_range": bool(getattr(scheme, "IsByRange", False)),
        "is_by_value": bool(getattr(scheme, "IsByValue", False)),
        "is_by_percentage": bool(getattr(scheme, "IsByPercentage", False)),
        "parameter_id": _scheme_parameter_id_int(scheme),
        "parameter_name": _scheme_parameter_label(doc, scheme),
        "entries": [],
    }
    try:
        entries = list(scheme.GetEntries())
    except Exception:
        entries = []
    for entry in entries:
        color = getattr(entry, "Color", None)
        payload["entries"].append({
            "caption": csu._entry_caption(entry),
            "storage_type": storage_type_token(getattr(entry, "StorageType", None)),
            "value": entry_value_to_text(entry, getattr(entry, "StorageType", None)),
            "color_r": "" if color is None else int(getattr(color, "Red", 0)),
            "color_g": "" if color is None else int(getattr(color, "Green", 0)),
            "color_b": "" if color is None else int(getattr(color, "Blue", 0)),
            "fill_pattern_id": elem_id_int(getattr(entry, "FillPatternId", None)),
            "is_visible": bool(getattr(entry, "IsVisible", getattr(entry, "Visible", True))),
        })
    return payload


def _com_get(obj, name):
    flags = _base_flags()
    if flags is None:
        raise Exception("Reflection BindingFlags are unavailable.")
    try:
        return obj.GetType().InvokeMember(name, flags | _binding_flag("GetProperty"), None, obj, None)
    except Exception:
        return obj.GetType().InvokeMember(name, flags | _binding_flag("InvokeMethod"), None, obj, [])


def _com_set(obj, name, value):
    flags = _base_flags()
    if flags is None:
        raise Exception("Reflection BindingFlags are unavailable.")
    try:
        obj.GetType().InvokeMember(name, flags | _binding_flag("SetProperty"), None, obj, [value])
    except Exception:
        obj.GetType().InvokeMember(name, flags | _binding_flag("InvokeMethod"), None, obj, [value])


def _com_call(obj, name, *args):
    flags = _base_flags()
    if flags is None:
        raise Exception("Reflection BindingFlags are unavailable.")
    try:
        return obj.GetType().InvokeMember(name, flags | _binding_flag("InvokeMethod"), None, obj, list(args))
    except Exception:
        return obj.GetType().InvokeMember(name, flags | _binding_flag("GetProperty"), None, obj, list(args))


def _com_call_strict(obj, name, *args):
    flags = _base_flags()
    if flags is None:
        raise Exception("Reflection BindingFlags are unavailable.")
    return obj.GetType().InvokeMember(name, flags | _binding_flag("InvokeMethod"), None, obj, list(args))


def _try_delete_file(path, log=None):
    try:
        if os.path.isfile(path):
            os.remove(path)
            _log(log, "Deleted existing file before save: {}".format(path))
    except Exception as ex:
        _log(log, "Could not delete existing file '{}': {}".format(path, str(ex)))


def _save_workbook_with_fallbacks(workbook, path, log=None):
    # 51 = xlOpenXMLWorkbook (.xlsx)
    save_attempts = [
        ("SaveAs-xlsx", lambda: _com_call_strict(workbook, "SaveAs", path, 51)),
        ("SaveAs-basic", lambda: _com_call_strict(workbook, "SaveAs", path)),
        ("SaveCopyAs", lambda: _com_call_strict(workbook, "SaveCopyAs", path)),
    ]
    for label, action in save_attempts:
        _try_delete_file(path, log=log)
        _log(log, "Trying workbook save strategy: {}".format(label))
        try:
            action()
        except Exception as ex:
            _log(log, "Save strategy {} failed: {}".format(label, str(ex)))
            continue
        if os.path.isfile(path):
            _log(log, "Save strategy {} succeeded.".format(label))
            return True
        _log(log, "Save strategy {} returned without error, but file is still missing.".format(label))
    return False


def _log(log, message):
    if callable(log):
        try:
            log(message)
        except Exception:
            pass


def export_payload_to_excel(payload, path, log=None):
    if Workbook is not None:
        _log(log, "Using openpyxl export path.")
        wb = Workbook()
        ws = wb.active
        ws.title = EXCEL_SHEET_NAME

        rows = [
            ["WWP Color Scheme Export", ""],
            ["FormatVersion", payload.get("format_version", FORMAT_VERSION)],
            ["SchemeName", payload.get("scheme_name", "")],
            ["CategoryName", payload.get("category_name", "")],
            ["CategoryId", payload.get("category_id", "")],
            ["AreaSchemeName", payload.get("area_scheme_name", "")],
            ["Title", payload.get("title", "")],
            ["ParameterId", payload.get("parameter_id", "")],
            ["ParameterName", payload.get("parameter_name", "")],
            ["IsByRange", "TRUE" if payload.get("is_by_range") else "FALSE"],
            ["IsByValue", "TRUE" if payload.get("is_by_value") else "FALSE"],
            ["IsByPercentage", "TRUE" if payload.get("is_by_percentage") else "FALSE"],
            [],
            ["Caption", "StorageType", "Value", "ColorR", "ColorG", "ColorB", "FillPatternId", "IsVisible"],
        ]
        for entry in payload.get("entries", []):
            rows.append([
                entry.get("caption", ""),
                entry.get("storage_type", ""),
                entry.get("value", ""),
                entry.get("color_r", ""),
                entry.get("color_g", ""),
                entry.get("color_b", ""),
                entry.get("fill_pattern_id", ""),
                "TRUE" if entry.get("is_visible", True) else "FALSE",
            ])

        _log(log, "Prepared {} rows for openpyxl export.".format(len(rows)))
        for row in rows:
            ws.append(row)
        for column_cells in ws.columns:
            width = 0
            for cell in column_cells:
                try:
                    width = max(width, len("" if cell.value is None else str(cell.value)))
                except Exception:
                    pass
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 60)

        directory = os.path.dirname(path)
        if directory and not os.path.isdir(directory):
            os.makedirs(directory)
            _log(log, "Created export directory: {}".format(directory))
        _try_delete_file(path, log=log)
        wb.save(path)
        if not os.path.isfile(path):
            raise Exception("openpyxl completed without creating '{}'.".format(path))
        _log(log, "Workbook saved successfully via openpyxl.")
        return

    _log(log, "openpyxl unavailable. Falling back to Excel COM export.")
    excel_app = None
    workbook = None
    worksheet = None
    used_range = None
    try:
        _log(log, "Starting Excel export.")
        excel_type = Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            raise Exception("Excel COM ProgID not found. Verify Microsoft Excel is installed.")
        _log(log, "Excel COM type acquired.")
        excel_app = Activator.CreateInstance(excel_type)
        _log(log, "Excel application created.")
        _com_set(excel_app, "Visible", False)
        _com_set(excel_app, "DisplayAlerts", False)
        workbooks = _com_get(excel_app, "Workbooks")
        _log(log, "Excel workbooks collection acquired.")
        workbook = _com_call(workbooks, "Add")
        _log(log, "Workbook created.")
        worksheets = _com_get(workbook, "Worksheets")
        worksheet = _com_call(worksheets, "Item", 1)
        _com_set(worksheet, "Name", EXCEL_SHEET_NAME)
        _log(log, "Worksheet prepared: {}".format(EXCEL_SHEET_NAME))

        rows = [
            ["WWP Color Scheme Export", ""],
            ["FormatVersion", payload.get("format_version", FORMAT_VERSION)],
            ["SchemeName", payload.get("scheme_name", "")],
            ["CategoryName", payload.get("category_name", "")],
            ["CategoryId", payload.get("category_id", "")],
            ["AreaSchemeName", payload.get("area_scheme_name", "")],
            ["Title", payload.get("title", "")],
            ["ParameterId", payload.get("parameter_id", "")],
            ["ParameterName", payload.get("parameter_name", "")],
            ["IsByRange", "TRUE" if payload.get("is_by_range") else "FALSE"],
            ["IsByValue", "TRUE" if payload.get("is_by_value") else "FALSE"],
            ["IsByPercentage", "TRUE" if payload.get("is_by_percentage") else "FALSE"],
            [],
            ["Caption", "StorageType", "Value", "ColorR", "ColorG", "ColorB", "FillPatternId", "IsVisible"],
        ]
        for entry in payload.get("entries", []):
            rows.append([
                entry.get("caption", ""),
                entry.get("storage_type", ""),
                entry.get("value", ""),
                entry.get("color_r", ""),
                entry.get("color_g", ""),
                entry.get("color_b", ""),
                entry.get("fill_pattern_id", ""),
                "TRUE" if entry.get("is_visible", True) else "FALSE",
            ])

        row_count = len(rows)
        col_count = max(len(r) for r in rows) if rows else 1
        _log(log, "Prepared {} rows x {} columns for export.".format(row_count, col_count))
        cells = _com_get(worksheet, "Cells")
        for r_idx, row in enumerate(rows, start=1):
            for c_idx in range(1, col_count + 1):
                value = row[c_idx - 1] if c_idx - 1 < len(row) else ""
                cell = _com_call(cells, "Item", r_idx, c_idx)
                _com_set(cell, "Value2", value)
        _log(log, "Cell values written to worksheet.")

        start_cell = _com_call(cells, "Item", 1, 1)
        end_cell = _com_call(cells, "Item", row_count, col_count)
        range_obj = _com_call(worksheet, "Range", start_cell, end_cell)
        used_range = _com_get(worksheet, "UsedRange")
        _com_call(_com_get(used_range, "Columns"), "AutoFit")
        _log(log, "Columns auto-fitted.")

        directory = os.path.dirname(path)
        if directory and not os.path.isdir(directory):
            os.makedirs(directory)
            _log(log, "Created export directory: {}".format(directory))
        _log(log, "Saving workbook to: {}".format(path))
        if not _save_workbook_with_fallbacks(workbook, path, log=log):
            raise Exception("Excel reported success but no file was created at '{}'.".format(path))
        _log(log, "Workbook saved successfully.")
    finally:
        try:
            if workbook is not None:
                _log(log, "Closing workbook.")
                _com_call_strict(workbook, "Close", False)
        except Exception:
            pass
        try:
            if excel_app is not None:
                _log(log, "Quitting Excel application.")
                _com_call_strict(excel_app, "Quit")
        except Exception:
            pass
        for obj in (used_range, worksheet, workbook, excel_app):
            try:
                if obj is not None:
                    Marshal.ReleaseComObject(obj)
            except Exception:
                pass


def _to_bool(text):
    return str(text or "").strip().lower() in ("true", "1", "yes", "y")


def _to_int(text, default=None):
    try:
        if text is None or str(text).strip() == "":
            return default
        return int(float(str(text).strip()))
    except Exception:
        return default


def _to_float(text, default=None):
    try:
        if text is None or str(text).strip() == "":
            return default
        return float(str(text).strip())
    except Exception:
        return default


def inspect_excel_workbook(path, max_preview_rows=8, max_columns=32):
    if load_workbook is None:
        raise Exception("openpyxl is not available for Excel column mapping.")

    wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        max_col = int(getattr(ws, "max_column", 0) or 0)
        max_row = int(getattr(ws, "max_row", 0) or 0)
        if max_col <= 0:
            continue
        max_col = min(max_col, max_columns)
        columns = []
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx) if callable(get_column_letter) else str(col_idx)
            header_value = ws.cell(row=1, column=col_idx).value
            header_text = "" if header_value is None else str(header_value).strip()
            columns.append({
                "index": col_idx - 1,
                "letter": letter,
                "header": header_text,
                "label": "{} - {}".format(letter, header_text or "Column {}".format(col_idx)),
            })

        preview_rows = []
        for row_idx in range(2, max_row + 1):
            row_values = []
            has_content = False
            for col_idx in range(1, max_col + 1):
                raw = ws.cell(row=row_idx, column=col_idx).value
                text = "" if raw is None else str(raw).strip()
                if text:
                    has_content = True
                row_values.append(text)
            if not has_content:
                continue
            preview_rows.append(row_values)
            if len(preview_rows) >= max_preview_rows:
                break

        sheets.append({
            "name": ws.title,
            "columns": columns,
            "preview_rows": preview_rows,
            "row_count": max_row,
        })

    if not sheets:
        raise Exception("No readable worksheets were found in the workbook.")
    return {"sheets": sheets}


def import_payload_from_excel(path):
    if load_workbook is not None:
        wb = load_workbook(path, data_only=True)
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            raise Exception("Workbook is not in WWP color scheme export format.")
        ws = wb[EXCEL_SHEET_NAME]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if value is None else str(value) for value in row])

        if not rows or not rows[0] or rows[0][0].strip() != "WWP Color Scheme Export":
            raise Exception("The workbook does not match the WWP color scheme export format.")

        metadata = {}
        entry_start = None
        for idx, row in enumerate(rows[1:], start=1):
            first = row[0].strip() if row else ""
            if first == "Caption":
                entry_start = idx + 1
                break
            if first:
                metadata[first] = row[1].strip() if len(row) > 1 else ""
        if entry_start is None:
            raise Exception("Entry header row was not found in the workbook.")

        payload = {
            "format_version": metadata.get("FormatVersion", FORMAT_VERSION),
            "scheme_name": metadata.get("SchemeName", "Imported Color Scheme"),
            "category_name": metadata.get("CategoryName", ""),
            "category_id": _to_int(metadata.get("CategoryId"), None),
            "area_scheme_name": metadata.get("AreaSchemeName", ""),
            "title": metadata.get("Title", ""),
            "parameter_id": _to_int(metadata.get("ParameterId"), None),
            "parameter_name": metadata.get("ParameterName", ""),
            "is_by_range": _to_bool(metadata.get("IsByRange")),
            "is_by_value": _to_bool(metadata.get("IsByValue")),
            "is_by_percentage": _to_bool(metadata.get("IsByPercentage")),
            "entries": [],
        }
        for row in rows[entry_start:]:
            if not any((cell or "").strip() for cell in row):
                continue
            payload["entries"].append({
                "caption": row[0].strip() if len(row) > 0 else "",
                "storage_type": row[1].strip() if len(row) > 1 else "",
                "value": row[2].strip() if len(row) > 2 else "",
                "color_r": _to_int(row[3] if len(row) > 3 else "", None),
                "color_g": _to_int(row[4] if len(row) > 4 else "", None),
                "color_b": _to_int(row[5] if len(row) > 5 else "", None),
                "fill_pattern_id": _to_int(row[6] if len(row) > 6 else "", None),
                "is_visible": _to_bool(row[7] if len(row) > 7 else "TRUE"),
            })
        return payload

    excel_app = None
    workbook = None
    worksheet = None
    used_range = None
    try:
        excel_type = Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            raise Exception("Excel COM ProgID not found. Verify Microsoft Excel is installed.")
        excel_app = Activator.CreateInstance(excel_type)
        _com_set(excel_app, "Visible", False)
        _com_set(excel_app, "DisplayAlerts", False)
        workbooks = _com_get(excel_app, "Workbooks")
        workbook = _com_call(workbooks, "Open", path)
        worksheets = _com_get(workbook, "Worksheets")
        worksheet = _com_call(worksheets, "Item", EXCEL_SHEET_NAME)
        used_range = _com_get(worksheet, "UsedRange")
        values = _com_get(used_range, "Value2")
        rows_count = int(_com_get(_com_get(used_range, "Rows"), "Count"))
        cols_count = int(_com_get(_com_get(used_range, "Columns"), "Count"))
        lower_r = int(values.GetLowerBound(0))
        lower_c = int(values.GetLowerBound(1))

        rows = []
        for r in range(rows_count):
            row = []
            for c in range(cols_count):
                value = values.GetValue(lower_r + r, lower_c + c)
                row.append("" if value is None else str(value))
            rows.append(row)

        if not rows or not rows[0] or rows[0][0].strip() != "WWP Color Scheme Export":
            raise Exception("The workbook does not match the WWP color scheme export format.")

        metadata = {}
        entry_start = None
        for idx, row in enumerate(rows[1:], start=1):
            first = row[0].strip() if row else ""
            if first == "Caption":
                entry_start = idx + 1
                break
            if first:
                metadata[first] = row[1].strip() if len(row) > 1 else ""
        if entry_start is None:
            raise Exception("Entry header row was not found in the workbook.")

        payload = {
            "format_version": metadata.get("FormatVersion", FORMAT_VERSION),
            "scheme_name": metadata.get("SchemeName", "Imported Color Scheme"),
            "category_name": metadata.get("CategoryName", ""),
            "category_id": _to_int(metadata.get("CategoryId"), None),
            "area_scheme_name": metadata.get("AreaSchemeName", ""),
            "title": metadata.get("Title", ""),
            "parameter_id": _to_int(metadata.get("ParameterId"), None),
            "parameter_name": metadata.get("ParameterName", ""),
            "is_by_range": _to_bool(metadata.get("IsByRange")),
            "is_by_value": _to_bool(metadata.get("IsByValue")),
            "is_by_percentage": _to_bool(metadata.get("IsByPercentage")),
            "entries": [],
        }
        for row in rows[entry_start:]:
            if not any((cell or "").strip() for cell in row):
                continue
            payload["entries"].append({
                "caption": row[0].strip() if len(row) > 0 else "",
                "storage_type": row[1].strip() if len(row) > 1 else "",
                "value": row[2].strip() if len(row) > 2 else "",
                "color_r": _to_int(row[3] if len(row) > 3 else "", None),
                "color_g": _to_int(row[4] if len(row) > 4 else "", None),
                "color_b": _to_int(row[5] if len(row) > 5 else "", None),
                "fill_pattern_id": _to_int(row[6] if len(row) > 6 else "", None),
                "is_visible": _to_bool(row[7] if len(row) > 7 else "TRUE"),
            })
        return payload
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        try:
            if excel_app is not None:
                excel_app.Quit()
        except Exception:
            pass
        for obj in (used_range, worksheet, workbook, excel_app):
            try:
                if obj is not None:
                    Marshal.ReleaseComObject(obj)
            except Exception:
                pass


def _hex_pair_to_int(value):
    try:
        return int(value, 16)
    except Exception:
        return None


def _parse_color_text(value):
    text = "" if value is None else str(value).strip()
    if not text:
        return None
    if text.startswith("#"):
        text = text[1:]
    text = text.strip()
    if re.match(r"^[0-9A-Fa-f]{6}$", text):
        return (
            _hex_pair_to_int(text[0:2]),
            _hex_pair_to_int(text[2:4]),
            _hex_pair_to_int(text[4:6]),
        )
    if re.match(r"^[0-9A-Fa-f]{8}$", text):
        return (
            _hex_pair_to_int(text[2:4]),
            _hex_pair_to_int(text[4:6]),
            _hex_pair_to_int(text[6:8]),
        )

    normalized = re.sub(r"(?i)rgb", "", text)
    parts = [part for part in re.split(r"[^0-9]+", normalized) if part != ""]
    if len(parts) >= 3:
        try:
            r = max(0, min(255, int(parts[0])))
            g = max(0, min(255, int(parts[1])))
            b = max(0, min(255, int(parts[2])))
            return (r, g, b)
        except Exception:
            return None
    return None


def _openpyxl_cell_fill_rgb(cell):
    try:
        fill = getattr(cell, "fill", None)
        if fill is None:
            return None
        if getattr(fill, "patternType", None) not in ("solid", "gray125"):
            return None
        for color_attr in ("fgColor", "start_color"):
            color = getattr(fill, color_attr, None)
            if color is None:
                continue
            rgb = getattr(color, "rgb", None)
            if rgb:
                parsed = _parse_color_text(rgb)
                if parsed:
                    return parsed
    except Exception:
        pass
    return None


def _coerce_mapped_value(raw_value, storage_type):
    if storage_type == "Integer":
        value = _to_int(raw_value, None)
        if value is None:
            raise Exception("Value '{}' is not a valid integer.".format(raw_value))
        return str(value)
    if storage_type == "Double":
        value = _to_float(raw_value, None)
        if value is None:
            raise Exception("Value '{}' is not a valid number.".format(raw_value))
        return "{:.12g}".format(value)
    if storage_type == "ElementId":
        value = _to_int(raw_value, None)
        if value is None:
            raise Exception(
                "Value '{}' is not a valid ElementId.\n"
                "The selected target color scheme uses an ElementId-backed parameter, so mapped values must be numeric Revit element ids.\n"
                "Choose a text/string-based target scheme instead, or supply numeric ids in the value column."
                .format(raw_value)
            )
        return str(value)
    return "" if raw_value is None else str(raw_value).strip()


def _normalize_match_text(value):
    return " ".join(str(value or "").strip().lower().split())


def _match_aliases(value):
    normalized = _normalize_match_text(value)
    if not normalized:
        return []
    aliases = [normalized]
    spaced = " ".join(normalized.replace("_", " ").replace("-", " ").split())
    compact = re.sub(r"[\s_-]+", "", normalized)
    for alias in (spaced, compact):
        if alias and alias not in aliases:
            aliases.append(alias)
    return aliases


def _elementid_entry_lookup(target_scheme):
    lookup = {}
    doc = getattr(target_scheme, "Document", None)
    try:
        entries = list(target_scheme.GetEntries())
    except Exception:
        entries = []
    for entry in entries:
        caption = csu._entry_caption(entry)
        storage_type = getattr(entry, "StorageType", None)
        value = csu._entry_get_value(entry, storage_type)
        value_int = elem_id_int(value)
        if value_int is None:
            continue
        keys = []
        resolved_value_text = entry_value_to_text(entry, storage_type, doc=doc)
        for candidate in (caption, resolved_value_text, csu._format_entry_value(entry)):
            for alias in _match_aliases(candidate):
                if alias not in keys:
                    keys.append(alias)
        for key in keys:
            if key and key not in lookup:
                lookup[key] = {"caption": caption, "value": value_int}
    return lookup


def _default_fill_pattern_id(doc, target_scheme):
    try:
        for entry in list(target_scheme.GetEntries()):
            fill_pattern_id = getattr(entry, "FillPatternId", None)
            if fill_pattern_id and fill_pattern_id != DB.ElementId.InvalidElementId:
                return elem_id_int(fill_pattern_id)
    except Exception:
        pass

    try:
        collector = DB.FilteredElementCollector(doc).OfClass(DB.FillPatternElement)
        for element in collector:
            pattern = element.GetFillPattern()
            if pattern is not None and getattr(pattern, "IsSolidFill", False):
                return elem_id_int(element.Id)
    except Exception:
        pass
    return None


def build_payload_from_mapped_excel(path, mapping, target_snapshot):
    if load_workbook is None:
        raise Exception("openpyxl is not available for Excel column mapping.")

    sheet_name = (mapping.get("sheet_name") or "").strip()
    if not sheet_name:
        raise Exception("No worksheet was selected.")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise Exception("Worksheet '{}' was not found in the workbook.".format(sheet_name))

    ws = wb[sheet_name]
    value_index = int(mapping.get("value_column_index", -1))
    if value_index < 0:
        raise Exception("No value column was selected.")
    color_index = int(mapping.get("color_column_index", -1))
    label_index = int(mapping.get("label_column_index", value_index))
    use_fill_color = bool(mapping.get("use_fill_color", True))
    storage_type = ""
    entry_storage_types = target_snapshot.get("entry_storage_types") or []
    if entry_storage_types:
        storage_type = entry_storage_types[0]
    if not storage_type:
        storage_type = "String"

    category_name_value = target_snapshot.get("category_name", "")
    area_scheme_name_value = target_snapshot.get("area_scheme_name", "")
    base_name = os.path.splitext(os.path.basename(path))[0]
    payload = {
        "format_version": FORMAT_VERSION,
        "scheme_name": "{} Import".format(base_name or "Color Scheme"),
        "category_name": category_name_value,
        "category_id": target_snapshot.get("category_id"),
        "area_scheme_name": area_scheme_name_value,
        "title": target_snapshot.get("title", ""),
        "parameter_id": target_snapshot.get("parameter_id"),
        "parameter_name": target_snapshot.get("parameter_name", ""),
        "is_by_range": bool(target_snapshot.get("is_by_range")),
        "is_by_value": bool(target_snapshot.get("is_by_value")),
        "is_by_percentage": bool(target_snapshot.get("is_by_percentage")),
        "entries": [],
    }

    default_fill_pattern_id = None
    try:
        from Autodesk.Revit import DB as _DB  # keep namespace local-safe
        _ = _DB
    except Exception:
        pass

    doc = mapping.get("doc")
    target_scheme = mapping.get("target_scheme")
    if doc is not None and target_scheme is not None:
        default_fill_pattern_id = _default_fill_pattern_id(doc, target_scheme)
    elementid_lookup = _elementid_entry_lookup(target_scheme) if storage_type == "ElementId" and target_scheme is not None else {}

    seen_values = set()
    for row_idx in range(2, int(getattr(ws, "max_row", 0) or 0) + 1):
        value_cell = ws.cell(row=row_idx, column=value_index + 1)
        raw_value = value_cell.value
        if raw_value is None or str(raw_value).strip() == "":
            continue
        raw_value_text = "" if raw_value is None else str(raw_value).strip()
        if storage_type == "ElementId":
            matched = None
            for alias in _match_aliases(raw_value_text):
                matched = elementid_lookup.get(alias)
                if matched is not None:
                    break
            if matched is None:
                raise Exception(
                    "Value '{}' did not match any existing displayed values in the selected ElementId-backed color scheme.\n"
                    "Pick a target scheme whose existing entries contain this key, or choose a text-based target scheme."
                    .format(raw_value_text)
                )
            value_text = str(matched.get("value"))
        else:
            value_text = _coerce_mapped_value(raw_value, storage_type)
        if value_text in seen_values:
            continue
        seen_values.add(value_text)

        label_cell = ws.cell(row=row_idx, column=label_index + 1) if label_index >= 0 else value_cell
        caption = "" if label_cell.value is None else str(label_cell.value).strip()
        if not caption:
            if storage_type == "ElementId":
                caption = (matched.get("caption") if matched is not None else "") or raw_value_text
            else:
                caption = value_text

        rgb = None
        if color_index >= 0:
            color_cell = ws.cell(row=row_idx, column=color_index + 1)
            if use_fill_color:
                rgb = _openpyxl_cell_fill_rgb(color_cell)
            if rgb is None:
                rgb = _parse_color_text(color_cell.value)
        if rgb is None:
            raise Exception("No usable color was found for row {}.".format(row_idx))

        payload["entries"].append({
            "caption": caption,
            "storage_type": storage_type,
            "value": value_text,
            "color_r": rgb[0],
            "color_g": rgb[1],
            "color_b": rgb[2],
            "fill_pattern_id": default_fill_pattern_id,
            "is_visible": True,
        })

    if not payload["entries"]:
        raise Exception("No mapped rows with both value and color data were found.")
    return payload


def build_color_map_from_excel(path, mapping):
    """Read an Excel file and return a {value_text: (r, g, b)} lookup."""
    if load_workbook is None:
        raise Exception("openpyxl is not available for Excel column mapping.")

    sheet_name = (mapping.get("sheet_name") or "").strip()
    if not sheet_name:
        raise Exception("No worksheet was selected.")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise Exception("Worksheet '{}' was not found in the workbook.".format(sheet_name))

    ws = wb[sheet_name]
    value_index = int(mapping.get("value_column_index", -1))
    if value_index < 0:
        raise Exception("No value column was selected.")
    color_index = int(mapping.get("color_column_index", -1))
    if color_index < 0:
        raise Exception("No color column was selected.")
    use_fill_color = bool(mapping.get("use_fill_color", True))

    color_map = {}
    for row_idx in range(2, int(getattr(ws, "max_row", 0) or 0) + 1):
        value_cell = ws.cell(row=row_idx, column=value_index + 1)
        raw_value = value_cell.value
        if raw_value is None or str(raw_value).strip() == "":
            continue
        value_text = str(raw_value).strip()

        color_cell = ws.cell(row=row_idx, column=color_index + 1)
        rgb = None
        if use_fill_color:
            rgb = _openpyxl_cell_fill_rgb(color_cell)
        if rgb is None:
            rgb = _parse_color_text(color_cell.value)
        if rgb is None:
            continue
        color_map[value_text] = rgb

    if not color_map:
        raise Exception("No rows with both a value and a color were found in the workbook.")
    return color_map





def apply_color_map_to_scheme(target, color_map, log=None):
    """Update only the colors of existing scheme entries whose value matches a key in color_map."""
    _log(log, "Applying to scheme: '{}' (Id={}) categoryId={} areaSchemeId={}".format(
        getattr(target, "Name", "<unknown>"),
        elem_id_int(getattr(target, "Id", None)),
        elem_id_int(getattr(target, "CategoryId", None)),
        elem_id_int(scheme_area_scheme_id(target)),
    ))
    try:
        entries = list(target.GetEntries())
    except Exception:
        return False, "Could not read existing entries from the color scheme."

    _log(log, "=== Color override: {} scheme entries, {} Excel keys ===".format(
        len(entries), len(color_map)))
    _log(log, "Excel keys (first 5): {}".format(list(color_map.keys())[:5]))

    doc = getattr(target, "Document", None)
    updated = 0
    skipped = []
    for i, entry in enumerate(entries):
        storage_type = getattr(entry, "StorageType", None)
        raw_value = csu._entry_get_value(entry, storage_type)
        caption = csu._entry_caption(entry)
        value_text = entry_value_to_text(entry, storage_type, doc=doc)
        if i < 5:
            _log(log, "Entry[{}] storage_type={} raw_value={!r} value_text={!r} caption={!r}".format(
                i, storage_type, raw_value, value_text, caption))

        # Guard against IronPython .NET null-to-string artefacts
        if not value_text or value_text == "None":
            value_text = ""
        # Fallback: match by entry caption (the name Revit shows in the color scheme dialog)
        if not value_text:
            value_text = caption
        if not value_text:
            continue

        rgb = color_map.get(value_text)
        if rgb is None:
            skipped.append(value_text)
            continue

        if hasattr(entry, "Color"):
            try:
                entry.Color = DB.Color(int(rgb[0]), int(rgb[1]), int(rgb[2]))
                updated += 1
            except Exception as ex:
                _log(log, "Failed to set color for '{}': {}".format(value_text, str(ex)))

    if updated == 0:
        return False, (
            "No scheme entries were matched to Excel Key Names.\n"
            "Verify that the Value column selected in the mapping dialog contains the same "
            "values as the color scheme (e.g. RES_AMENITY_INDOOR_GF).\n"
            "Unmatched scheme values: {}".format(", ".join(skipped[:10]) if skipped else "<none>")
        )

    set_entries = getattr(target, "SetEntries", None)
    if not callable(set_entries):
        return False, "Target scheme API does not support SetEntries."

    set_entries(csu._to_entry_collection(entries))
    csu._regenerate_scheme_document(target)
    refreshed = list(target.GetEntries())
    if refreshed:
        # After SetEntries + Regenerate, Revit may blend/reset colors for "In Use"
        # entries and clears their ElementId values, so value/caption matching is
        # not reliable here. Copy colors directly by index position instead.
        # Do NOT regenerate after this second SetEntries — a second Regenerate
        # causes Revit to blend the colors again.
        if len(refreshed) == len(entries):
            for src, tgt in zip(entries, refreshed):
                csu._copy_entry_visuals(src, tgt)
        else:
            csu._patch_entry_colors(refreshed, entries, log=log, stage="color-override-post-set")
        set_entries(csu._to_entry_collection(refreshed))
    _log(log, "Color override: {} entries updated, {} skipped (not in Excel — colors left unchanged).".format(
        updated, len(skipped)))
    return True, None


def build_entry_from_payload(item):
    storage_type = parse_storage_type(item.get("storage_type"))
    if storage_type is None:
        raise Exception("Unsupported storage type '{}'.".format(item.get("storage_type", "")))
    entry = DB.ColorFillSchemeEntry(storage_type)
    raw_value = item.get("value")
    if storage_type == DB.StorageType.String:
        value = "" if raw_value is None else str(raw_value)
    elif storage_type == DB.StorageType.Integer:
        value = _to_int(raw_value, 0)
    elif storage_type == DB.StorageType.Double:
        value = _to_float(raw_value, 0.0)
    elif storage_type == DB.StorageType.ElementId:
        value = DB.ElementId(_to_int(raw_value, -1))
    else:
        value = raw_value
    if not csu._entry_set_value(entry, storage_type, value):
        raise Exception("Failed to set value '{}' on entry.".format(raw_value))

    r = item.get("color_r")
    g = item.get("color_g")
    b = item.get("color_b")
    if r is not None and g is not None and b is not None and hasattr(entry, "Color"):
        try:
            entry.Color = DB.Color(int(r), int(g), int(b))
        except Exception:
            pass
    fill_pattern_id = item.get("fill_pattern_id")
    if fill_pattern_id is not None and hasattr(entry, "FillPatternId"):
        try:
            entry.FillPatternId = DB.ElementId(int(fill_pattern_id))
        except Exception:
            pass
    for attr in ("IsVisible", "Visible"):
        if hasattr(entry, attr):
            try:
                setattr(entry, attr, bool(item.get("is_visible", True)))
            except Exception:
                pass
    caption = item.get("caption", "")
    if caption:
        try:
            setter = getattr(entry, "SetCaption", None)
            if callable(setter):
                setter(caption)
            elif hasattr(entry, "Caption"):
                entry.Caption = caption
        except Exception:
            pass
    return entry


def _payload_item_match_keys(item):
    keys = []
    for candidate in (item.get("caption", ""), item.get("value", "")):
        for alias in _match_aliases(candidate):
            if alias not in keys:
                keys.append(alias)
    return keys


def _existing_entry_match_keys(entry, doc=None):
    keys = []
    storage_type = getattr(entry, "StorageType", None)
    raw_value = csu._entry_get_value(entry, storage_type)
    candidates = [
        csu._entry_caption(entry),
        entry_value_to_text(entry, storage_type, doc=doc),
        csu._format_entry_value(entry),
    ]
    if storage_type == DB.StorageType.ElementId:
        candidates.append(str(elem_id_int(raw_value)))
    for candidate in candidates:
        for alias in _match_aliases(candidate):
            if alias not in keys:
                keys.append(alias)
    return keys


def merge_payload_into_scheme(target, payload, log=None):
    for attr, key in (("Title", "title"), ("IsByRange", "is_by_range"), ("IsByValue", "is_by_value"), ("IsByPercentage", "is_by_percentage")):
        if hasattr(target, attr) and key in payload:
            try:
                setattr(target, attr, payload.get(key))
            except Exception:
                pass

    try:
        entries = list(target.GetEntries())
    except Exception:
        return False, "Could not read existing entries from the target scheme."

    doc = getattr(target, "Document", None)
    lookup = {}
    for idx, entry in enumerate(entries):
        for key in _existing_entry_match_keys(entry, doc=doc):
            if key and key not in lookup:
                lookup[key] = idx

    updated = 0
    added_entries = []
    desired_entries = list(entries)

    for item in payload.get("entries", []):
        match_index = None
        for key in _payload_item_match_keys(item):
            if key in lookup:
                match_index = lookup.get(key)
                break

        if match_index is not None:
            entry = desired_entries[match_index]
            r = item.get("color_r")
            g = item.get("color_g")
            b = item.get("color_b")
            if r is not None and g is not None and b is not None and hasattr(entry, "Color"):
                try:
                    entry.Color = DB.Color(int(r), int(g), int(b))
                    updated += 1
                except Exception as ex:
                    _log(log, "Failed to set merged color for '{}': {}".format(item.get("caption", "") or item.get("value", ""), str(ex)))
            continue

        try:
            new_entry = build_entry_from_payload(item)
            desired_entries.append(new_entry)
            added_entries.append(new_entry)
        except Exception as ex:
            return False, "Failed to build a new entry for '{}': {}".format(item.get("caption", "") or item.get("value", ""), str(ex))

    if not updated and not added_entries:
        return False, "No matching entries were updated and no new entries were added."

    set_entries = getattr(target, "SetEntries", None)
    if not callable(set_entries):
        return False, "Target scheme API does not support SetEntries."

    set_entries(csu._to_entry_collection(desired_entries))
    csu._regenerate_scheme_document(target)
    refreshed = list(target.GetEntries())
    if refreshed:
        if len(refreshed) == len(desired_entries):
            for src, tgt in zip(desired_entries, refreshed):
                csu._copy_entry_visuals(src, tgt)
        else:
            csu._patch_entry_colors(refreshed, desired_entries, log=log, stage="merge-post-set")
        set_entries(csu._to_entry_collection(refreshed))
        csu._regenerate_scheme_document(target)

    _log(log, "Merged payload into scheme: {} existing entries recolored, {} new entries added.".format(updated, len(added_entries)))
    return True, None


def apply_payload_to_scheme(target, payload, log=None):
    for attr, key in (("Title", "title"), ("IsByRange", "is_by_range"), ("IsByValue", "is_by_value"), ("IsByPercentage", "is_by_percentage")):
        if hasattr(target, attr) and key in payload:
            try:
                setattr(target, attr, payload.get(key))
            except Exception:
                pass

    entries_to_set = [build_entry_from_payload(item) for item in payload.get("entries", [])]
    set_entries = getattr(target, "SetEntries", None)
    if callable(set_entries):
        set_entries(csu._to_entry_collection(entries_to_set))
        csu._regenerate_scheme_document(target)
        refreshed = list(target.GetEntries())
        if refreshed:
            csu._patch_entry_colors(refreshed, entries_to_set, log=log, stage="import-post-set")
            set_entries(csu._to_entry_collection(refreshed))
            csu._regenerate_scheme_document(target)
        return True, None

    add_entry = getattr(target, "AddEntry", None)
    if not callable(add_entry):
        return False, "Target scheme API does not support SetEntries/AddEntry."
    failures = []
    for entry in entries_to_set:
        try:
            add_entry(entry)
        except Exception as ex:
            failures.append(str(ex))
    if failures:
        return False, "Failed to add one or more entries: {}".format(" | ".join(failures[:5]))
    csu._regenerate_scheme_document(target)
    return True, None


def get_scheme_definition_snapshot(doc, scheme):
    entry_storage_types = []
    try:
        entries = list(scheme.GetEntries())
    except Exception:
        entries = []
    for entry in entries:
        entry_storage_types.append(storage_type_token(getattr(entry, "StorageType", None)))
    unique_storage_types = sorted(set([s for s in entry_storage_types if s]))
    return {
        "scheme_name": getattr(scheme, "Name", "") or "Color Scheme",
        "category_id": elem_id_int(getattr(scheme, "CategoryId", None)),
        "category_name": category_name(doc, getattr(scheme, "CategoryId", None)),
        "area_scheme_name": scheme_area_scheme_name(doc, scheme),
        "title": getattr(scheme, "Title", "") or "",
        "parameter_id": _scheme_parameter_id_int(scheme),
        "parameter_name": _scheme_parameter_label(doc, scheme),
        "is_by_range": bool(getattr(scheme, "IsByRange", False)),
        "is_by_value": bool(getattr(scheme, "IsByValue", False)),
        "is_by_percentage": bool(getattr(scheme, "IsByPercentage", False)),
        "entry_storage_types": unique_storage_types,
    }


def validate_payload_against_scheme(doc, payload, target_scheme):
    target = get_scheme_definition_snapshot(doc, target_scheme)
    payload_storage_types = sorted(set([
        str(item.get("storage_type", "")).strip()
        for item in payload.get("entries", [])
        if str(item.get("storage_type", "")).strip()
    ]))
    problems = []
    if payload_storage_types and target.get("entry_storage_types") and payload_storage_types != target.get("entry_storage_types"):
        problems.append(
            "Entry storage type mismatch. Workbook={} Target={}".format(
                ", ".join(payload_storage_types),
                ", ".join(target.get("entry_storage_types") or []),
            )
        )
    payload_pid = payload.get("parameter_id")
    target_pid = target.get("parameter_id")
    if payload_pid is not None and target_pid is not None and payload_pid != target_pid:
        problems.append(
            "Scheme parameter mismatch. Workbook={} ({}) Target={} ({})".format(
                payload_pid,
                payload.get("parameter_name", ""),
                target_pid,
                target.get("parameter_name", ""),
            )
        )
    for key, label in (
        ("is_by_range", "By Range"),
        ("is_by_value", "By Value"),
        ("is_by_percentage", "By Percentage"),
    ):
        if key in payload and bool(payload.get(key)) != bool(target.get(key)):
            problems.append(
                "Scheme mode mismatch for {}. Workbook={} Target={}".format(
                    label,
                    bool(payload.get(key)),
                    bool(target.get(key)),
                )
            )
    ok = len(problems) == 0
    return ok, problems, target


def unique_scheme_name_in_scope(doc, scope_seed_scheme, base_name):
    base = (base_name or "Color Scheme").strip()
    if not base:
        base = "Color Scheme"
    all_schemes = csu.collect_color_fill_schemes(doc)
    if csu.find_scheme_in_scope_by_name(all_schemes, scope_seed_scheme, base) is None:
        return base
    idx = 1
    while idx < 1000:
        candidate = "{} ({})".format(base, idx)
        if csu.find_scheme_in_scope_by_name(all_schemes, scope_seed_scheme, candidate) is None:
            return candidate
        idx += 1
    return "{} ({})".format(base, "Copy")


def build_import_target_choices(doc, schemes):
    scope_map = {}
    for scheme in schemes:
        key = (elem_id_int(getattr(scheme, "CategoryId", None)), elem_id_int(scheme_area_scheme_id(scheme)))
        if key not in scope_map:
            scope_map[key] = scheme
    create_choices = [{
        "mode": "create",
        "label": "Create New In {}".format(scheme_scope_label(doc, seed)),
        "scheme": seed,
    } for seed in scope_map.values()]
    overwrite_choices = [{
        "mode": "overwrite",
        "label": "Overwrite {}".format(scheme_display_name(doc, scheme)),
        "scheme": scheme,
    } for scheme in schemes]
    create_choices.sort(key=lambda x: x["label"].lower())
    overwrite_choices.sort(key=lambda x: x["label"].lower())
    return create_choices + overwrite_choices


def choose_default_import_target_index(doc, payload, choices):
    payload_area = (payload.get("area_scheme_name", "") or "").strip().lower()
    payload_category = (payload.get("category_name", "") or "").strip().lower()
    payload_scheme = (payload.get("scheme_name", "") or "").strip().lower()
    for idx, choice in enumerate(choices):
        scheme = choice.get("scheme")
        if not scheme:
            continue
        area_name = scheme_area_scheme_name(doc, scheme).strip().lower()
        cat_name = category_name(doc, getattr(scheme, "CategoryId", None)).strip().lower()
        if payload_area:
            if area_name != payload_area:
                continue
        elif payload_category and cat_name != payload_category:
            continue
        if choice.get("mode") == "overwrite":
            name = (getattr(scheme, "Name", "") or "").strip().lower()
            if name == payload_scheme:
                return idx
        else:
            return idx
    return 0
