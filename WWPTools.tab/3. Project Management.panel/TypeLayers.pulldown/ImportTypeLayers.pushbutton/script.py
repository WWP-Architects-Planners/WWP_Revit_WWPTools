#! python3
from __future__ import annotations

import os
import re
import sys

from pyrevit import DB, revit
from System.Windows.Forms import DialogResult, MessageBox, OpenFileDialog


def add_lib_path():
    script_dir = os.path.dirname(__file__)
    lib_path = os.path.abspath(os.path.join(script_dir, "..", "..", "..", "..", "lib"))
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def load_uiutils():
    add_lib_path()
    import WWP_uiUtils as ui
    return ui


def element_id_value(elem_id):
    if elem_id is None:
        return -1
    if hasattr(elem_id, "IntegerValue"):
        return elem_id.IntegerValue
    if hasattr(elem_id, "Value"):
        return elem_id.Value
    try:
        return int(elem_id)
    except Exception:
        return -1


def pick_excel_path(doc):
    dlg = OpenFileDialog()
    dlg.Title = "Select Type Layers Excel Workbook"
    dlg.Filter = "Excel Files (*.xlsx)|*.xlsx|All Files (*.*)|*.*"
    dlg.Multiselect = False
    try:
        if doc.PathName:
            dlg.InitialDirectory = os.path.dirname(doc.PathName)
    except Exception:
        pass
    if dlg.ShowDialog() != DialogResult.OK:
        return None
    return dlg.FileName


def get_length_unit_id(doc):
    try:
        units = doc.GetUnits()
        return units.GetFormatOptions(DB.SpecTypeId.Length).GetUnitTypeId()
    except Exception:
        return None


def to_internal_length(doc, value):
    unit_id = get_length_unit_id(doc)
    if unit_id:
        try:
            return DB.UnitUtils.ConvertToInternalUnits(value, unit_id)
        except Exception:
            pass
    try:
        return DB.UnitUtils.ConvertToInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def to_internal_length_mm(value):
    try:
        return DB.UnitUtils.ConvertToInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def parse_display_length(doc, value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        units = doc.GetUnits()
        return DB.UnitFormatUtils.Parse(units, DB.SpecTypeId.Length, text)
    except Exception:
        pass
    cleaned = text.replace(",", "")
    match = re.search(r"-?\d+(\.\d+)?", cleaned)
    if not match:
        return None
    try:
        return to_internal_length(doc, float(match.group(0)))
    except Exception:
        return None


def parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in ("true", "yes", "1")


def parse_function(value):
    if value is None or value == "":
        return DB.MaterialFunctionAssignment.Structure
    if isinstance(value, (int, float)):
        try:
            return DB.MaterialFunctionAssignment(int(value))
        except Exception:
            pass
    try:
        int_val = int(str(value).strip())
        return DB.MaterialFunctionAssignment(int_val)
    except Exception:
        pass
    name = str(value).strip()
    if not name:
        return DB.MaterialFunctionAssignment.Structure
    if "." in name:
        name = name.split(".")[-1]
    try:
        return getattr(DB.MaterialFunctionAssignment, name)
    except Exception:
        return DB.MaterialFunctionAssignment.Structure


def find_material_id(doc, mat_id_value, mat_name):
    if mat_id_value and mat_id_value > 0:
        try:
            mat = doc.GetElement(DB.ElementId(mat_id_value))
            if mat:
                return mat.Id
        except Exception:
            pass
    if mat_name:
        try:
            mat = next(
                (m for m in DB.FilteredElementCollector(doc).OfClass(DB.Material) if m.Name == mat_name),
                None,
            )
            if mat:
                return mat.Id
        except Exception:
            pass
    return DB.ElementId.InvalidElementId


def read_workbook(path):
    add_lib_path()
    try:
        import openpyxl
    except Exception as exc:
        MessageBox.Show("openpyxl is not available.\n{}".format(exc), "Import Type Layers")
        return None
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        MessageBox.Show("Failed to open workbook.\n{}".format(exc), "Import Type Layers")
        return None


def extract_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(val).strip() if val is not None else "" for val in rows[0]]
    aliases = {
        "category": "Category",
        "typename": "TypeName",
        "type name": "TypeName",
        "sourcetypename": "SourceTypeName",
        "source type name": "SourceTypeName",
        "originaltypename": "SourceTypeName",
        "original type name": "SourceTypeName",
        "familyname": "FamilyName",
        "family name": "FamilyName",
        "typeid": "TypeId",
        "type id": "TypeId",
        "typeuniqueid": "TypeUniqueId",
        "type unique id": "TypeUniqueId",
        "layerindex": "LayerIndex",
        "layer index": "LayerIndex",
        "layerfunction": "LayerFunction",
        "layer function": "LayerFunction",
        "layerfunctionid": "LayerFunctionId",
        "layer function id": "LayerFunctionId",
        "materialname": "MaterialName",
        "material name": "MaterialName",
        "materialid": "MaterialId",
        "material id": "MaterialId",
        "thicknessinternal": "ThicknessInternal",
        "thickness internal": "ThicknessInternal",
        "thicknessmetric": "ThicknessMetric",
        "thickness metric": "ThicknessMetric",
        "thicknessdisplay": "ThicknessDisplay",
        "thickness display": "ThicknessDisplay",
        "thicknessproject": "ThicknessProject",
        "thickness project": "ThicknessProject",
        "isvariable": "IsVariable",
        "is variable": "IsVariable",
        "hascompoundstructure": "HasCompoundStructure",
        "has compound structure": "HasCompoundStructure",
    }
    header_map = {}
    for idx, name in enumerate(headers):
        if not name:
            continue
        key = name.strip()
        lower = key.lower()
        canonical = aliases.get(lower, key)
        header_map[canonical] = idx
    data = []
    for raw in rows[1:]:
        if raw is None:
            continue
        if all(cell in (None, "") for cell in raw):
            continue
        row = {}
        for key, idx in header_map.items():
            if idx < len(raw):
                row[key] = raw[idx]
            else:
                row[key] = None
        data.append(row)
    return data


def map_category(sheet_name):
    name = sheet_name.strip().lower()
    if name in ("wall types", "walls", "wall"):
        return ("Wall Types", DB.WallType)
    if name in ("floor types", "floors", "floor"):
        return ("Floor Types", DB.FloorType)
    if name in ("roof types", "roofs", "roof"):
        return ("Roof Types", DB.RoofType)
    return None


def collect_types_by_category(doc, type_class):
    items = list(DB.FilteredElementCollector(doc).OfClass(type_class).WhereElementIsElementType())
    by_unique = {}
    by_id = {}
    by_name = {}
    for item in items:
        try:
            by_unique[item.UniqueId] = item
        except Exception:
            pass
        by_id[element_id_value(item.Id)] = item
        by_name[item.Name] = item
    return items, by_unique, by_id, by_name


def build_layer(layer_row, doc):
    width = None
    metric = parse_float(layer_row.get("ThicknessMetric"))
    if metric is not None:
        width = to_internal_length_mm(metric)
    if width is None:
        width = parse_float(layer_row.get("ThicknessInternal"))
    if width is None:
        width = parse_float(layer_row.get("ThicknessProject"))
        if width is not None:
            width = to_internal_length(doc, width)
    if width is None:
        width = parse_display_length(doc, layer_row.get("ThicknessDisplay"))
    if width is None:
        return None
    if width <= 0:
        return None

    func_value = layer_row.get("LayerFunctionId")
    if func_value in (None, ""):
        func_value = layer_row.get("LayerFunction")
    func = parse_function(func_value)

    mat_id_val = parse_float(layer_row.get("MaterialId"))
    mat_id_val = int(mat_id_val) if mat_id_val is not None else None
    mat_name = layer_row.get("MaterialName")
    mat_id = find_material_id(doc, mat_id_val, mat_name)

    layer = DB.CompoundStructureLayer(width, func, mat_id)
    if "IsVariable" in layer_row:
        try:
            layer.IsVariable = parse_bool(layer_row.get("IsVariable"))
        except Exception:
            pass
    return layer


def update_type_layers(doc, elem_type, layers):
    if not layers:
        return False, "No valid layers found."

    compound = None
    try:
        compound = elem_type.GetCompoundStructure()
    except Exception:
        compound = None

    if compound is not None:
        try:
            compound.SetLayers(layers)
        except Exception:
            compound = None

    if compound is None:
        try:
            compound = DB.CompoundStructure.CreateSimpleCompoundStructure(layers)
        except Exception:
            compound = None

    if compound is None:
        return False, "Compound structure creation failed."

    try:
        elem_type.SetCompoundStructure(compound)
    except Exception as exc:
        return False, str(exc)
    return True, ""


def rename_type(elem_type, new_name):
    if not new_name:
        return
    try:
        if elem_type.Name != new_name:
            elem_type.Name = new_name
    except Exception:
        pass


def get_existing_layers(elem_type):
    try:
        compound = elem_type.GetCompoundStructure()
    except Exception:
        compound = None
    if compound is None:
        return []
    try:
        return list(compound.GetLayers())
    except Exception:
        return []


def layers_thickness_changed(existing_layers, new_layers, tol=1e-4):
    if len(existing_layers) != len(new_layers):
        return True
    for old, new in zip(existing_layers, new_layers):
        try:
            if abs(old.Width - new.Width) > tol:
                return True
        except Exception:
            return True
    return False


def group_rows(rows):
    grouped = {}
    for row in rows:
        unique_id = row.get("TypeUniqueId")
        type_id = row.get("TypeId")
        type_name = row.get("TypeName")
        source_name = row.get("SourceTypeName")
        key = unique_id or type_id or source_name or type_name
        if not key:
            continue
        group = grouped.get(key)
        if group is None:
            group = {
                "TypeUniqueId": unique_id,
                "TypeId": type_id,
                "TypeName": type_name,
                "SourceTypeName": source_name,
                "FamilyName": row.get("FamilyName"),
                "HasCompoundStructure": parse_bool(row.get("HasCompoundStructure")),
                "Layers": [],
            }
            grouped[key] = group
        group["Layers"].append(row)
    return grouped


def main():
    doc = revit.doc
    ui = load_uiutils()
    path = pick_excel_path(doc)
    if not path:
        return

    workbook = read_workbook(path)
    if workbook is None:
        return

    category_data = []
    for sheet_name in workbook.sheetnames:
        mapping = map_category(sheet_name)
        if not mapping:
            continue
        label, type_class = mapping
        rows = extract_rows(workbook[sheet_name])
        if rows:
            category_data.append((label, type_class, rows))

    if not category_data:
        MessageBox.Show("No matching category sheets found.", "Import Type Layers")
        return

    updated = 0
    created = 0
    skipped = 0
    errors = []
    missing_types_summary = []
    overwrite_skipped = 0

    with revit.Transaction("Import Type Layers"):
        for label, type_class, rows in category_data:
            types, by_unique, by_id, by_name = collect_types_by_category(doc, type_class)
            grouped = group_rows(rows)
            excel_ids = set()

            for key, group in grouped.items():
                unique_id = group.get("TypeUniqueId")
                type_id = group.get("TypeId")
                type_name = group.get("TypeName")
                source_name = group.get("SourceTypeName")
                elem_type = None

                if unique_id and unique_id in by_unique:
                    elem_type = by_unique[unique_id]
                else:
                    type_id_val = parse_float(type_id)
                    if type_id_val is not None and int(type_id_val) in by_id:
                        elem_type = by_id[int(type_id_val)]
                if elem_type is None and source_name and source_name in by_name:
                    elem_type = by_name[source_name]
                if elem_type is None and type_name and type_name in by_name:
                    elem_type = by_name[type_name]

                if elem_type is None:
                    base_type = types[0] if types else None
                    if base_type and type_name:
                        try:
                            new_id = base_type.Duplicate(type_name)
                            elem_type = doc.GetElement(new_id)
                            created += 1
                        except Exception as exc:
                            errors.append("{}: create failed ({})".format(type_name, exc))
                            skipped += 1
                            continue
                    else:
                        skipped += 1
                        continue

                excel_ids.add(element_id_value(elem_type.Id))
                existing_name = elem_type.Name

                sorted_rows = sorted(
                    group["Layers"],
                    key=lambda r: parse_float(r.get("LayerIndex")) or 0,
                )
                layers = []
                for row in sorted_rows:
                    layer = build_layer(row, doc)
                    if layer:
                        layers.append(layer)

                name_changed = bool(type_name and existing_name != type_name)

                if not layers:
                    if not group.get("HasCompoundStructure", True):
                        skipped += 1
                        continue
                    if name_changed:
                        prompt = (
                            "Type '{}' has a name change ({} -> {}).\n"
                            "No valid layer data was found.\n\n"
                            "Rename anyway?"
                        ).format(existing_name, existing_name, type_name)
                        if ui.uiUtils_confirm(prompt, title="Import Type Layers"):
                            rename_type(elem_type, type_name)
                        else:
                            overwrite_skipped += 1
                            skipped += 1
                            continue
                    skipped += 1
                    errors.append("{}: No valid layers found.".format(type_name or source_name or existing_name))
                    continue

                existing_layers = get_existing_layers(elem_type)
                thickness_changed = layers_thickness_changed(existing_layers, layers)
                if name_changed or thickness_changed:
                    prompt_parts = []
                    if name_changed:
                        prompt_parts.append("name ({} -> {})".format(existing_name, type_name))
                    if thickness_changed:
                        prompt_parts.append("layer thickness")
                    prompt = (
                        "Type '{}' has changes in {}.\n\n"
                        "Overwrite from Excel?"
                    ).format(elem_type.Name, " and ".join(prompt_parts))
                    if not ui.uiUtils_confirm(prompt, title="Import Type Layers"):
                        overwrite_skipped += 1
                        skipped += 1
                        continue

                if name_changed:
                    rename_type(elem_type, type_name)

                ok, err = update_type_layers(doc, elem_type, layers)
                if ok:
                    updated += 1
                else:
                    skipped += 1
                    if err:
                        errors.append("{}: {}".format(elem_type.Name, err))

            missing = [t for t in types if element_id_value(t.Id) not in excel_ids]
            if missing:
                missing_types_summary.append(
                    (label, [t.Name for t in missing])
                )

    if missing_types_summary:
        confirm = ui.uiUtils_confirm(
            "Some Revit types are not listed in the Excel file.\n"
            "Do you want to delete those types?\n\n"
            "This cannot be undone.",
            title="Import Type Layers",
        )
        if confirm:
            with revit.Transaction("Delete Types Not In Excel"):
                for label, names in missing_types_summary:
                    items, _, _, by_name = collect_types_by_category(
                        doc,
                        DB.WallType if label == "Wall Types" else
                        DB.FloorType if label == "Floor Types" else
                        DB.RoofType
                    )
                    for name in names:
                        elem = by_name.get(name)
                        if elem:
                            try:
                                doc.Delete(elem.Id)
                            except Exception:
                                pass

    summary = "Updated: {}\nCreated: {}\nSkipped: {}".format(updated, created, skipped)
    if overwrite_skipped:
        summary += "\nSkipped (user chose not to overwrite): {}".format(overwrite_skipped)
    if errors:
        summary += "\n\nErrors:\n" + "\n".join(errors[:10])
        if len(errors) > 10:
            summary += "\n... ({}) more".format(len(errors) - 10)
    MessageBox.Show(summary, "Import Type Layers")


if __name__ == "__main__":
    main()
