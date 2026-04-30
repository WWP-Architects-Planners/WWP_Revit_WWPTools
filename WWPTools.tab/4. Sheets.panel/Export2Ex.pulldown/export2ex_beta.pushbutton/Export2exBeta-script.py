#! python3

import json
import os
import re
import sys
import time
import traceback

import clr
from System import String
from System import Object
from System.Collections.Generic import List
from System.IO import File
from System.Windows.Media import Brushes
from System.Windows.Controls import ListBoxItem

from pyrevit import DB
from WWP_settings import get_tool_settings
from WWP_versioning import apply_window_title


CONFIG_LAST_EXCEL_PATH = "last_excel_path"
CONFIG_LAST_MODE = "last_mode"
CONFIG_LAST_SCHEDULE_ID = "last_schedule_id"
CONFIG_LAST_CATEGORY_ID = "last_category_id"
CONFIG_LAST_PARAM_NAMES = "last_param_names"
CONFIG_LAST_SHEET_NAME = "last_sheet_name"
PARAM_SAVED_SETS = "! P_STATS_Export"
SAVED_SET_NAMESPACE = "export2ex_beta"
LOG_FILE_NAME = "Export2ExBeta.log"
ALLOWED_EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
MODE_FROM_SCHEDULE = "schedule"
MODE_BY_CATEGORY = "category"
EMBEDDED_EXPORT_DIALOG_XAML = r'''<Window xmlns="http://schemas.microsoft.com/winfx/2006/xaml/presentation"
        xmlns:x="http://schemas.microsoft.com/winfx/2006/xaml"
        Title="Export2Ex Beta"
        Height="820"
        Width="1100"
        MinHeight="640"
        MinWidth="900"
        WindowStartupLocation="CenterScreen"
        ResizeMode="CanResizeWithGrip"
        FontFamily="Segoe UI"
        FontSize="14"
        Background="#F3F5F7">
    <Window.Resources>
        <Style x:Key="PrimaryButtonStyle" TargetType="Button">
            <Setter Property="Background" Value="#3F9AD9"/>
            <Setter Property="Foreground" Value="White"/>
            <Setter Property="BorderBrush" Value="#3F9AD9"/>
            <Setter Property="BorderThickness" Value="1"/>
            <Setter Property="Height" Value="30"/>
            <Setter Property="Padding" Value="18,0"/>
            <Setter Property="FontWeight" Value="SemiBold"/>
        </Style>
        <Style x:Key="SecondaryButtonStyle" TargetType="Button" BasedOn="{StaticResource PrimaryButtonStyle}">
            <Setter Property="Background" Value="#FFFFFF"/>
            <Setter Property="Foreground" Value="#1F2937"/>
            <Setter Property="BorderBrush" Value="#CBD5E1"/>
        </Style>
        <Style TargetType="TextBox">
            <Setter Property="Background" Value="#FFFFFF"/>
            <Setter Property="BorderBrush" Value="#CBD5E1"/>
            <Setter Property="BorderThickness" Value="1"/>
            <Setter Property="Padding" Value="6,4"/>
            <Setter Property="MinHeight" Value="28"/>
        </Style>
        <Style TargetType="ListBox">
            <Setter Property="Background" Value="#FFFFFF"/>
            <Setter Property="BorderBrush" Value="#CBD5E1"/>
            <Setter Property="BorderThickness" Value="1"/>
            <Setter Property="Padding" Value="2"/>
        </Style>
    </Window.Resources>
    <Grid Margin="16">
        <Border Background="#FFFFFF"
                BorderBrush="#D7DEE6"
                BorderThickness="1"
                Padding="20">
            <Grid>
                <Grid.RowDefinitions>
                    <RowDefinition Height="Auto"/>
                    <RowDefinition Height="*"/>
                    <RowDefinition Height="Auto"/>
                    <RowDefinition Height="Auto"/>
                    <RowDefinition Height="Auto"/>
                </Grid.RowDefinitions>

                <!-- Saved sets -->
                <Grid Grid.Row="0" Margin="0,0,0,8">
                    <Grid.ColumnDefinitions>
                        <ColumnDefinition Width="Auto"/>
                        <ColumnDefinition Width="*"/>
                        <ColumnDefinition Width="Auto"/>
                        <ColumnDefinition Width="Auto"/>
                        <ColumnDefinition Width="Auto"/>
                    </Grid.ColumnDefinitions>
                    <TextBlock Grid.Column="0"
                               Text="Saved set:"
                               Foreground="#374151"
                               VerticalAlignment="Center"
                               FontSize="12"
                               Margin="0,0,8,0"/>
                    <ComboBox Name="SavedSetBox"
                              Grid.Column="1"
                              IsEditable="True"
                              Margin="0,0,8,0"/>
                    <Button Name="LoadSetButton"
                            Grid.Column="2"
                            Content="Load"
                            Width="64"
                            Margin="0,0,6,0"
                            Style="{StaticResource SecondaryButtonStyle}"/>
                    <Button Name="SaveSetButton"
                            Grid.Column="3"
                            Content="Save"
                            Width="64"
                            Margin="0,0,6,0"
                            Style="{StaticResource SecondaryButtonStyle}"/>
                    <Button Name="DeleteSetButton"
                            Grid.Column="4"
                            Content="Delete"
                            Width="64"
                            Style="{StaticResource SecondaryButtonStyle}"/>
                </Grid>

                <!-- Main 3-column area -->
                <Grid Grid.Row="1">
                    <Grid.ColumnDefinitions>
                        <ColumnDefinition Width="340"/>
                        <ColumnDefinition Width="130"/>
                        <ColumnDefinition Width="*"/>
                    </Grid.ColumnDefinitions>

                    <!-- LEFT: Available properties (stretches) + source selector at bottom -->
                    <Grid Grid.Column="0">
                        <Grid.RowDefinitions>
                            <RowDefinition Height="Auto"/>
                            <RowDefinition Height="*"/>
                            <RowDefinition Height="Auto"/>
                        </Grid.RowDefinitions>

                        <TextBlock Grid.Row="0"
                                   Text="Available properties:"
                                   Foreground="#374151"
                                   FontWeight="SemiBold"
                                   Margin="0,0,0,4"/>

                        <ListBox Name="ParameterList"
                                 Grid.Row="1"
                                 SelectionMode="Extended"
                                 MinHeight="80"/>

                        <Border Grid.Row="2"
                                BorderBrush="#E2E8F0"
                                BorderThickness="0,1,0,0"
                                Margin="0,10,0,0"
                                Padding="0,8,0,0">
                            <StackPanel>
                                <DockPanel Margin="0,0,0,4">
                                    <TextBlock DockPanel.Dock="Left"
                                               Text="Select from:"
                                               Foreground="#6B7280"
                                               FontSize="12"
                                               VerticalAlignment="Center"
                                               Margin="0,0,6,0"/>
                                    <ComboBox Name="ModeBox" SelectedIndex="0">
                                        <ComboBoxItem Content="From Schedule"/>
                                        <ComboBoxItem Content="By Category"/>
                                    </ComboBox>
                                </DockPanel>
                                <TextBlock Name="SourceLabel"
                                           Text="Search schedules:"
                                           Foreground="#6B7280"
                                           FontSize="12"
                                           Margin="0,0,0,2"/>
                                <TextBox Name="SourceSearchBox" Margin="0,0,0,4"/>
                                <TextBlock Name="SourceListLabel"
                                           Text="Schedules"
                                           Foreground="#6B7280"
                                           FontSize="12"
                                           Margin="0,0,0,2"/>
                                <ListBox Name="SourceList"
                                         Height="140"
                                         SelectionMode="Single"/>
                            </StackPanel>
                        </Border>
                    </Grid>

                    <!-- MIDDLE: Add/Remove + Display filters + Search -->
                    <StackPanel Grid.Column="1"
                                Margin="10,28,10,0"
                                VerticalAlignment="Top">
                        <Button Name="AddParameterButton"
                                Content="Add -->"
                                Margin="0,0,0,6"
                                Style="{StaticResource SecondaryButtonStyle}"/>
                        <Button Name="RemoveParameterButton"
                                Content="&lt;-- Remove"
                                Style="{StaticResource SecondaryButtonStyle}"/>

                        <Rectangle Height="1" Fill="#E2E8F0" Margin="0,14,0,10"/>

                        <TextBlock Text="Display filters"
                                   Foreground="#6B7280"
                                   FontSize="12"
                                   FontWeight="SemiBold"
                                   Margin="0,0,0,6"/>
                        <CheckBox Name="ShowReadOnlyFilter"
                                  Content="Read-only"
                                  IsChecked="True"
                                  FontSize="12"
                                  Foreground="#374151"/>

                        <Rectangle Height="1" Fill="#E2E8F0" Margin="0,12,0,10"/>

                        <TextBlock Text="Search:"
                                   Foreground="#6B7280"
                                   FontSize="12"
                                   Margin="0,0,0,4"/>
                        <TextBox Name="ParameterSearchBox"/>
                    </StackPanel>

                    <!-- RIGHT: Linked properties (stretches) + Move buttons at bottom -->
                    <Grid Grid.Column="2">
                        <Grid.RowDefinitions>
                            <RowDefinition Height="Auto"/>
                            <RowDefinition Height="*"/>
                            <RowDefinition Height="Auto"/>
                        </Grid.RowDefinitions>

                        <TextBlock Grid.Row="0"
                                   Text="Linked properties:"
                                   Foreground="#374151"
                                   FontWeight="SemiBold"
                                   Margin="0,0,0,4"/>

                        <ListBox Name="SelectedParameterList"
                                 Grid.Row="1"
                                 SelectionMode="Extended"
                                 MinHeight="80"/>

                        <StackPanel Grid.Row="2"
                                    Orientation="Horizontal"
                                    Margin="0,8,0,0">
                            <Button Name="MoveParameterUpButton"
                                    Content="Move Up"
                                    Width="110"
                                    Style="{StaticResource SecondaryButtonStyle}"/>
                            <Button Name="MoveParameterDownButton"
                                    Content="Move Down"
                                    Width="110"
                                    Margin="8,0,0,0"
                                    Style="{StaticResource SecondaryButtonStyle}"/>
                        </StackPanel>
                    </Grid>
                </Grid>

                <!-- Sheet name -->
                <Grid Grid.Row="2" Margin="0,10,0,0">
                    <Grid.ColumnDefinitions>
                        <ColumnDefinition Width="Auto"/>
                        <ColumnDefinition Width="*"/>
                    </Grid.ColumnDefinitions>
                    <TextBlock Grid.Column="0"
                               Text="Sheet name:"
                               Foreground="#374151"
                               VerticalAlignment="Center"
                               FontSize="12"
                               Margin="0,0,8,0"/>
                    <TextBox Name="SheetNameBox" Grid.Column="1"/>
                </Grid>

                <!-- Excel path + Browse -->
                <Grid Grid.Row="3" Margin="0,8,0,0">
                    <Grid.ColumnDefinitions>
                        <ColumnDefinition Width="*"/>
                        <ColumnDefinition Width="90"/>
                    </Grid.ColumnDefinitions>
                    <TextBox Name="ExcelPath" Margin="0,0,8,0"/>
                    <Button Name="BrowseExcel"
                            Grid.Column="1"
                            Content="Browse"
                            Style="{StaticResource SecondaryButtonStyle}"/>
                </Grid>

                <!-- Footer: logo + Export + Cancel -->
                <DockPanel Grid.Row="4" Margin="0,16,0,0">
                    <Image Name="LogoImage"
                           DockPanel.Dock="Left"
                           Width="56"
                           Height="56"
                           VerticalAlignment="Bottom"
                           HorizontalAlignment="Left"
                           IsHitTestVisible="False"/>
                    <StackPanel DockPanel.Dock="Right"
                                Orientation="Horizontal"
                                HorizontalAlignment="Right">
                        <Button Name="OkButton"
                                Style="{StaticResource PrimaryButtonStyle}"
                                Width="150"
                                Margin="0,0,8,0"
                                Content="Export"/>
                        <Button Name="CancelButton"
                                Style="{StaticResource SecondaryButtonStyle}"
                                Width="150"
                                Content="Cancel"/>
                    </StackPanel>
                </DockPanel>
            </Grid>
        </Border>
    </Grid>
</Window>'''




def _elem_id_int(eid):
    try:
        return int(eid.Value)      # Revit 2024+
    except AttributeError:
        return int(eid.Value)  # Revit 2023-

def sanitize_sheet_name(name):
    safe = re.sub(r"[:\\/?*\[\]]", "_", (name or "").strip())
    return (safe or "Schedule")[:31]


def _pluralize(name):
    if not name:
        return name
    parts = name.rsplit(" ", 1)
    last = parts[-1]
    if last.lower().endswith(("s", "x", "z", "ch", "sh")):
        last = last + "es"
    else:
        last = last + "s"
    parts[-1] = last
    return " ".join(parts)


def normalize_excel_output_path(path, default_ext=".xlsx"):
    value = (path or "").strip()
    if not value:
        return ""
    root, ext = os.path.splitext(value)
    if not ext:
        return value + default_ext
    if ext.lower() in ALLOWED_EXCEL_EXTENSIONS:
        return value
    return ""


def element_id_value(elem_id):
    if elem_id is None:
        return -1
    if hasattr(elem_id, "IntegerValue"):
        return _elem_id_int(elem_id)
    if hasattr(elem_id, "Value"):
        return elem_id.Value
    try:
        return int(elem_id)
    except Exception:
        return -1


def get_active_doc():
    try:
        uidoc = __revit__.ActiveUIDocument
        if uidoc:
            return uidoc.Document
    except Exception:
        pass
    return None


def read_saved_sets(doc):
    try:
        proj_info = doc.ProjectInformation
        if proj_info is None:
            return {}
        param = proj_info.LookupParameter(PARAM_SAVED_SETS)
        if param is None:
            return {}
        raw = (param.AsString() or "").strip()
        if not raw:
            return {}
        data = json.loads(raw)
        if isinstance(data, dict) and isinstance(data.get(SAVED_SET_NAMESPACE), dict):
            return data.get(SAVED_SET_NAMESPACE) or {}
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _looks_like_legacy_saved_sets(data):
    if not isinstance(data, dict) or not data:
        return False
    if "mass_stats" in data or SAVED_SET_NAMESPACE in data:
        return False
    return all(isinstance(v, dict) for v in data.values())


def write_saved_sets(doc, sets_dict):
    try:
        proj_info = doc.ProjectInformation
        if proj_info is None:
            return False
        param = proj_info.LookupParameter(PARAM_SAVED_SETS)
        if param is None or param.IsReadOnly:
            return False
        raw = (param.AsString() or "").strip()
        try:
            payload = json.loads(raw) if raw else {}
        except Exception:
            payload = {}
        if not isinstance(payload, dict):
            payload = {}
        if _looks_like_legacy_saved_sets(payload):
            payload = {SAVED_SET_NAMESPACE: payload}
        payload[SAVED_SET_NAMESPACE] = sets_dict
        t = DB.Transaction(doc, "Save Export2Ex Settings")
        t.Start()
        param.Set(json.dumps(payload, ensure_ascii=False, indent=2))
        t.Commit()
        return True
    except Exception as exc:
        log_exception("write_saved_sets", exc)
        return False


def _log_file_path():
    appdata = os.environ.get("APPDATA")
    if not appdata:
        appdata = os.path.join(os.path.expanduser("~"), "AppData", "Roaming")
    return os.path.join(appdata, "pyRevit", "WWPTools", "Logs", LOG_FILE_NAME)


def log_message(message):
    try:
        log_path = _log_file_path()
        folder = os.path.dirname(log_path)
        if folder and not os.path.isdir(folder):
            os.makedirs(folder)
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        with open(log_path, "a") as fp:
            fp.write("[{}] {}\n".format(timestamp, message))
    except Exception:
        pass


def log_exception(context, exc):
    try:
        detail = traceback.format_exc()
    except Exception:
        detail = str(exc)
    log_message("{}: {}\n{}".format(context, str(exc), detail))


def add_lib_path():
    lib_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "lib"))
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def load_uiutils():
    add_lib_path()
    import WWP_uiUtils as ui
    return ui


def get_config_and_saver():
    return get_tool_settings("Export2ExBeta", doc=get_active_doc())


def config_get(config, name, default=None):
    try:
        value = getattr(config, name)
    except Exception:
        return default
    return default if value is None else value


def _coerce_int(value, default=None):
    if value is None:
        return default
    if isinstance(value, (list, tuple)):
        for item in value:
            coerced = _coerce_int(item, None)
            if coerced is not None:
                return coerced
        return default
    try:
        return int(value)
    except Exception:
        pass
    text = str(value).strip()
    if not text:
        return default
    match = re.search(r"-?\d+", text)
    if not match:
        return default
    try:
        return int(match.group(0))
    except Exception:
        return default


def _coerce_string_list(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(item) for item in value if str(item).strip()]
    text = str(value).strip()
    return [text] if text else []


def _normalize_mode(value):
    return MODE_BY_CATEGORY if value == MODE_BY_CATEGORY else MODE_FROM_SCHEDULE


def _pick_save_file(title, filter_text, default_extension, initial_directory, file_name):
    clr.AddReference("PresentationFramework")
    from Microsoft.Win32 import SaveFileDialog

    dialog = SaveFileDialog()
    dialog.Title = title or "Save File"
    dialog.Filter = filter_text or "All files (*.*)|*.*"
    if default_extension:
        dialog.DefaultExt = default_extension
        dialog.AddExtension = True
    if initial_directory and os.path.isdir(initial_directory):
        dialog.InitialDirectory = initial_directory
    if file_name:
        dialog.FileName = file_name
    result = dialog.ShowDialog()
    if result:
        return dialog.FileName
    return None


def get_default_dir(doc):
    if doc.IsWorkshared:
        try:
            central = doc.GetWorksharingCentralModelPath()
            if central:
                return os.path.dirname(DB.ModelPathUtils.ConvertModelPathToUserVisiblePath(central))
        except Exception:
            pass
    if doc.PathName:
        return os.path.dirname(doc.PathName)
    return os.path.expanduser("~")


def ensure_existing_dir(path, fallback=""):
    if path and os.path.isdir(path):
        return path
    if fallback and os.path.isdir(fallback):
        return fallback
    return ""


def collect_schedules(doc):
    schedules = []
    for view in DB.FilteredElementCollector(doc).OfClass(DB.ViewSchedule):
        if view.IsTemplate or view.IsTitleblockRevisionSchedule:
            continue
        schedules.append(view)
    schedules.sort(key=lambda item: item.Name)
    return schedules


def collect_category_records(doc):
    records = {}
    collector = DB.FilteredElementCollector(doc).WhereElementIsNotElementType()
    for element in collector:
        try:
            category = element.Category
        except Exception:
            category = None
        if category is None:
            continue
        name = (category.Name or "").strip()
        if not name:
            continue
        category_id = element_id_value(category.Id)
        if category_id == -1:
            continue
        record = records.get(category_id)
        if record is None:
            record = {
                "id": category.Id,
                "id_value": category_id,
                "name": name,
                "count": 0,
            }
            records[category_id] = record
        record["count"] += 1
    result = list(records.values())
    result.sort(key=lambda item: item["name"].lower())
    return result


def get_elements_by_category(doc, category_id):
    try:
        collector = (
            DB.FilteredElementCollector(doc)
            .WherePasses(DB.ElementCategoryFilter(category_id))
            .WhereElementIsNotElementType()
        )
        elements = list(collector.ToElements())
    except Exception:
        elements = []
    return [elem for elem in elements if elem is not None]


def get_schedule_category_id(schedule):
    try:
        category_id = schedule.Definition.CategoryId
    except Exception:
        category_id = None
    if category_id is None or element_id_value(category_id) == -1:
        return None
    return category_id


def iter_parameter_names(element):
    if element is None:
        return
    for param in getattr(element, "Parameters", []):
        try:
            definition = param.Definition
            name = definition.Name if definition else None
        except Exception:
            name = None
        if name:
            yield name


def _iter_element_parameters(element):
    if element is None:
        return
    for param in getattr(element, "Parameters", []):
        try:
            definition = param.Definition
            name = definition.Name if definition else None
        except Exception:
            name = None
        if name:
            yield name, param


def get_parameter_options_for_category(doc, category_id, sample_limit=400):
    options = {}
    seen_types = set()
    for index, element in enumerate(get_elements_by_category(doc, category_id)):
        if index >= sample_limit:
            break
        for name, param in _iter_element_parameters(element):
            record = options.setdefault(name, {"name": name, "editable": False})
            try:
                if param is not None and not param.IsReadOnly:
                    record["editable"] = True
            except Exception:
                pass
        elem_type = get_element_type(doc, element)
        if elem_type is None:
            continue
        type_id = element_id_value(elem_type.Id)
        if type_id in seen_types:
            continue
        seen_types.add(type_id)
        for name, param in _iter_element_parameters(elem_type):
            record = options.setdefault(name, {"name": name, "editable": False})
            try:
                if param is not None and not param.IsReadOnly:
                    record["editable"] = True
            except Exception:
                pass
    try:
        proj_info = doc.ProjectInformation
        if proj_info is not None:
            for name, param in _iter_element_parameters(proj_info):
                record = options.setdefault(name, {"name": name, "editable": False})
                try:
                    if param is not None and not param.IsReadOnly:
                        record["editable"] = True
                except Exception:
                    pass
    except Exception:
        pass
    result = list(options.values())
    result.sort(key=lambda item: item["name"].lower())
    return result


class ScheduleItem(object):
    def __init__(self, view):
        self.view = view
        self.id_value = element_id_value(view.Id)
        self.category_id = get_schedule_category_id(view)
        self.display_name = "{} [id:{}]".format(view.Name, self.id_value).replace("_", "__")

    def __str__(self):
        return self.display_name


class CategoryItem(object):
    def __init__(self, record):
        self.record = record
        self.id_value = record["id_value"]
        self.display_name = "{} ({})".format(record["name"], record["count"]).replace("_", "__")

    def __str__(self):
        return self.display_name


def _to_net_list(values):
    result = List[String]()
    for value in values:
        result.Add("" if value is None else str(value))
    return result


def _to_net_object_list(values):
    result = List[Object]()
    for value in values:
        result.Add(value)
    return result


def _parameter_label(option):
    if option.get("editable", False):
        return option.get("name", "")
    return "{} (read-only)".format(option.get("name", ""))


def _make_parameter_list_item(option):
    item = ListBoxItem()
    item.Content = _parameter_label(option)
    item.Tag = option
    if not option.get("editable", False):
        item.Foreground = Brushes.Gray
        item.ToolTip = "This parameter is read-only in the sampled category elements/types."
    return item


def _get_parameter_name_from_item(item):
    if item is None:
        return ""
    try:
        tag = item.Tag
    except Exception:
        tag = None
    if isinstance(tag, dict):
        return tag.get("name", "")
    return str(item)


def _load_export_window():
    from System.IO import StringReader
    from System.Windows.Markup import XamlReader
    from System.Xml import XmlReader

    xaml_path = os.path.join(os.path.dirname(__file__), "ExportSchedulesDialog.xaml")
    sources = [
        ("file", lambda: File.ReadAllText(xaml_path)),
        ("embedded", lambda: EMBEDDED_EXPORT_DIALOG_XAML),
    ]
    last_exc = None
    for source_name, source_loader in sources:
        try:
            xaml_text = source_loader()
            reader = XmlReader.Create(StringReader(xaml_text))
            window = XamlReader.Load(reader)
            if source_name != "file":
                log_message("Loaded Export2Ex Beta dialog from {} fallback".format(source_name))
            return window
        except Exception as exc:
            last_exc = exc
            log_exception("Failed to load Export2Ex Beta dialog from {}".format(source_name), exc)
    raise last_exc


def show_export_form(ui, doc, schedules, categories, init_excel_path, initial_mode, initial_source_id, initial_category_id, initial_param_names, initial_sheet_name=""):
    clr.AddReference("PresentationFramework")
    clr.AddReference("PresentationCore")
    clr.AddReference("WindowsBase")
    from System import Uri
    from System.Windows.Media.Imaging import BitmapCacheOption, BitmapImage

    window = _load_export_window()
    apply_window_title(window, "Export2Ex Beta")

    mode_box = window.FindName("ModeBox")
    source_search_box = window.FindName("SourceSearchBox")
    source_label = window.FindName("SourceLabel")
    source_list_label = window.FindName("SourceListLabel")
    source_list = window.FindName("SourceList")
    parameter_search_box = window.FindName("ParameterSearchBox")
    parameter_list = window.FindName("ParameterList")
    selected_parameter_list = window.FindName("SelectedParameterList")
    add_parameter_button = window.FindName("AddParameterButton")
    remove_parameter_button = window.FindName("RemoveParameterButton")
    move_parameter_up_button = window.FindName("MoveParameterUpButton")
    move_parameter_down_button = window.FindName("MoveParameterDownButton")
    excel_path = window.FindName("ExcelPath")
    browse_excel = window.FindName("BrowseExcel")
    ok_button = window.FindName("OkButton")
    cancel_button = window.FindName("CancelButton")
    logo_image = window.FindName("LogoImage")
    show_read_only_filter = window.FindName("ShowReadOnlyFilter")
    sheet_name_box = window.FindName("SheetNameBox")
    saved_set_box = window.FindName("SavedSetBox")
    load_set_button = window.FindName("LoadSetButton")
    save_set_button = window.FindName("SaveSetButton")
    delete_set_button = window.FindName("DeleteSetButton")

    excel_path.Text = init_excel_path or ""
    schedule_items = schedules or []
    category_items = categories or []
    parameter_cache = {}
    selected_params_by_category = {}
    initial_category_id = _coerce_int(initial_category_id, None)
    initial_source_id = _coerce_int(initial_source_id, None)
    initial_param_names = _coerce_string_list(initial_param_names)
    if initial_category_id is not None and initial_param_names:
        selected_params_by_category[initial_category_id] = list(initial_param_names)

    def _current_mode():
        try:
            return MODE_BY_CATEGORY if mode_box.SelectedIndex == 1 else MODE_FROM_SCHEDULE
        except Exception:
            return MODE_FROM_SCHEDULE

    if initial_mode == MODE_BY_CATEGORY:
        mode_box.SelectedIndex = 1
    else:
        mode_box.SelectedIndex = 0

    def _get_source_items():
        return category_items if _current_mode() == MODE_BY_CATEGORY else schedule_items

    def _resolve_category_id(item):
        if item is None:
            return None
        if _current_mode() == MODE_BY_CATEGORY:
            return item.id_value
        if item.category_id is None:
            return None
        return element_id_value(item.category_id)

    def _get_parameter_names(category_id):
        if category_id is None:
            return []
        if category_id not in parameter_cache:
            parameter_cache[category_id] = get_parameter_options_for_category(doc, DB.ElementId(category_id))
        return parameter_cache[category_id]

    def _get_selected_parameter_names(category_id):
        if category_id is None:
            return []
        selected = selected_params_by_category.get(category_id)
        if selected is None:
            selected = []
            selected_params_by_category[category_id] = selected
        return selected

    def _refresh_selected_parameter_list(category_id, preserve_selection=None):
        selected_names = _get_selected_parameter_names(category_id)
        selected_parameter_list.Items.Clear()
        option_map = dict((option["name"], option) for option in _get_parameter_names(category_id))
        for name in selected_names:
            option = option_map.get(name, {"name": name, "editable": True})
            selected_parameter_list.Items.Add(_make_parameter_list_item(option))
        preserve = list(preserve_selection or [])
        try:
            selected_parameter_list.SelectedItems.Clear()
            for item in selected_parameter_list.Items:
                if _get_parameter_name_from_item(item) in preserve:
                    selected_parameter_list.SelectedItems.Add(item)
        except Exception:
            pass

    def _refresh_parameter_list():
        selected_item = source_list.SelectedItem
        category_id = _resolve_category_id(selected_item)
        all_options = _get_parameter_names(category_id)
        selected_names = set(_get_selected_parameter_names(category_id))
        text = (parameter_search_box.Text or "").strip().lower()
        filtered = [option for option in all_options if option["name"] not in selected_names]
        try:
            if not show_read_only_filter.IsChecked:
                filtered = [option for option in filtered if option.get("editable", False)]
        except Exception:
            pass
        if text:
            filtered = [option for option in filtered if text in option["name"].lower()]
        parameter_list.Items.Clear()
        for option in filtered:
            parameter_list.Items.Add(_make_parameter_list_item(option))
        _refresh_selected_parameter_list(category_id)

    def _refresh_source_list():
        mode = _current_mode()
        items = _get_source_items()
        text = (source_search_box.Text or "").strip().lower()
        filtered = items if not text else [item for item in items if text in item.display_name.lower()]
        source_list.ItemsSource = _to_net_object_list(filtered)
        source_label.Text = "Search categories" if mode == MODE_BY_CATEGORY else "Search schedules"
        source_list_label.Text = "Categories" if mode == MODE_BY_CATEGORY else "Schedules"
        target_id = initial_source_id
        selected = None
        for item in filtered:
            item_id = item.id_value if mode == MODE_BY_CATEGORY else item.id_value
            if target_id is not None and item_id == target_id:
                selected = item
                break
        if selected is None and filtered:
            selected = filtered[0]
        source_list.SelectedItem = selected
        _refresh_parameter_list()

    _initialized = [False]

    def _auto_populate_schedule_fields():
        if not _initialized[0] or _current_mode() != MODE_FROM_SCHEDULE:
            return
        selected_item = source_list.SelectedItem
        if selected_item is None:
            return
        category_id = _resolve_category_id(selected_item)
        if category_id is None:
            return
        try:
            fields = get_visible_schedule_fields(selected_item.view)
        except Exception:
            return
        names = []
        for f in fields:
            try:
                name = f["field"].GetName()
                if name and name not in names:
                    names.append(name)
            except Exception:
                pass
        selected_params_by_category[category_id] = names

    def _source_selection_changed(_sender, _args):
        _auto_populate_schedule_fields()
        _refresh_parameter_list()

    def _add_parameters(_sender=None, _args=None):
        selected_item = source_list.SelectedItem
        category_id = _resolve_category_id(selected_item)
        if category_id is None:
            return
        selected_names = _get_selected_parameter_names(category_id)
        for item in parameter_list.SelectedItems:
            name = _get_parameter_name_from_item(item)
            if name not in selected_names:
                selected_names.append(name)
        _refresh_parameter_list()

    def _remove_parameters(_sender=None, _args=None):
        selected_item = source_list.SelectedItem
        category_id = _resolve_category_id(selected_item)
        if category_id is None:
            return
        selected_names = _get_selected_parameter_names(category_id)
        removed = set(_get_parameter_name_from_item(item) for item in selected_parameter_list.SelectedItems)
        if not removed:
            return
        selected_params_by_category[category_id] = [name for name in selected_names if name not in removed]
        _refresh_parameter_list()

    def _move_selected_parameters(direction):
        selected_item = source_list.SelectedItem
        category_id = _resolve_category_id(selected_item)
        if category_id is None:
            return
        selected_names = _get_selected_parameter_names(category_id)
        selected_now = [_get_parameter_name_from_item(item) for item in selected_parameter_list.SelectedItems]
        if not selected_now:
            return
        indices = [idx for idx, name in enumerate(selected_names) if name in selected_now]
        if direction < 0:
            for idx in indices:
                if idx <= 0:
                    continue
                selected_names[idx - 1], selected_names[idx] = selected_names[idx], selected_names[idx - 1]
        else:
            for idx in reversed(indices):
                if idx >= len(selected_names) - 1:
                    continue
                selected_names[idx + 1], selected_names[idx] = selected_names[idx], selected_names[idx + 1]
        _refresh_selected_parameter_list(category_id, preserve_selection=selected_now)

    def _move_up(_sender=None, _args=None):
        _move_selected_parameters(-1)

    def _move_down(_sender=None, _args=None):
        _move_selected_parameters(1)

    def _browse_excel(_sender, _args):
        current = excel_path.Text or ""
        file_name = os.path.basename(current) if current else "CategoryExport.xlsx"
        initial_directory = ensure_existing_dir(
            os.path.dirname(current) if current else "",
            os.path.dirname(init_excel_path) if init_excel_path else get_default_dir(get_active_doc()),
        )
        try:
            file_path = _pick_save_file(
                title="Export Category Data to Excel",
                filter_text="Excel Workbook (*.xlsx;*.xlsm)|*.xlsx;*.xlsm",
                default_extension="xlsx",
                initial_directory=initial_directory,
                file_name=file_name,
            )
        except Exception as exc:
            log_exception("Native browse Excel dialog failed", exc)
            try:
                file_path = _pick_save_file(
                    title="Export Category Data to Excel",
                    filter_text="Excel Workbook (*.xlsx;*.xlsm)|*.xlsx;*.xlsm",
                    default_extension="xlsx",
                    initial_directory="",
                    file_name=file_name,
                )
            except Exception as retry_exc:
                log_exception("Native browse Excel dialog retry failed", retry_exc)
                ui.uiUtils_alert(
                    "Could not open the Excel save dialog. Check the suggested path and try again.",
                    title="Export2Ex Beta",
                )
                return
        if file_path:
            excel_path.Text = file_path

    saved_sets = read_saved_sets(doc)

    def _refresh_saved_set_dropdown():
        if saved_set_box is None:
            return
        current_text = saved_set_box.Text or ""
        saved_set_box.Items.Clear()
        for name in sorted(saved_sets.keys()):
            saved_set_box.Items.Add(name)
        saved_set_box.Text = current_text

    def _get_current_set_data():
        selected_item = source_list.SelectedItem
        category_id = _resolve_category_id(selected_item)
        source_id = selected_item.id_value if selected_item is not None else None
        return {
            "mode": _current_mode(),
            "source_id": source_id,
            "category_id": category_id,
            "param_names": list(selected_params_by_category.get(category_id, [])),
            "sheet_name": (sheet_name_box.Text or "").strip() if sheet_name_box is not None else "",
        }

    def _apply_saved_set(set_data):
        _initialized[0] = False
        try:
            mode = set_data.get("mode", MODE_FROM_SCHEDULE)
            source_id = _coerce_int(set_data.get("source_id"), None)
            category_id = _coerce_int(set_data.get("category_id"), None)
            param_names = set_data.get("param_names") or []
            saved_sheet = (set_data.get("sheet_name") or "").strip()
            if category_id is not None:
                selected_params_by_category[category_id] = list(param_names)
            mode_box.SelectedIndex = 1 if mode == MODE_BY_CATEGORY else 0
            source_search_box.Text = ""
            for item in _get_source_items():
                if item.id_value == source_id:
                    source_list.SelectedItem = item
                    break
            if sheet_name_box is not None and saved_sheet:
                sheet_name_box.Text = saved_sheet
            _refresh_parameter_list()
        finally:
            _initialized[0] = True

    def _load_set(_sender=None, _args=None):
        if saved_set_box is None:
            return
        name = (saved_set_box.Text or "").strip()
        if not name or name not in saved_sets:
            ui.uiUtils_alert("Select a saved set from the dropdown to load.", title="Export2Ex Beta")
            return
        _apply_saved_set(saved_sets[name])

    def _save_set(_sender=None, _args=None):
        if saved_set_box is None:
            return
        name = (saved_set_box.Text or "").strip()
        if not name:
            ui.uiUtils_alert("Type a name for the saved set before saving.", title="Export2Ex Beta")
            return
        saved_sets[name] = _get_current_set_data()
        if not write_saved_sets(doc, saved_sets):
            ui.uiUtils_alert(
                "Could not write to '{}' on Project Information.\nCheck the parameter exists and is not read-only.".format(PARAM_SAVED_SETS),
                title="Export2Ex Beta",
            )
            return
        _refresh_saved_set_dropdown()

    def _delete_set(_sender=None, _args=None):
        if saved_set_box is None:
            return
        name = (saved_set_box.Text or "").strip()
        if not name or name not in saved_sets:
            ui.uiUtils_alert("Select a saved set from the dropdown to delete.", title="Export2Ex Beta")
            return
        del saved_sets[name]
        if not write_saved_sets(doc, saved_sets):
            ui.uiUtils_alert(
                "Could not update '{}' on Project Information.".format(PARAM_SAVED_SETS),
                title="Export2Ex Beta",
            )
            return
        saved_set_box.Text = ""
        _refresh_saved_set_dropdown()

    def _ok(_sender, _args):
        window.DialogResult = True
        window.Close()

    def _cancel(_sender, _args):
        window.DialogResult = False
        window.Close()

    try:
        logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "lib", "WWPtools-logo.png"))
        if logo_image is not None and os.path.isfile(logo_path):
            bitmap = BitmapImage()
            bitmap.BeginInit()
            bitmap.UriSource = Uri(logo_path)
            bitmap.CacheOption = BitmapCacheOption.OnLoad
            bitmap.EndInit()
            logo_image.Source = bitmap
    except Exception:
        pass

    mode_box.SelectionChanged += lambda _sender, _args: _refresh_source_list()
    source_search_box.TextChanged += lambda _sender, _args: _refresh_source_list()
    source_list.SelectionChanged += _source_selection_changed
    parameter_search_box.TextChanged += lambda _sender, _args: _refresh_parameter_list()
    show_read_only_filter.Checked += lambda _sender, _args: _refresh_parameter_list()
    show_read_only_filter.Unchecked += lambda _sender, _args: _refresh_parameter_list()
    add_parameter_button.Click += _add_parameters
    remove_parameter_button.Click += _remove_parameters
    move_parameter_up_button.Click += _move_up
    move_parameter_down_button.Click += _move_down
    browse_excel.Click += _browse_excel
    ok_button.Click += _ok
    cancel_button.Click += _cancel
    if load_set_button is not None:
        load_set_button.Click += _load_set
    if save_set_button is not None:
        save_set_button.Click += _save_set
    if delete_set_button is not None:
        delete_set_button.Click += _delete_set
    _refresh_saved_set_dropdown()
    _refresh_source_list()
    _initialized[0] = True

    if sheet_name_box is not None:
        initial_sheet_name_stripped = (initial_sheet_name or "").strip()
        if initial_sheet_name_stripped:
            sheet_name_box.Text = initial_sheet_name_stripped
        else:
            selected = source_list.SelectedItem
            if selected is not None:
                if _current_mode() == MODE_BY_CATEGORY:
                    auto_name = _pluralize(selected.record["name"])
                else:
                    auto_name = selected.view.Name
                sheet_name_box.Text = sanitize_sheet_name(auto_name)

    if not window.ShowDialog():
        return None

    selected_item = source_list.SelectedItem
    category_id = _resolve_category_id(selected_item)
    source_id = None
    source_name = ""
    if selected_item is not None:
        if _current_mode() == MODE_BY_CATEGORY:
            source_id = selected_item.id_value
            source_name = selected_item.record["name"]
        else:
            source_id = selected_item.id_value
            source_name = selected_item.view.Name
    return {
        "mode": _current_mode(),
        "source_id": source_id,
        "source_name": source_name,
        "category_id": category_id,
        "selected_param_names": list(selected_params_by_category.get(category_id, [])),
        "excel_path": excel_path.Text or "",
        "sheet_name": (sheet_name_box.Text or "").strip() if sheet_name_box is not None else "",
    }


def get_section_data(schedule, section_type):
    try:
        return schedule.GetTableData().GetSectionData(section_type)
    except Exception:
        return None


def get_cell_text(schedule, section_type, row, col):
    try:
        return schedule.GetCellText(section_type, row, col) or ""
    except Exception:
        pass
    try:
        section = get_section_data(schedule, section_type)
        if section is not None:
            return section.GetCellText(row, col) or ""
    except Exception:
        pass
    return ""


def get_body_row_element_ids(schedule):
    section = get_section_data(schedule, DB.SectionType.Body)
    if section is None:
        return []
    ids = []
    row_count = int(getattr(section, "NumberOfRows", 0))
    col_count = int(getattr(section, "NumberOfColumns", 0))
    for row in range(row_count):
        found = -1
        for col in range(col_count):
            try:
                elem_id = section.GetCellElementId(row, col)
            except Exception:
                elem_id = None
            found = element_id_value(elem_id)
            if found != -1:
                break
        ids.append(found)
    return ids


def collect_schedule_elements(schedule):
    try:
        elements = list(
            DB.FilteredElementCollector(schedule.Document, schedule.Id)
            .WhereElementIsNotElementType()
            .ToElements()
        )
    except Exception:
        elements = []
    return [elem for elem in elements if elem is not None]


def get_visible_schedule_fields(schedule):
    fields = []
    try:
        definition = schedule.Definition
        field_ids = list(definition.GetFieldOrder())
    except Exception:
        return fields
    for col_index, field_id in enumerate(field_ids):
        try:
            field = definition.GetField(field_id)
        except Exception:
            continue
        if field is None:
            continue
        try:
            if field.IsHidden:
                continue
        except Exception:
            pass
        try:
            heading = (field.ColumnHeading or "").strip()
        except Exception:
            heading = ""
        if not heading:
            try:
                heading = (field.GetName() or "").strip()
            except Exception:
                heading = ""
        fields.append(
            {
                "field": field,
                "column_index": len(fields),
                "heading": heading or "Column {}".format(col_index + 1),
                "param_id": safe_field_parameter_id(field),
            }
        )
    return fields


def safe_field_parameter_id(field):
    try:
        param_id = field.ParameterId
    except Exception:
        param_id = None
    if param_id is not None and element_id_value(param_id) == -1:
        return None
    return param_id


def is_key_schedule(view):
    try:
        definition = view.Definition
    except Exception:
        definition = None
    try:
        return bool(definition and definition.IsKeySchedule)
    except Exception:
        return False


def get_element_type(doc, element):
    try:
        type_id = element.GetTypeId()
    except Exception:
        type_id = None
    if type_id is None or element_id_value(type_id) == -1:
        return None
    try:
        return doc.GetElement(type_id)
    except Exception:
        return None


def get_parameter_from_element_or_type(doc, element, param_id):
    if element is None or param_id is None:
        return None
    try:
        param = element.get_Parameter(param_id)
    except Exception:
        param = None
    if param:
        return param
    elem_type = get_element_type(doc, element)
    if elem_type is None:
        return None
    try:
        return elem_type.get_Parameter(param_id)
    except Exception:
        return None


def parameter_to_text(doc, param):
    if not param:
        return ""
    try:
        value = param.AsValueString()
        if value not in (None, ""):
            return value
    except Exception:
        pass
    try:
        storage = param.StorageType
    except Exception:
        storage = None
    try:
        if storage == DB.StorageType.String:
            return param.AsString() or ""
        if storage == DB.StorageType.Integer:
            return str(param.AsInteger())
        if storage == DB.StorageType.Double:
            return str(param.AsDouble())
        if storage == DB.StorageType.ElementId:
            ref_id = param.AsElementId()
            ref_val = element_id_value(ref_id)
            if ref_val == -1:
                return ""
            ref_elem = doc.GetElement(ref_id)
            if ref_elem is not None:
                for attr in ("Name",):
                    try:
                        value = getattr(ref_elem, attr)
                        if value:
                            return str(value)
                    except Exception:
                        pass
            return str(ref_val)
    except Exception:
        pass
    return ""


def get_parameter_by_name(doc, element, param_name):
    if element is None or not param_name:
        return None
    try:
        param = element.LookupParameter(param_name)
    except Exception:
        param = None
    if param:
        return param
    elem_type = get_element_type(doc, element)
    if elem_type is not None:
        try:
            param = elem_type.LookupParameter(param_name)
        except Exception:
            param = None
        if param:
            return param
    try:
        proj_info = doc.ProjectInformation
        if proj_info is not None:
            return proj_info.LookupParameter(param_name)
    except Exception:
        pass
    return None


def parameter_to_export_value(doc, param):
    if not param:
        return ""
    try:
        storage = param.StorageType
    except Exception:
        storage = None
    try:
        if storage == DB.StorageType.String:
            return param.AsString() or ""
        if storage == DB.StorageType.Double:
            return param.AsDouble()
        if storage == DB.StorageType.Integer:
            value_string = param.AsValueString()
            if value_string not in (None, "") and value_string not in ("0", "1"):
                return value_string
            return param.AsInteger()
        if storage == DB.StorageType.ElementId:
            ref_id = param.AsElementId()
            ref_value = element_id_value(ref_id)
            if ref_value == -1:
                return ""
            ref_elem = doc.GetElement(ref_id)
            if ref_elem is not None:
                try:
                    name = getattr(ref_elem, "Name", None)
                    if name:
                        return name
                except Exception:
                    pass
            return str(ref_value)
    except Exception:
        pass
    try:
        value = param.AsValueString()
        if value not in (None, ""):
            return value
    except Exception:
        pass
    return ""


def build_category_export_rows(doc, category_id, param_names):
    elements = get_elements_by_category(doc, category_id)
    elements.sort(key=lambda item: element_id_value(item.Id))
    headers = ["Id"] + list(param_names or [])
    rows = []
    for element in elements:
        row = [element_id_value(element.Id)]
        for param_name in param_names or []:
            row.append(parameter_to_export_value(doc, get_parameter_by_name(doc, element, param_name)))
        rows.append(row)
    return headers, rows, len(elements)


def build_schedule_body_rows(schedule):
    section = get_section_data(schedule, DB.SectionType.Body)
    if section is None:
        return []
    row_count = int(getattr(section, "NumberOfRows", 0))
    col_count = int(getattr(section, "NumberOfColumns", 0))
    rows = []
    for row in range(row_count):
        rows.append([get_cell_text(schedule, DB.SectionType.Body, row, col) for col in range(col_count)])
    return rows


def build_schedule_row_lookup(schedule):
    row_ids = get_body_row_element_ids(schedule)
    row_values = build_schedule_body_rows(schedule)
    lookup = {}
    for row_id, values in zip(row_ids, row_values):
        if row_id == -1:
            continue
        lookup.setdefault(row_id, []).append(values)
    return row_ids, lookup


def build_table_headers(schedule, column_count):
    section = get_section_data(schedule, DB.SectionType.Header)
    row_count = int(getattr(section, "NumberOfRows", 0)) if section is not None else 0
    if row_count > 0:
        last_row = row_count - 1
        headers = [get_cell_text(schedule, DB.SectionType.Header, last_row, col) for col in range(column_count)]
        if any(header.strip() for header in headers):
            return [header or "Column {}".format(idx + 1) for idx, header in enumerate(headers)]
    return ["Column {}".format(idx + 1) for idx in range(column_count)]


def order_schedule_elements(schedule, elements, row_ids):
    elem_by_id = {}
    ordered = []
    seen = set()
    for elem in elements:
        elem_by_id[element_id_value(elem.Id)] = elem
    for row_id in row_ids:
        if row_id in seen:
            continue
        elem = elem_by_id.get(row_id)
        if elem is not None:
            ordered.append(elem)
            seen.add(row_id)
    for elem in sorted(elements, key=lambda item: element_id_value(item.Id)):
        elem_id = element_id_value(elem.Id)
        if elem_id not in seen:
            ordered.append(elem)
            seen.add(elem_id)
    return ordered


def get_field_text_for_element(doc, schedule, element, field_info, row_lookup):
    param_id = field_info.get("param_id")
    value = ""
    if param_id is not None:
        value = parameter_to_text(doc, get_parameter_from_element_or_type(doc, element, param_id))
    if value not in (None, ""):
        return value
    row_queue = row_lookup.get(element_id_value(element.Id)) or []
    if row_queue:
        column_index = int(field_info.get("column_index", 0))
        if column_index < len(row_queue[0]):
            return row_queue[0][column_index]
    return ""


def _inject_element_ids(headers, rows, ids):
    """Prepend an 'Element ID' column to headers and rows.

    `ids` is a parallel list of integer element IDs (or -1 for rows without
    a traceable element, e.g. group-header rows).  Rows with id == -1 get an
    empty cell in the Element ID column.
    """
    new_headers = ["Element ID"] + list(headers)
    new_rows = []
    for i, row in enumerate(rows):
        eid = ids[i] if i < len(ids) else -1
        new_rows.append(([str(eid)] if eid != -1 else [""]) + list(row))
    return new_headers, new_rows


def strip_leading_header_and_blanks(body_rows, headers=None):
    """Remove duplicate column-header rows and blank separator rows from the
    top of the body data.

    Some schedule types (key schedules, schedules with the 'blank row before
    data' appearance option) include the column-header row and/or a blank row
    as the first rows of the body section.  Because the exporter already writes
    its own bold header row, those body rows must be dropped to avoid
    duplication and empty rows in the output.
    """
    if not body_rows:
        return body_rows
    result = list(body_rows)
    headers_lower = [h.strip().lower() for h in headers] if headers else []
    while result:
        row = result[0]
        row_lower = [cell.strip().lower() for cell in row]
        if headers_lower and row_lower == headers_lower:
            result.pop(0)
        elif all(cell == "" for cell in row):
            result.pop(0)
        else:
            break
    return result


def _strip_body_rows(body_rows, headers, raw_ids):
    """Strip leading header/blank rows and keep raw_ids in sync.

    Returns (stripped_rows, aligned_ids) where aligned_ids[i] is the element
    ID that corresponds to stripped_rows[i].
    """
    n_before = len(body_rows)
    stripped = strip_leading_header_and_blanks(body_rows, headers)
    n_stripped = n_before - len(stripped)
    aligned_ids = (raw_ids[n_stripped:n_stripped + len(stripped)]
                   if raw_ids else [-1] * len(stripped))
    return stripped, aligned_ids


def build_schedule_export_rows(doc, schedule):
    fields = get_visible_schedule_fields(schedule)
    elements = collect_schedule_elements(schedule)
    key_sched = is_key_schedule(schedule)

    # GetCellText-based approach is the most reliable primary source because it
    # reflects what Revit actually renders in the schedule, covering calculated
    # value fields and schedules where GetCellElementId returns invalid IDs
    # (grouped schedules, key schedules, etc.).
    body_rows = build_schedule_body_rows(schedule)
    # Fetch element IDs now, before any stripping, so indices stay aligned.
    raw_ids = [] if key_sched else get_body_row_element_ids(schedule)

    if not fields:
        headers = build_table_headers(schedule, len(body_rows[0]) if body_rows else 0)
        body_rows, aligned_ids = _strip_body_rows(body_rows, headers, raw_ids)
        if not key_sched:
            headers, body_rows = _inject_element_ids(headers, body_rows, aligned_ids)
        return headers, body_rows, len(elements)

    headers = [field["heading"] for field in fields]

    # Use direct cell text when available and column count matches visible fields.
    if body_rows and len(body_rows[0]) == len(fields):
        body_rows, aligned_ids = _strip_body_rows(body_rows, headers, raw_ids)
        if not key_sched:
            headers, body_rows = _inject_element_ids(headers, body_rows, aligned_ids)
        return headers, body_rows, len(elements)

    # Fall back to element-based parameter lookup when GetCellText gives nothing
    # or the column count doesn't match.
    row_ids, row_lookup = build_schedule_row_lookup(schedule)
    ordered_elements = order_schedule_elements(schedule, elements, row_ids)
    if not ordered_elements:
        body_rows, aligned_ids = _strip_body_rows(body_rows, headers, raw_ids)
        if not key_sched:
            headers, body_rows = _inject_element_ids(headers, body_rows, aligned_ids)
        return headers, body_rows, len(elements)

    rows = []
    for element in ordered_elements:
        row = [get_field_text_for_element(doc, schedule, element, field, row_lookup) for field in fields]
        rows.append(row)

    # If the element-based approach also produces all-empty data, prefer the
    # raw cell-text rows even when the column count differs.
    if rows and all(all(cell == "" for cell in row) for row in rows) and body_rows:
        body_rows, aligned_ids = _strip_body_rows(body_rows, headers, raw_ids)
        if not key_sched:
            headers, body_rows = _inject_element_ids(headers, body_rows, aligned_ids)
        return headers, body_rows, len(elements)

    # Element-based path: IDs come from the ordered elements directly.
    if not key_sched:
        elem_ids = [element_id_value(e.Id) for e in ordered_elements]
        headers, rows = _inject_element_ids(headers, rows, elem_ids)
    return headers, rows, len(elements)


def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def make_unique_name(base, used):
    candidate = base
    if candidate not in used:
        used.add(candidate)
        return candidate
    idx = 1
    while True:
        suffix = "_{}".format(idx)
        trimmed = candidate[: 31 - len(suffix)]
        attempt = "{}{}".format(trimmed, suffix)
        if attempt not in used:
            used.add(attempt)
            return attempt
        idx += 1


def export_to_excel(doc, category_name, category_id, param_names, file_path, ui, sheet_name=None):
    add_lib_path()
    try:
        import openpyxl
        from openpyxl.styles import Font
    except Exception as exc:
        ui.uiUtils_alert("openpyxl is not available.\n{}".format(exc), title="Export2Ex Beta")
        return False

    if os.path.exists(file_path):
        load_kwargs = {}
        if os.path.splitext(file_path)[1].lower() == ".xlsm":
            load_kwargs["keep_vba"] = True
        workbook = openpyxl.load_workbook(file_path, **load_kwargs)
    else:
        workbook = openpyxl.Workbook()

    base_name = sanitize_sheet_name(sheet_name if sheet_name else _pluralize(category_name))
    sheet_name = base_name
    if sheet_name in workbook.sheetnames:
        existing = workbook[sheet_name]
        sheet_index = workbook.worksheets.index(existing)
        workbook.remove(existing)
        sheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    headers, rows, elem_count = build_category_export_rows(doc, category_id, param_names)
    log_message(
        "Category '{}' resolved {} elements and {} export columns".format(
            category_name, elem_count, len(headers)
        )
    )

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    if headers:
        sheet.freeze_panes = "A2"
    auto_fit_columns(sheet)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])

    workbook.save(file_path)
    return True


def show_error_report(ui, exc):
    report = "Export2Ex Beta failed.\n\n{}\n\nLog File\n{}".format(str(exc), _log_file_path())
    try:
        ui.uiUtils_alert(report, title="Export2Ex Beta")
    except Exception:
        pass


def main():
    log_message("main start")
    doc = get_active_doc()
    ui = load_uiutils()
    if doc is None:
        ui.uiUtils_alert("No active Revit document found.", title="Export2Ex Beta")
        return

    schedules = [ScheduleItem(view) for view in collect_schedules(doc)]
    categories = [CategoryItem(record) for record in collect_category_records(doc)]
    if not schedules and not categories:
        ui.uiUtils_alert("No schedules or categories with elements were found.", title="Export2Ex Beta")
        return

    config, save_config = get_config_and_saver()
    default_dir = get_default_dir(doc)
    last_excel_path = config_get(config, CONFIG_LAST_EXCEL_PATH, "")
    init_excel_path = last_excel_path or os.path.join(default_dir, "CategoryExport.xlsx")
    last_mode = _normalize_mode(config_get(config, CONFIG_LAST_MODE, MODE_FROM_SCHEDULE))
    last_source_id = config_get(
        config,
        CONFIG_LAST_CATEGORY_ID if last_mode == MODE_BY_CATEGORY else CONFIG_LAST_SCHEDULE_ID,
        None,
    )
    last_category_id = _coerce_int(config_get(config, CONFIG_LAST_CATEGORY_ID, None), None)
    last_param_names = _coerce_string_list(config_get(config, CONFIG_LAST_PARAM_NAMES, []))
    last_sheet_name = str(config_get(config, CONFIG_LAST_SHEET_NAME, "") or "").strip()
    result = show_export_form(
        ui,
        doc,
        schedules,
        categories,
        init_excel_path,
        last_mode,
        last_source_id,
        last_category_id,
        last_param_names,
        last_sheet_name,
    )
    if not result:
        return

    category_id_value = _coerce_int(result.get("category_id"), None)
    if category_id_value in (None, -1):
        ui.uiUtils_alert("Select a schedule or category with a valid category.", title="Export2Ex Beta")
        return

    selected_param_names = result.get("selected_param_names") or []
    if not selected_param_names:
        ui.uiUtils_alert("Select at least one parameter to export.", title="Export2Ex Beta")
        return

    file_path = normalize_excel_output_path(result.get("excel_path"))
    if not file_path:
        ui.uiUtils_alert(
            "Choose an Excel file path ending with .xlsx or .xlsm.",
            title="Export2Ex Beta",
        )
        return

    category_record = None
    for item in categories:
        if item.id_value == category_id_value:
            category_record = item.record
            break
    category_name = category_record["name"] if category_record else (result.get("source_name") or "Category Export")

    sheet_name_raw = (result.get("sheet_name") or "").strip()
    if not export_to_excel(doc, category_name, DB.ElementId(category_id_value), selected_param_names, file_path, ui, sheet_name=sheet_name_raw):
        return

    config.last_mode = _normalize_mode(result.get("mode"))
    config.last_schedule_id = result.get("source_id") if config.last_mode == MODE_FROM_SCHEDULE else None
    config.last_category_id = category_id_value
    config.last_param_names = list(selected_param_names)
    config.last_excel_path = file_path
    config.last_sheet_name = sheet_name_raw
    save_config()

    try:
        os.startfile(file_path)
    except Exception:
        pass
    ui.uiUtils_alert("Export complete.", title="Export2Ex Beta")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log_exception("Unhandled exception in Export2Ex Beta", exc)
        try:
            show_error_report(load_uiutils(), exc)
        except Exception:
            pass
        raise
