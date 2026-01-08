#! python3
from __future__ import annotations

import csv
import os
import re
import shutil
import sys
import tempfile

from pyrevit import DB, revit, script

import System
from System.Drawing import Point, Size
from System.Windows.Forms import (
    Button,
    CheckedListBox,
    DialogResult,
    Form,
    Label,
    MessageBox,
    RadioButton,
    SaveFileDialog,
    FolderBrowserDialog,
)


CONFIG_LAST_EXCEL_PATH = "last_excel_path"
CONFIG_LAST_CSV_DIR = "last_csv_dir"
CONFIG_LAST_SCHEDULE_IDS = "last_schedule_ids"


def sanitize_sheet_name(name):
    invalid = r"[:\\/?*\[\]]"
    safe = re.sub(invalid, "_", name)
    safe = safe.strip()
    if not safe:
        safe = "Schedule"
    return safe[:31]


def sanitize_file_name(name):
    invalid = r'[<>:"/\\|?*]'
    safe = re.sub(invalid, "_", name).strip()
    return safe or "Schedule"


def get_default_dir(doc):
    if doc.IsWorkshared:
        try:
            central = doc.GetWorksharingCentralModelPath()
            if central:
                return os.path.dirname(DB.ModelPathUtils.ConvertModelPathToUserVisiblePath(central))
        except Exception:
            pass
    if doc.PathName:
        return os.path.dirname(doc.PathName)
    return os.path.expanduser("~")


def collect_schedules(doc):
    schedules = []
    for view in DB.FilteredElementCollector(doc).OfClass(DB.ViewSchedule):
        if view.IsTemplate:
            continue
        try:
            if view.ViewType == DB.ViewType.Legend:
                continue
        except Exception:
            pass
        if view.IsTitleblockRevisionSchedule:
            continue
        if view.Definition and view.Definition.IsKeySchedule:
            continue
        schedules.append(view)
    schedules.sort(key=lambda v: v.Name)
    return schedules


class ScheduleItem(object):
    def __init__(self, view):
        self.view = view
        self.display_name = "{} [id:{}]".format(view.Name, view.Id.IntegerValue)


def add_lib_path():
    lib_path = os.path.join(os.path.dirname(__file__), "lib")
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def read_csv_rows(path):
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                return [row for row in reader]
        except Exception:
            continue
    return []


def normalize_table_data(data):
    return data


def get_section_data(view, section_type):
    table = view.GetTableData()
    try:
        return table.GetSectionData(section_type)
    except Exception:
        return None


def get_section_row_count(view, section_type):
    section = get_section_data(view, section_type)
    if section is None:
        return 0
    try:
        return section.NumberOfRows
    except Exception:
        return 0


def get_body_row_element_ids(view):
    section = get_section_data(view, DB.SectionType.Body)
    if section is None:
        return []
    ids = []
    for row in range(section.NumberOfRows):
        elem_value = ""
        try:
            col_count = section.NumberOfColumns
        except Exception:
            col_count = 0
        for col in range(col_count):
            try:
                elem_id = section.GetCellElementId(row, col)
            except Exception:
                elem_id = None
            if elem_id and getattr(elem_id, "IntegerValue", -1) != -1:
                elem_value = str(elem_id.IntegerValue)
                break
        ids.append(elem_value)
    return ids


def inject_element_id_column(data, view):
    if not data:
        return data
    header_rows = get_section_row_count(view, DB.SectionType.Header)
    body_ids = get_body_row_element_ids(view)
    total_rows = len(data)
    if header_rows > total_rows:
        header_rows = total_rows
    body_rows = min(len(body_ids), max(0, total_rows - header_rows))
    for idx, row in enumerate(data):
        if row is None:
            row = []
        if idx < header_rows:
            if idx == max(0, header_rows - 1):
                row.insert(0, "ElementId")
            else:
                row.insert(0, "")
        elif idx < header_rows + body_rows:
            row.insert(0, body_ids[idx - header_rows])
        else:
            row.insert(0, "")
        data[idx] = row
    return data




def write_table_to_sheet(sheet, data, start_row):
    if not data:
        return
    row_idx = start_row
    for row in data:
        col_idx = 1
        for value in row:
            sheet.cell(row=row_idx, column=col_idx, value=value)
            col_idx += 1
        row_idx += 1


def make_unique_name(base, used, max_len=None):
    candidate = base
    if max_len:
        candidate = candidate[:max_len]
    if candidate not in used:
        used.add(candidate)
        return candidate
    idx = 1
    while True:
        suffix = "_{}".format(idx)
        trunc = candidate
        if max_len:
            trunc = candidate[: max_len - len(suffix)]
        name = "{}{}".format(trunc, suffix)
        if name not in used:
            used.add(name)
            return name
        idx += 1


def export_to_excel(doc, schedules, file_path):
    add_lib_path()
    try:
        import openpyxl
    except Exception as exc:
        MessageBox.Show("openpyxl is not available.\n{}".format(exc), "Multiple Schedules Exporter")
        return False

    used_names = set()
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    existing_names = set(workbook.sheetnames)
    existing_list = list(workbook.sheetnames)
    temp_dir = tempfile.mkdtemp(prefix="wwp_schedules_")

    options = DB.ViewScheduleExportOptions()
    options.FieldDelimiter = ","
    try:
        for view in schedules:
            base_name = sanitize_sheet_name(view.Name)
            if base_name in existing_names and base_name not in used_names:
                sheet_name = base_name
            else:
                candidates = [
                    name for name in existing_list
                    if name.startswith(base_name) and name not in used_names
                ]
                if len(candidates) == 1:
                    sheet_name = candidates[0]
                else:
                    used_pool = set(existing_names)
                    used_pool.update(used_names)
                    sheet_name = make_unique_name(base_name, used_pool, max_len=31)
            used_names.add(sheet_name)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.max_row:
                    sheet.delete_rows(1, sheet.max_row)
            else:
                sheet = workbook.create_sheet(title=sheet_name)

            temp_name = "{}.csv".format(sanitize_file_name(view.Name))
            view.Export(temp_dir, temp_name, options)
            csv_path = os.path.join(temp_dir, temp_name)
            data = normalize_table_data(read_csv_rows(csv_path))
            write_table_to_sheet(sheet, data, 1)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

    workbook.save(file_path)
    return True


def export_to_csv(doc, schedules, folder):
    if not os.path.isdir(folder):
        os.makedirs(folder)
    options = DB.ViewScheduleExportOptions()
    options.FieldDelimiter = ","
    used_names = set()
    for view in schedules:
        base_name = sanitize_file_name(view.Name)
        unique_name = make_unique_name(base_name, used_names)
        file_name = "{}.csv".format(unique_name)
        view.Export(folder, file_name, options)
        csv_path = os.path.join(folder, file_name)
        rows = normalize_table_data(read_csv_rows(csv_path))
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)
    return True


class ExportSchedulesForm(Form):
    def __init__(self, items, prechecked_ids):
        Form.__init__(self)
        self.Text = "Export Schedules"
        self.Size = Size(680, 620)
        self.MinimumSize = Size(620, 520)
        self.StartPosition = System.Windows.Forms.FormStartPosition.CenterScreen
        self.selected_items = []
        self._items = list(items)
        self.export_mode = None
        self.result = DialogResult.Cancel

        title = Label()
        title.Text = "Select schedules to export:"
        title.Location = Point(12, 12)
        title.AutoSize = True
        self.Controls.Add(title)

        self.listbox = CheckedListBox()
        self.listbox.CheckOnClick = True
        self.listbox.Location = Point(12, 36)
        self.listbox.Size = Size(640, 420)
        for idx, item in enumerate(self._items):
            self.listbox.Items.Add(item.display_name)
            if item.view.Id.IntegerValue in prechecked_ids:
                self.listbox.SetItemChecked(idx, True)
        self.Controls.Add(self.listbox)

        self.rb_excel = RadioButton()
        self.rb_excel.Text = "Export to Excel"
        self.rb_excel.Location = Point(12, 468)
        self.rb_excel.Checked = True
        self.Controls.Add(self.rb_excel)

        self.rb_csv = RadioButton()
        self.rb_csv.Text = "Export to CSV"
        self.rb_csv.Location = Point(160, 468)
        self.Controls.Add(self.rb_csv)

        export_btn = Button()
        export_btn.Text = "Export"
        export_btn.Location = Point(474, 520)
        export_btn.Click += self.on_export
        self.Controls.Add(export_btn)

        cancel_btn = Button()
        cancel_btn.Text = "Cancel"
        cancel_btn.Location = Point(566, 520)
        cancel_btn.Click += self.on_cancel
        self.Controls.Add(cancel_btn)

    def on_export(self, sender, args):
        checked = [self._items[i] for i in self.listbox.CheckedIndices]
        if not checked:
            MessageBox.Show("Select at least one schedule.", "Multiple Schedules Exporter")
            return
        self.selected_items = checked
        self.export_mode = "excel" if self.rb_excel.Checked else "csv"
        self.result = DialogResult.OK
        self.Close()

    def on_cancel(self, sender, args):
        self.result = DialogResult.Cancel
        self.Close()


def main():
    doc = revit.doc
    config = script.get_config()

    schedules = collect_schedules(doc)
    if not schedules:
        MessageBox.Show("No schedules found.", "Multiple Schedules Exporter")
        return

    items = [ScheduleItem(v) for v in schedules]
    last_ids = getattr(config, CONFIG_LAST_SCHEDULE_IDS, [])
    try:
        prechecked_ids = set(int(x) for x in last_ids)
    except Exception:
        prechecked_ids = set()
    form = ExportSchedulesForm(items, prechecked_ids)
    form.ShowDialog()
    if form.result != DialogResult.OK:
        return
    selected_views = [item.view for item in form.selected_items]
    config.last_schedule_ids = [v.Id.IntegerValue for v in selected_views]

    default_dir = get_default_dir(doc)
    if form.export_mode == "excel":
        last_excel_path = getattr(config, CONFIG_LAST_EXCEL_PATH, "")
        init_dir = os.path.dirname(last_excel_path) if last_excel_path else default_dir
        dialog = SaveFileDialog()
        dialog.Filter = "Excel Workbook (*.xlsx)|*.xlsx"
        dialog.DefaultExt = "xlsx"
        dialog.AddExtension = True
        dialog.CheckPathExists = True
        dialog.ValidateNames = True
        dialog.FileName = os.path.basename(last_excel_path) if last_excel_path else "Schedules.xlsx"
        dialog.InitialDirectory = init_dir
        dialog.OverwritePrompt = False
        if dialog.ShowDialog() != DialogResult.OK:
            return
        file_path = dialog.FileName
        if not file_path.lower().endswith(".xlsx"):
            file_path = "{}.xlsx".format(file_path)
        success = export_to_excel(doc, selected_views, file_path)
        if not success:
            return
        config.last_excel_path = file_path
    else:
        last_csv_dir = getattr(config, CONFIG_LAST_CSV_DIR, "")
        init_dir = last_csv_dir or default_dir
        dialog = FolderBrowserDialog()
        dialog.SelectedPath = init_dir
        if dialog.ShowDialog() != DialogResult.OK:
            return
        folder = dialog.SelectedPath
        export_to_csv(doc, selected_views, folder)
        config.last_csv_dir = folder

    script.save_config()
    MessageBox.Show("Export complete.", "Multiple Schedules Exporter")


if __name__ == "__main__":
    main()
