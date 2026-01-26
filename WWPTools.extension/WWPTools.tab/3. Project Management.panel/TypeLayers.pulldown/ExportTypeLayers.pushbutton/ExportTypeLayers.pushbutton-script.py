#! python3
from __future__ import annotations

import os
import sys

from pyrevit import DB, revit, script


def add_lib_path():
    script_dir = os.path.dirname(__file__)
    lib_path = os.path.abspath(os.path.join(script_dir, "..", "..", "..", "..", "lib"))
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def load_uiutils():
    add_lib_path()
    import WWP_uiUtils as ui
    return ui


def sanitize_sheet_name(name):
    invalid = r"[:\\/?*\[\]]"
    safe = name
    for ch in invalid:
        safe = safe.replace(ch, "_")
    safe = safe.strip()
    if not safe:
        safe = "Types"
    return safe[:31]


def element_id_value(elem_id):
    if elem_id is None:
        return -1
    if hasattr(elem_id, "IntegerValue"):
        return elem_id.IntegerValue
    if hasattr(elem_id, "Value"):
        return elem_id.Value
    try:
        return int(elem_id)
    except Exception:
        return -1


def get_family_name(elem_type):
    try:
        return elem_type.FamilyName
    except Exception:
        return ""


def get_length_unit_id(doc):
    try:
        units = doc.GetUnits()
        return units.GetFormatOptions(DB.SpecTypeId.Length).GetUnitTypeId()
    except Exception:
        return None


def format_length(doc, value):
    try:
        units = doc.GetUnits()
        try:
            return DB.UnitFormatUtils.Format(units, DB.SpecTypeId.Length, value, False)
        except Exception:
            return DB.UnitFormatUtils.Format(units, DB.UnitType.UT_Length, value, False)
    except Exception:
        return str(value)


def convert_length(doc, value):
    unit_id = get_length_unit_id(doc)
    if unit_id:
        try:
            return DB.UnitUtils.ConvertFromInternalUnits(value, unit_id)
        except Exception:
            pass
    try:
        return DB.UnitUtils.ConvertFromInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def convert_length_mm(value):
    try:
        return DB.UnitUtils.ConvertFromInternalUnits(value, DB.DisplayUnitType.DUT_MILLIMETERS)
    except Exception:
        return value


def pick_export_path(doc, ui):
    initial_dir = ""
    try:
        if doc.PathName:
            initial_dir = os.path.dirname(doc.PathName)
    except Exception:
        initial_dir = ""
    path = ui.uiUtils_save_file_dialog(
        title="Export Type Layers",
        filter_text="Excel Workbook (*.xlsx)|*.xlsx",
        default_extension="xlsx",
        initial_directory=initial_dir,
        file_name="Type_Layers.xlsx",
    )
    if not path:
        return None
    if not path.lower().endswith(".xlsx"):
        path = "{}.xlsx".format(path)
    return path


def collect_types(doc, type_class):
    return list(DB.FilteredElementCollector(doc).OfClass(type_class).WhereElementIsElementType())


def layer_material_name(doc, mat_id):
    if mat_id is None:
        return ""
    mat_int = element_id_value(mat_id)
    if mat_int == -1:
        return ""
    try:
        mat = doc.GetElement(mat_id)
        if mat:
            return mat.Name
    except Exception:
        pass
    return ""


def layer_function_name(layer):
    try:
        return str(layer.Function)
    except Exception:
        return ""


def layer_function_id(layer):
    try:
        return int(layer.Function)
    except Exception:
        return ""


def layer_is_variable(layer):
    try:
        return bool(layer.IsVariable)
    except Exception:
        return False


def type_base_row(category_label, elem_type):
    return {
        "Category": category_label,
        "TypeName": elem_type.Name,
        "SourceTypeName": elem_type.Name,
        "FamilyName": get_family_name(elem_type),
        "TypeId": element_id_value(elem_type.Id),
        "TypeUniqueId": elem_type.UniqueId,
    }


def collect_type_rows(doc, category_label, types):
    rows = []
    for elem_type in types:
        base = type_base_row(category_label, elem_type)
        comp = None
        try:
            comp = elem_type.GetCompoundStructure()
        except Exception:
            comp = None

        if comp is None:
            row = dict(base)
            row.update({
                "LayerIndex": "",
                "LayerFunction": "",
                "LayerFunctionId": "",
                "MaterialName": "",
                "MaterialId": "",
                "ThicknessInternal": "",
                "ThicknessMetric": "",
                "IsVariable": "",
                "HasCompoundStructure": False,
            })
            rows.append(row)
            continue

        try:
            layers = list(comp.GetLayers())
        except Exception:
            layers = []

        if not layers:
            row = dict(base)
            row.update({
                "LayerIndex": "",
                "LayerFunction": "",
                "LayerFunctionId": "",
                "MaterialName": "",
                "MaterialId": "",
                "ThicknessInternal": "",
                "ThicknessMetric": "",
                "IsVariable": "",
                "HasCompoundStructure": True,
            })
            rows.append(row)
            continue

        for idx, layer in enumerate(layers):
            width = 0.0
            try:
                width = layer.Width
            except Exception:
                pass
            mat_id = None
            try:
                mat_id = layer.MaterialId
            except Exception:
                pass
            row = dict(base)
            row.update({
                "LayerIndex": idx,
                "LayerFunction": layer_function_name(layer),
                "LayerFunctionId": layer_function_id(layer),
                "MaterialName": layer_material_name(doc, mat_id),
                "MaterialId": element_id_value(mat_id),
                "ThicknessInternal": width,
                "ThicknessMetric": convert_length_mm(width),
                "IsVariable": layer_is_variable(layer),
                "HasCompoundStructure": True,
            })
            rows.append(row)
    return rows


def export_to_excel(doc, selections, file_path, ui):
    add_lib_path()
    try:
        import openpyxl
    except Exception as exc:
        ui.uiUtils_alert("openpyxl is not available.\n{}".format(exc), title="Export Type Layers")
        return False

    workbook = openpyxl.Workbook()
    try:
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    except Exception:
        pass

    headers = [
        "Category",
        "TypeName",
        "SourceTypeName",
        "FamilyName",
        "TypeId",
        "TypeUniqueId",
        "LayerIndex",
        "LayerFunction",
        "LayerFunctionId",
        "MaterialName",
        "MaterialId",
        "ThicknessInternal",
        "ThicknessMetric",
        "IsVariable",
        "HasCompoundStructure",
    ]

    for label, type_class in selections:
        types = collect_types(doc, type_class)
        rows = collect_type_rows(doc, label, types)

        sheet = workbook.create_sheet(title=sanitize_sheet_name(label))
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            if header in ("TypeId", "TypeUniqueId", "SourceTypeName"):
                cell.number_format = "@"
        row_idx = 2
        for row in rows:
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header, "")
                if header in ("TypeId", "TypeUniqueId", "SourceTypeName"):
                    value = "" if value is None else str(value)
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if header in ("TypeId", "TypeUniqueId", "SourceTypeName"):
                    cell.number_format = "@"
            row_idx += 1

    workbook.save(file_path)
    return True


def main():
    doc = revit.doc
    ui = load_uiutils()

    options = [
        ("Wall Types", DB.WallType),
        ("Floor Types", DB.FloorType),
        ("Roof Types", DB.RoofType),
    ]
    labels = [opt[0] for opt in options]
    selected_indices = ui.uiUtils_select_indices(
        labels,
        title="Export Type Layers",
        prompt="Select categories to export:",
        multiselect=True,
    )
    if not selected_indices:
        return

    selections = [options[idx] for idx in selected_indices if 0 <= idx < len(options)]
    if not selections:
        return

    file_path = pick_export_path(doc, ui)
    if not file_path:
        return

    if export_to_excel(doc, selections, file_path, ui):
        ui.uiUtils_alert("Export complete:\n{}".format(file_path), title="Export Type Layers")


if __name__ == "__main__":
    main()
