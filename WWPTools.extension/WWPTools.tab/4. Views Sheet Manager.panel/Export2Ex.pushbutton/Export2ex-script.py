#! python3
from __future__ import annotations

import csv
import importlib
import os
import re
import shutil
import sys
import tempfile

import clr
from System import String
from System.Collections.Generic import List

from pyrevit import DB, revit, script



CONFIG_LAST_EXCEL_PATH = "last_excel_path"
CONFIG_LAST_CSV_DIR = "last_csv_dir"
CONFIG_LAST_SCHEDULE_IDS = "last_schedule_ids"
CONFIG_LAST_CSV_MODE = "last_csv_mode"
CONFIG_LAST_CSV_DELIM = "last_csv_delim"
CONFIG_LAST_EXPORT_MODE = "last_export_mode"
CONFIG_LAST_CSV_EXPORT_TITLE = "last_csv_export_title"
CONFIG_LAST_CSV_COLUMN_HEADERS = "last_csv_column_headers"
CONFIG_LAST_CSV_GROUP_HEADERS = "last_csv_group_headers"
CONFIG_LAST_CSV_GROUPED_COLUMN_HEADERS = "last_csv_grouped_column_headers"
CONFIG_LAST_CSV_TEXT_QUALIFIER = "last_csv_text_qualifier"


def sanitize_sheet_name(name):
    invalid = r"[:\\/?*\[\]]"
    safe = re.sub(invalid, "_", name)
    safe = safe.strip()
    if not safe:
        safe = "Schedule"
    return safe[:31]


def sanitize_file_name(name):
    invalid = r'[<>:"/\\|?*]'
    safe = re.sub(invalid, "_", name).strip()
    return safe or "Schedule"


def get_default_dir(doc):
    if doc.IsWorkshared:
        try:
            central = doc.GetWorksharingCentralModelPath()
            if central:
                return os.path.dirname(DB.ModelPathUtils.ConvertModelPathToUserVisiblePath(central))
        except Exception:
            pass
    if doc.PathName:
        return os.path.dirname(doc.PathName)
    return os.path.expanduser("~")


def ensure_existing_dir(path, fallback=""):
    if path and os.path.isdir(path):
        return path
    if fallback and os.path.isdir(fallback):
        return fallback
    return ""


def collect_schedules(doc):
    schedules = []
    for view in DB.FilteredElementCollector(doc).OfClass(DB.ViewSchedule):
        if view.IsTemplate:
            continue
        try:
            if view.ViewType == DB.ViewType.Legend:
                continue
        except Exception:
            pass
        if view.IsTitleblockRevisionSchedule:
            continue
        schedules.append(view)
    schedules.sort(key=lambda v: v.Name)
    return schedules


def element_id_value(elem_id):
    if elem_id is None:
        return -1
    if hasattr(elem_id, "IntegerValue"):
        return elem_id.IntegerValue
    if hasattr(elem_id, "Value"):
        return elem_id.Value
    try:
        return int(elem_id)
    except Exception:
        return -1


class ScheduleItem(object):
    def __init__(self, view):
        self.view = view
        display_name = "{} [id:{}]".format(view.Name, element_id_value(view.Id))
        self.display_name = display_name.replace("_", "__")


def add_lib_path():
    lib_path = os.path.join(os.path.dirname(__file__), "lib")
    if lib_path not in sys.path:
        sys.path.append(lib_path)


def load_uiutils():
    script_dir = os.path.dirname(__file__)
    lib_path = os.path.abspath(os.path.join(script_dir, "..", "..", "..", "lib"))
    if lib_path not in sys.path:
        sys.path.append(lib_path)
    import WWP_uiUtils as ui
    if not hasattr(ui, "uiUtils_select_items_with_mode"):
        try:
            ui = importlib.reload(ui)
        except Exception:
            pass
    return ui


def _show_export_form(
    ui,
    items,
    prechecked_indices,
    init_excel_path,
    init_csv_dir,
    last_csv_delim,
    last_csv_mode,
    last_export_mode,
    last_export_title,
    last_column_headers,
    last_group_headers,
    last_grouped_column_headers,
    last_text_qualifier,
):
    clr.AddReference("PresentationFramework")
    clr.AddReference("PresentationCore")
    clr.AddReference("WindowsBase")
    from System.IO import StringReader
    from System.Windows.Markup import XamlReader
    from System.Xml import XmlReader
    from System.Windows.Controls import SelectionMode

    def _to_net_list(values):
        net_list = List[String]()
        for value in values:
            net_list.Add("" if value is None else str(value))
        return net_list

    xaml = """
    <Window xmlns="http://schemas.microsoft.com/winfx/2006/xaml/presentation"
            xmlns:x="http://schemas.microsoft.com/winfx/2006/xaml"
            Title="Export Schedules" Height="720" Width="880"
            WindowStartupLocation="CenterScreen" ResizeMode="CanResizeWithGrip">
        <Window.Resources>
            <Style x:Key="PrimaryButton" TargetType="Button">
                <Setter Property="Background" Value="#1976D2"/>
                <Setter Property="Foreground" Value="White"/>
                <Setter Property="FontWeight" Value="SemiBold"/>
                <Setter Property="Padding" Value="14,6"/>
                <Setter Property="Height" Value="40"/>
                <Setter Property="BorderThickness" Value="0"/>
                <Setter Property="Template">
                    <Setter.Value>
                        <ControlTemplate TargetType="Button">
                            <Border Background="{TemplateBinding Background}"
                                    CornerRadius="18">
                                <ContentPresenter HorizontalAlignment="Center"
                                                  VerticalAlignment="Center"/>
                            </Border>
                        </ControlTemplate>
                    </Setter.Value>
                </Setter>
            </Style>
            <Style x:Key="SecondaryButton" TargetType="Button" BasedOn="{StaticResource PrimaryButton}">
                <Setter Property="Background" Value="#E0E0E0"/>
                <Setter Property="Foreground" Value="#424242"/>
            </Style>
            <Style x:Key="BlueToggle" TargetType="ToggleButton">
                <Setter Property="Width" Value="42"/>
                <Setter Property="Height" Value="21"/>
                <Setter Property="FontWeight" Value="Bold"/>
                <Setter Property="FontSize" Value="9"/>
                <Setter Property="Foreground" Value="White"/>
                <Setter Property="Template">
                    <Setter.Value>
                        <ControlTemplate TargetType="ToggleButton">
                            <Grid>
                                <Border x:Name="SwitchTrack"
                                        CornerRadius="10.5"
                                        Background="#BDBDBD"
                                        BorderBrush="#9E9E9E"
                                        BorderThickness="1"/>
                                <Grid>
                                    <TextBlock x:Name="OnText"
                                               Text="ON"
                                               Foreground="White"
                                               FontWeight="Bold"
                                               VerticalAlignment="Center"
                                               HorizontalAlignment="Left"
                                               Margin="4,0,0,0"
                                               Visibility="Collapsed"/>
                                    <TextBlock x:Name="OffText"
                                               Text="OFF"
                                               Foreground="White"
                                               FontWeight="Bold"
                                               VerticalAlignment="Center"
                                               HorizontalAlignment="Right"
                                               Margin="0,0,5,0"
                                               Visibility="Visible"/>
                                </Grid>
                                <Ellipse x:Name="Thumb"
                                         Width="14"
                                         Height="14"
                                         Fill="White"
                                         Stroke="#9E9E9E"
                                         StrokeThickness="1"
                                         HorizontalAlignment="Left"
                                         Margin="3,3,0,3"/>
                            </Grid>
                            <ControlTemplate.Triggers>
                                <Trigger Property="IsChecked" Value="True">
                                    <Setter TargetName="SwitchTrack" Property="Background" Value="#1976D2"/>
                                    <Setter TargetName="SwitchTrack" Property="BorderBrush" Value="#0F5BA8"/>
                                    <Setter TargetName="Thumb" Property="HorizontalAlignment" Value="Right"/>
                                    <Setter TargetName="Thumb" Property="Margin" Value="0,3,3,3"/>
                                    <Setter TargetName="OnText" Property="Visibility" Value="Visible"/>
                                    <Setter TargetName="OffText" Property="Visibility" Value="Collapsed"/>
                                </Trigger>
                                <Trigger Property="IsMouseOver" Value="True">
                                    <Setter TargetName="SwitchTrack" Property="Opacity" Value="0.95"/>
                                </Trigger>
                                <Trigger Property="IsEnabled" Value="False">
                                    <Setter TargetName="SwitchTrack" Property="Opacity" Value="0.5"/>
                                    <Setter TargetName="Thumb" Property="Opacity" Value="0.6"/>
                                </Trigger>
                            </ControlTemplate.Triggers>
                        </ControlTemplate>
                    </Setter.Value>
                </Setter>
            </Style>
        </Window.Resources>
        <Grid Margin="12">
            <Grid.RowDefinitions>
                <RowDefinition Height="Auto"/>
                <RowDefinition Height="*"/>
                <RowDefinition Height="Auto"/>
            </Grid.RowDefinitions>
            <StackPanel Grid.Row="0">
                <TextBlock Text="Search schedules:"/>
                <TextBox Name="SearchBox" Margin="0,4,0,8"/>
            </StackPanel>
            <Grid Grid.Row="1">
                <Grid.ColumnDefinitions>
                    <ColumnDefinition Width="*"/>
                    <ColumnDefinition Width="320"/>
                </Grid.ColumnDefinitions>
                <ListBox Name="ScheduleList" Grid.Column="0" SelectionMode="Extended"/>
                <StackPanel Grid.Column="1" Margin="12,0,0,0">
                    <TextBlock Text="Export mode:"/>
                    <StackPanel Orientation="Horizontal" Margin="0,4,0,8">
                        <RadioButton Name="ExcelMode" Content="Excel" GroupName="ModeGroup" Margin="0,0,12,0"/>
                        <RadioButton Name="CsvMode" Content="CSV" GroupName="ModeGroup"/>
                    </StackPanel>

                    <GroupBox Header="Schedule appearance" Margin="0,4,0,8">
                        <StackPanel Margin="8,6,8,8">
                            <CheckBox Name="ExportTitle" Content="Export title"/>
                            <CheckBox Name="ExportColumnHeaders" Content="Export column headers" IsChecked="True"/>
                            <CheckBox Name="ExportGroupedColumnHeaders" Content="Include grouped column headers" Margin="18,0,0,0"/>
                            <CheckBox Name="ExportGroupHeaders" Content="Export group headers, footers, and blank lines"/>
                        </StackPanel>
                    </GroupBox>

                    <TextBlock Text="Excel file path:"/>
                    <DockPanel Margin="0,4,0,8">
                        <TextBox Name="ExcelPath" DockPanel.Dock="Left" Width="220" Margin="0,0,6,0"/>
                        <Button Name="BrowseExcel" Content="Browse" Width="70"/>
                    </DockPanel>

                    <GroupBox Header="Output options" Margin="0,4,0,8">
                        <StackPanel Margin="8,6,8,8">
                            <TextBlock Text="CSV folder:"/>
                            <DockPanel Margin="0,4,0,8">
                                <TextBox Name="CsvFolder" DockPanel.Dock="Left" Width="220" Margin="0,0,6,0"/>
                                <Button Name="BrowseCsv" Content="Browse" Width="70"/>
                            </DockPanel>

                            <TextBlock Text="CSV delimiter:"/>
                            <ComboBox Name="CsvDelimiter" Margin="0,4,0,8"/>

                            <TextBlock Text="Text qualifier:"/>
                            <ComboBox Name="TextQualifier" Margin="0,4,0,8"/>

                            <StackPanel Orientation="Horizontal" Margin="0,2,0,0" VerticalAlignment="Center">
                                <TextBlock Text="Quote all fields" VerticalAlignment="Center" Margin="0,0,10,0"/>
                                <ToggleButton Name="QuoteAll" Style="{StaticResource BlueToggle}" />
                            </StackPanel>
                        </StackPanel>
                    </GroupBox>
                </StackPanel>
            </Grid>
            <StackPanel Grid.Row="2" Orientation="Horizontal" HorizontalAlignment="Right" Margin="0,12,0,0">
                <Button Name="OkButton" Style="{StaticResource PrimaryButton}" Width="150" Margin="0,0,8,0">
                    <StackPanel Orientation="Horizontal" HorizontalAlignment="Center">
                        <TextBlock Text="&#xE105;" FontFamily="Segoe MDL2 Assets" FontSize="16" Margin="0,0,8,0"/>
                        <TextBlock Text="Export" FontSize="14"/>
                    </StackPanel>
                </Button>
                <Button Name="CancelButton" Style="{StaticResource SecondaryButton}" Width="150">
                    <StackPanel Orientation="Horizontal" HorizontalAlignment="Center">
                        <TextBlock Text="&#xE10A;" FontFamily="Segoe MDL2 Assets" FontSize="16" Margin="0,0,8,0"/>
                        <TextBlock Text="Cancel" FontSize="14"/>
                    </StackPanel>
                </Button>
            </StackPanel>
        </Grid>
    </Window>
    """
    reader = XmlReader.Create(StringReader(xaml))
    window = XamlReader.Load(reader)

    search_box = window.FindName("SearchBox")
    schedule_list = window.FindName("ScheduleList")
    excel_mode = window.FindName("ExcelMode")
    csv_mode = window.FindName("CsvMode")
    excel_path = window.FindName("ExcelPath")
    browse_excel = window.FindName("BrowseExcel")
    csv_folder = window.FindName("CsvFolder")
    browse_csv = window.FindName("BrowseCsv")
    csv_delim = window.FindName("CsvDelimiter")
    text_qualifier = window.FindName("TextQualifier")
    quote_all = window.FindName("QuoteAll")
    export_title = window.FindName("ExportTitle")
    export_column_headers = window.FindName("ExportColumnHeaders")
    export_group_headers = window.FindName("ExportGroupHeaders")
    export_grouped_column_headers = window.FindName("ExportGroupedColumnHeaders")
    ok_button = window.FindName("OkButton")
    cancel_button = window.FindName("CancelButton")

    schedule_list.ItemsSource = _to_net_list(items)
    delimiter_items = [
        "Comma (,)",
        "Semicolon (;)",
        "Tab (\\t)",
    ]
    csv_delim.ItemsSource = _to_net_list(delimiter_items)

    qualifier_items = [
        "(none)",
        'Double Quote (")',
        "Single Quote (')",
    ]
    text_qualifier.ItemsSource = _to_net_list(qualifier_items)

    delimiter_values = {
        "Comma (,)": ",",
        "Semicolon (;)": ";",
        "Tab (\\t)": "\t",
    }
    default_delim_label = "Comma (,)"
    for label, value in delimiter_values.items():
        if value == last_csv_delim:
            default_delim_label = label
            break

    excel_path.Text = init_excel_path or ""
    csv_folder.Text = init_csv_dir or ""
    quote_all.IsChecked = bool(last_csv_mode == 1)
    csv_delim.SelectedItem = default_delim_label

    export_title.IsChecked = bool(last_export_title)
    export_column_headers.IsChecked = bool(last_column_headers)
    export_group_headers.IsChecked = bool(last_group_headers)
    export_grouped_column_headers.IsChecked = bool(last_grouped_column_headers)
    qualifier_label = "(none)"
    if last_text_qualifier == "\"":
        qualifier_label = 'Double Quote (")'
    elif last_text_qualifier == "'":
        qualifier_label = "Single Quote (')"
    text_qualifier.SelectedItem = qualifier_label

    if last_export_mode == 1:
        csv_mode.IsChecked = True
    else:
        excel_mode.IsChecked = True

    selected_names = set()
    if prechecked_indices:
        for idx in prechecked_indices:
            if 0 <= idx < len(items):
                selected_names.add(items[idx])

    def _apply_selection():
        schedule_list.SelectedItems.Clear()
        for item in schedule_list.Items:
            if str(item) in selected_names:
                schedule_list.SelectedItems.Add(item)

    _apply_selection()

    def _update_enabled_state():
        is_excel = bool(excel_mode.IsChecked)
        excel_path.IsEnabled = is_excel
        browse_excel.IsEnabled = is_excel
        csv_folder.IsEnabled = not is_excel
        browse_csv.IsEnabled = not is_excel

    def _filter_list(_sender=None, _args=None):
        text = search_box.Text or ""
        text = text.strip().lower()
        if not text:
            filtered = items
        else:
            filtered = [item for item in items if text in item.lower()]
        schedule_list.ItemsSource = _to_net_list(filtered)
        _apply_selection()

    def _selection_changed(_sender, _args):
        # Keep export selection aligned with what is currently highlighted in the UI.
        selected_names.clear()
        for item in schedule_list.SelectedItems:
            selected_names.add(str(item))

    def _browse_excel(_sender, _args):
        current = excel_path.Text or ""
        init_dir = os.path.dirname(current) if current else ""
        file_path = ui.uiUtils_save_file_dialog(
            title="Export Schedules",
            filter_text="Excel Workbook (*.xlsx)|*.xlsx",
            default_extension="xlsx",
            initial_directory=init_dir,
            file_name=os.path.basename(current) if current else "Schedules.xlsx",
        )
        if file_path:
            excel_path.Text = file_path

    def _browse_csv(_sender, _args):
        init_dir = csv_folder.Text or ""
        folder = ui.uiUtils_select_folder_dialog(
            title="Select CSV Folder",
            initial_directory=init_dir,
        )
        if folder:
            csv_folder.Text = folder

    def _ok(_sender, _args):
        window.DialogResult = True
        window.Close()

    def _cancel(_sender, _args):
        window.DialogResult = False
        window.Close()

    excel_mode.Checked += lambda s, e: _update_enabled_state()
    csv_mode.Checked += lambda s, e: _update_enabled_state()
    search_box.TextChanged += _filter_list
    schedule_list.SelectionChanged += _selection_changed
    browse_excel.Click += _browse_excel
    browse_csv.Click += _browse_csv
    ok_button.Click += _ok
    cancel_button.Click += _cancel

    _update_enabled_state()

    if not window.ShowDialog():
        return None

    selected_indices = [idx for idx, item in enumerate(items) if item in selected_names]
    selected_mode = 0 if excel_mode.IsChecked else 1
    delimiter_label = str(csv_delim.SelectedItem) if csv_delim.SelectedItem else default_delim_label
    delimiter = delimiter_values.get(delimiter_label, ",")
    qualifier_label = str(text_qualifier.SelectedItem) if text_qualifier.SelectedItem else "(none)"
    qualifier_value = ""
    if qualifier_label.startswith("Double"):
        qualifier_value = "\""
    elif qualifier_label.startswith("Single"):
        qualifier_value = "'"

    return {
        "selected_indices": selected_indices,
        "mode": selected_mode,
        "excel_path": excel_path.Text or "",
        "csv_folder": csv_folder.Text or "",
        "csv_delimiter": delimiter,
        "csv_quote_all": bool(quote_all.IsChecked),
        "csv_text_qualifier": qualifier_value,
        "export_title": bool(export_title.IsChecked),
        "export_column_headers": bool(export_column_headers.IsChecked),
        "export_group_headers": bool(export_group_headers.IsChecked),
        "export_grouped_column_headers": bool(export_grouped_column_headers.IsChecked),
    }


def read_csv_rows(path, delimiter=",", quotechar=""):
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                if quotechar in ("\"", "'"):
                    reader = csv.reader(handle, delimiter=delimiter, quotechar=quotechar)
                else:
                    reader = csv.reader(handle, delimiter=delimiter)
                return [row for row in reader]
        except Exception:
            continue
    return []


def normalize_table_data(data):
    return data


def get_section_data(view, section_type):
    table = view.GetTableData()
    try:
        return table.GetSectionData(section_type)
    except Exception:
        return None


def get_section_row_count(view, section_type):
    section = get_section_data(view, section_type)
    if section is None:
        return 0
    try:
        return section.NumberOfRows
    except Exception:
        return 0


def is_key_schedule(view):
    try:
        definition = view.Definition
    except Exception:
        definition = None
    try:
        return bool(definition and definition.IsKeySchedule)
    except Exception:
        return False


def get_body_row_element_ids(view):
    section = get_section_data(view, DB.SectionType.Body)
    if section is None:
        return []
    ids = []
    for row in range(section.NumberOfRows):
        elem_value = ""
        try:
            col_count = section.NumberOfColumns
        except Exception:
            col_count = 0
        for col in range(col_count):
            try:
                elem_id = section.GetCellElementId(row, col)
            except Exception:
                elem_id = None
            elem_int = element_id_value(elem_id)
            if elem_int != -1:
                elem_value = str(elem_int)
                break
        ids.append(elem_value)
    return ids


def inject_element_id_column(data, view, csv_text=False):
    if not data:
        return data
    header_rows = get_section_row_count(view, DB.SectionType.Header)
    body_ids = get_body_row_element_ids(view)
    total_rows = len(data)
    if header_rows > total_rows:
        header_rows = total_rows
    body_rows = min(len(body_ids), max(0, total_rows - header_rows))
    for idx, row in enumerate(data):
        if row is None:
            row = []
        if idx < header_rows:
            if idx == max(0, header_rows - 1):
                row.append("ElementId")
            else:
                row.append("")
        elif idx < header_rows + body_rows:
            elem_value = body_ids[idx - header_rows]
            if csv_text and elem_value and elem_value.isdigit() and len(elem_value) > 11:
                elem_value = "'" + elem_value
            row.append(elem_value)
        else:
            row.append("")
        data[idx] = row
    return data




def write_table_to_sheet(sheet, data, start_row, header_rows=0, column_specs=None, doc=None):
    if not data:
        return
    has_specs = bool(column_specs) and any(spec is not None for spec in column_specs)
    row_idx = start_row
    for row_offset, row in enumerate(data):
        col_idx = 1
        for value in row:
            spec = None
            if has_specs and column_specs and (col_idx - 1) < len(column_specs):
                spec = column_specs[col_idx - 1]
            if row_offset < header_rows:
                cell_value = value
            elif has_specs:
                if spec is None:
                    cell_value = value
                else:
                    cell_value = coerce_cell_value(value, spec=spec, doc=doc, numeric_fallback=True)
            else:
                cell_value = coerce_cell_value(value, spec=None, doc=None, numeric_fallback=False)
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            if row_offset >= header_rows and col_idx == len(row):
                if cell_value is None:
                    cell_value = ""
                cell.value = str(cell_value)
                cell.number_format = "@"
            col_idx += 1
        row_idx += 1


_NUMERIC_RE = re.compile(r"^-?\d+(\.\d+)?$")


def coerce_cell_value(value, spec=None, doc=None, numeric_fallback=True):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return ""
        if not numeric_fallback:
            return value
        if "'" in text or '"' in text or "/" in text:
            return value
        if re.search(r"\d[A-Za-z]", text):
            return value
        if _NUMERIC_RE.match(text):
            if spec is None:
                if text.startswith("0") and len(text) > 1 and not text.startswith("0."):
                    return value
                if text.startswith("-0") and len(text) > 2 and not text.startswith("-0."):
                    return value
            try:
                return int(text)
            except Exception:
                pass
            try:
                return float(text)
            except Exception:
                return value
    return value


def get_column_specs(view):
    section = get_section_data(view, DB.SectionType.Body)
    if section is None:
        return []
    try:
        col_count = int(section.NumberOfColumns)
    except Exception:
        col_count = 0
    specs = [None] * col_count
    try:
        definition = view.Definition
        field_ids = list(definition.GetFieldOrder())
        col = 0
        for field_id in field_ids:
            try:
                field = definition.GetField(field_id)
            except Exception:
                continue
            try:
                if field.IsHidden:
                    continue
            except Exception:
                pass
            if col >= col_count:
                break
            spec = None
            try:
                spec = field.GetSpecTypeId()
            except Exception:
                pass
            if spec is None:
                try:
                    spec = field.UnitType
                except Exception:
                    pass
            specs[col] = spec
            col += 1
    except Exception:
        pass
    return specs


def make_unique_name(base, used, max_len=None):
    candidate = base
    if max_len:
        candidate = candidate[:max_len]
    if candidate not in used:
        used.add(candidate)
        return candidate
    idx = 1
    while True:
        suffix = "_{}".format(idx)
        trunc = candidate
        if max_len:
            trunc = candidate[: max_len - len(suffix)]
        name = "{}{}".format(trunc, suffix)
        if name not in used:
            used.add(name)
            return name
        idx += 1


def _try_set_option(options, names, value):
    for name in names:
        if not hasattr(options, name):
            continue
        try:
            setattr(options, name, value)
            return True
        except Exception:
            continue
    return False


def apply_schedule_export_options(
    options,
    delimiter=",",
    export_title=False,
    export_column_headers=True,
    export_group_headers=False,
    export_grouped_column_headers=False,
    text_qualifier="",
):
    options.FieldDelimiter = delimiter
    # Revit API property names differ by version; support both.
    _try_set_option(options, ("ExportTitle", "Title"), bool(export_title))
    _try_set_option(options, ("ExportColumnHeaders", "ColumnHeaders"), bool(export_column_headers))
    _try_set_option(options, ("ExportGroupHeaders", "HeadersFootersBlanks"), bool(export_group_headers))
    _try_set_option(options, ("ExportGroupedColumnHeaders",), bool(export_grouped_column_headers))

    if text_qualifier in ("\"", "'"):
        _try_set_option(options, ("TextQualifier",), text_qualifier)
    else:
        _try_set_option(options, ("TextQualifier",), "")


def export_to_excel(
    doc,
    schedules,
    file_path,
    ui,
    export_title=False,
    export_column_headers=True,
    export_group_headers=False,
    export_grouped_column_headers=False,
    text_qualifier="",
    delimiter=",",
):
    add_lib_path()
    try:
        import openpyxl
    except Exception as exc:
        ui.uiUtils_alert(
            "openpyxl is not available.\n{}".format(exc),
            title="Multiple Schedules Exporter",
        )
        return False

    used_names = set()
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
    temp_dir = tempfile.mkdtemp(prefix="wwp_schedules_")

    options = DB.ViewScheduleExportOptions()
    apply_schedule_export_options(
        options,
        delimiter=delimiter,
        export_title=export_title,
        export_column_headers=export_column_headers,
        export_group_headers=export_group_headers,
        export_grouped_column_headers=export_grouped_column_headers,
        text_qualifier=text_qualifier,
    )
    try:
        for view in schedules:
            key_schedule = is_key_schedule(view)
            base_name = sanitize_sheet_name(view.Name)
            if base_name in workbook.sheetnames and base_name not in used_names:
                sheet_name = base_name
            else:
                used_pool = set(workbook.sheetnames)
                used_pool.update(used_names)
                sheet_name = make_unique_name(base_name, used_pool, max_len=31)
            used_names.add(sheet_name)
            if sheet_name in workbook.sheetnames:
                existing = workbook[sheet_name]
                sheet_idx = workbook.worksheets.index(existing)
                workbook.remove(existing)
                sheet = workbook.create_sheet(title=sheet_name, index=sheet_idx)
            else:
                sheet = workbook.create_sheet(title=sheet_name)

            temp_name = "{}.csv".format(sanitize_file_name(view.Name))
            view.Export(temp_dir, temp_name, options)
            csv_path = os.path.join(temp_dir, temp_name)
            data = normalize_table_data(read_csv_rows(csv_path, delimiter=delimiter, quotechar=text_qualifier))
            if not key_schedule:
                data = inject_element_id_column(data, view, csv_text=False)
            header_rows = get_section_row_count(view, DB.SectionType.Header)
            column_specs = get_column_specs(view)
            if key_schedule:
                # Preserve key schedule values exactly as exported (avoid numeric coercion).
                column_specs = None
            elif column_specs is not None:
                column_specs = [None] + column_specs
            write_table_to_sheet(
                sheet,
                data,
                1,
                header_rows=header_rows,
                column_specs=column_specs,
                doc=doc,
            )
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

    workbook.save(file_path)
    return True


def export_to_csv(
    doc,
    schedules,
    folder,
    quote_all=False,
    delimiter=",",
    export_title=False,
    export_column_headers=True,
    export_group_headers=False,
    export_grouped_column_headers=False,
    text_qualifier="",
):
    if not os.path.isdir(folder):
        os.makedirs(folder)
    options = DB.ViewScheduleExportOptions()
    apply_schedule_export_options(
        options,
        delimiter=delimiter,
        export_title=export_title,
        export_column_headers=export_column_headers,
        export_group_headers=export_group_headers,
        export_grouped_column_headers=export_grouped_column_headers,
        text_qualifier=text_qualifier,
    )
    used_names = set()
    for view in schedules:
        key_schedule = is_key_schedule(view)
        base_name = sanitize_file_name(view.Name)
        unique_name = make_unique_name(base_name, used_names)
        file_name = "{}.csv".format(unique_name)
        view.Export(folder, file_name, options)
        csv_path = os.path.join(folder, file_name)
        rows = normalize_table_data(read_csv_rows(csv_path, delimiter=delimiter, quotechar=text_qualifier))
        if not key_schedule:
            rows = inject_element_id_column(rows, view, csv_text=True)
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as handle:
            if text_qualifier in ("\"", "'"):
                writer = csv.writer(
                    handle,
                    delimiter=delimiter,
                    quoting=csv.QUOTE_ALL if quote_all else csv.QUOTE_MINIMAL,
                    quotechar=text_qualifier,
                )
            else:
                writer = csv.writer(
                    handle,
                    delimiter=delimiter,
                    quoting=csv.QUOTE_ALL if quote_all else csv.QUOTE_MINIMAL,
                )
            writer.writerows(rows)
    return True


def select_csv_mode(ui, default_mode=0):
    options = [
        "Standard CSV (comma, minimal quotes)",
        "CSV (Quoted - all fields)",
    ]
    try:
        selected = ui.uiUtils_select_indices(
            options,
            title="CSV Export Mode",
            prompt="Choose CSV export format:",
            multiselect=False,
            width=520,
            height=260,
        )
    except Exception:
        selected = []
    if selected is None or len(selected) == 0:
        return None
    if selected[0] < 0 or selected[0] >= len(options):
        return default_mode
    return int(selected[0])


def select_csv_delimiter(ui, default_delimiter=","):
    options = [
        ("Comma (,)", ","),
        ("Semicolon (;)", ";"),
        ("Tab (\\t)", "\t"),
    ]
    labels = [opt[0] for opt in options]
    default_index = 0
    for idx, opt in enumerate(options):
        if opt[1] == default_delimiter:
            default_index = idx
            break
    try:
        selected = ui.uiUtils_select_indices(
            labels,
            title="CSV Delimiter",
            prompt="Choose delimiter:",
            multiselect=False,
            width=520,
            height=260,
        )
    except Exception:
        selected = []
    if selected is None or len(selected) == 0:
        return None
    sel_idx = int(selected[0]) if selected else default_index
    if sel_idx < 0 or sel_idx >= len(options):
        sel_idx = default_index
    return options[sel_idx][1]


def main():
    doc = revit.doc
    config = script.get_config()
    ui = load_uiutils()

    schedules = collect_schedules(doc)
    if not schedules:
        ui.uiUtils_alert("No schedules found.", title="Multiple Schedules Exporter")
        return

    items = [ScheduleItem(v) for v in schedules]
    last_ids = getattr(config, CONFIG_LAST_SCHEDULE_IDS, [])
    try:
        prechecked_ids = set(int(x) for x in last_ids)
    except Exception:
        prechecked_ids = set()
    prechecked_indices = [
        idx for idx, item in enumerate(items)
        if element_id_value(item.view.Id) in prechecked_ids
    ]
    default_dir = get_default_dir(doc)
    last_excel_path = getattr(config, CONFIG_LAST_EXCEL_PATH, "")
    last_csv_dir = getattr(config, CONFIG_LAST_CSV_DIR, "")
    last_csv_mode = getattr(config, CONFIG_LAST_CSV_MODE, 0)
    last_csv_delim = getattr(config, CONFIG_LAST_CSV_DELIM, ",")
    last_export_mode = getattr(config, CONFIG_LAST_EXPORT_MODE, 0)
    last_export_title = getattr(config, CONFIG_LAST_CSV_EXPORT_TITLE, False)
    last_column_headers = getattr(config, CONFIG_LAST_CSV_COLUMN_HEADERS, True)
    last_group_headers = getattr(config, CONFIG_LAST_CSV_GROUP_HEADERS, False)
    last_grouped_column_headers = getattr(config, CONFIG_LAST_CSV_GROUPED_COLUMN_HEADERS, False)
    last_text_qualifier = getattr(config, CONFIG_LAST_CSV_TEXT_QUALIFIER, "")

    init_excel_path = last_excel_path or os.path.join(default_dir, "Schedules.xlsx")
    init_csv_dir = ensure_existing_dir(last_csv_dir, default_dir)
    inputs = _show_export_form(
        ui,
        [item.display_name for item in items],
        prechecked_indices,
        init_excel_path,
        init_csv_dir,
        last_csv_delim,
        last_csv_mode,
        last_export_mode,
        last_export_title,
        last_column_headers,
        last_group_headers,
        last_grouped_column_headers,
        last_text_qualifier,
    )
    if inputs is not False:
        if not inputs:
            return
        selected_indices = inputs.get("selected_indices") or []
        if not selected_indices:
            ui.uiUtils_alert("Select at least one schedule.", title="Multiple Schedules Exporter")
            return
        selected_views = [items[i].view for i in selected_indices]
        config.last_schedule_ids = [element_id_value(v.Id) for v in selected_views]
        export_title = bool(inputs.get("export_title"))
        export_column_headers = bool(inputs.get("export_column_headers"))
        export_group_headers = bool(inputs.get("export_group_headers"))
        export_grouped_column_headers = bool(inputs.get("export_grouped_column_headers"))
        csv_text_qualifier = inputs.get("csv_text_qualifier") or ""
        csv_delim = inputs.get("csv_delimiter") or last_csv_delim
        quote_all = bool(inputs.get("csv_quote_all"))
        mode = int(inputs.get("mode", 0))
        if mode == 0:
            file_path = (inputs.get("excel_path") or "").strip()
            if not file_path:
                ui.uiUtils_alert("Choose an Excel file path.", title="Multiple Schedules Exporter")
                return
            if not file_path.lower().endswith(".xlsx"):
                file_path = "{}.xlsx".format(file_path)
            success = export_to_excel(
                doc,
                selected_views,
                file_path,
                ui,
                export_title=export_title,
                export_column_headers=export_column_headers,
                export_group_headers=export_group_headers,
                export_grouped_column_headers=export_grouped_column_headers,
                text_qualifier=csv_text_qualifier,
                delimiter=csv_delim,
            )
            if not success:
                return
            config.last_excel_path = file_path
            try:
                os.startfile(file_path)
            except Exception:
                pass
        else:
            folder = (inputs.get("csv_folder") or "").strip()
            if not folder:
                ui.uiUtils_alert("Choose a CSV folder.", title="Multiple Schedules Exporter")
                return
            export_to_csv(
                doc,
                selected_views,
                folder,
                quote_all=quote_all,
                delimiter=csv_delim,
                export_title=export_title,
                export_column_headers=export_column_headers,
                export_group_headers=export_group_headers,
                export_grouped_column_headers=export_grouped_column_headers,
                text_qualifier=csv_text_qualifier,
            )
            config.last_csv_dir = folder
        config.last_export_mode = mode
        config.last_csv_mode = 1 if quote_all else 0
        config.last_csv_delim = csv_delim
        config.last_csv_text_qualifier = csv_text_qualifier
        config.last_csv_export_title = export_title
        config.last_csv_column_headers = export_column_headers
        config.last_csv_group_headers = export_group_headers
        config.last_csv_grouped_column_headers = export_grouped_column_headers
        script.save_config()
        ui.uiUtils_alert("Export complete.", title="Multiple Schedules Exporter")
        return

    if hasattr(ui, "uiUtils_select_items_with_mode"):
        selected_indices, mode = ui.uiUtils_select_items_with_mode(
            [item.display_name for item in items],
            title="Export Schedules",
            prompt="Select schedules to export:",
            mode_labels=("Export to Excel", "Export to CSV"),
            default_mode=0,
            prechecked_indices=prechecked_indices,
            width=680,
            height=620,
        )
    else:
        ui.uiUtils_alert(
            "UI helper uiUtils_select_items_with_mode is unavailable. Restart pyRevit or update WWP_uiUtils.",
            title="Multiple Schedules Exporter",
        )
        return
    if mode is None:
        return
    if not selected_indices:
        ui.uiUtils_alert("Select at least one schedule.", title="Multiple Schedules Exporter")
        return
    selected_views = [items[i].view for i in selected_indices]
    config.last_schedule_ids = [element_id_value(v.Id) for v in selected_views]

    if mode == 0:
        last_excel_dir = os.path.dirname(last_excel_path) if last_excel_path else ""
        init_dir = ensure_existing_dir(last_excel_dir, default_dir)
        file_path = ui.uiUtils_save_file_dialog(
            title="Export Schedules",
            filter_text="Excel Workbook (*.xlsx)|*.xlsx",
            default_extension="xlsx",
            initial_directory=init_dir,
            file_name=os.path.basename(last_excel_path) if last_excel_path else "Schedules.xlsx",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path = "{}.xlsx".format(file_path)
        success = export_to_excel(doc, selected_views, file_path, ui)
        if not success:
            return
        config.last_excel_path = file_path
        try:
            os.startfile(file_path)
        except Exception:
            pass
    else:
        init_dir = ensure_existing_dir(last_csv_dir, default_dir)
        csv_mode = select_csv_mode(ui, default_mode=last_csv_mode)
        if csv_mode is None:
            return
        csv_delim = select_csv_delimiter(ui, default_delimiter=last_csv_delim)
        if csv_delim is None:
            return
        folder = ui.uiUtils_select_folder_dialog(
            title="Select CSV Folder",
            initial_directory=init_dir,
        )
        if not folder:
            return
        export_to_csv(doc, selected_views, folder, quote_all=(csv_mode == 1), delimiter=csv_delim)
        config.last_csv_dir = folder
        config.last_csv_mode = csv_mode
        config.last_csv_delim = csv_delim

    script.save_config()
    ui.uiUtils_alert("Export complete.", title="Multiple Schedules Exporter")


if __name__ == "__main__":
    main()
